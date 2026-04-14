

# -*- coding: utf-8 -*-
"""SPF production simulation with Rule base / GA modes.

- Rule base 모드: 기존 계획 수립 로직 유지
- 최적화 모드: Job 순서와 2D 용접장 배정을 Genetic Algorithm으로 최적화
"""

# =====================================================
# 초보자용 코드 읽기 안내
# -----------------------------------------------------
# 이 파일은 크게 5단계로 나뉜다.
# 1) 입력 파일(CycleTime / Product / Plan) 읽기
# 2) 제품 정보를 내부 계획 데이터로 변환
# 3) SimPy 자원/이벤트를 사용해 생산 흐름 시뮬레이션
# 4) 결과를 간트 차트와 엑셀(ProcessTime, Plan)로 저장
# 5) Tkinter 화면(UI)에서 버튼으로 위 기능 실행
#
# 따라서 코드를 볼 때는
# - 먼저 '입력/전처리 함수'
# - 다음으로 '공정 시간 조회 함수'
# - 그다음 '제품 타입별 run_... 함수'
# - 마지막으로 'run_simulation / UI 클래스'
# 순서로 따라가면 전체 구조를 이해하기 쉽다.
# =====================================================


from __future__ import annotations

import math
import os
import platform
import random
import sys
import tempfile
import traceback
import xml.etree.ElementTree as ET
import zipfile
import base64
import zlib
import types
from collections import defaultdict, deque
from copy import copy
from dataclasses import dataclass
from datetime import datetime
import time
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import pandas as pd
import simpy
import tkinter as tk
from contextlib import redirect_stderr, redirect_stdout
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
from matplotlib.offsetbox import AnchoredText

try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None
from matplotlib.ticker import FuncFormatter
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    import win32com.client as win32
except Exception:
    win32 = None

try:
    import pythoncom
except Exception:
    pythoncom = None


# =====================================================
# 데이터 묶음 클래스
# -----------------------------------------------------
# 여러 값을 함수마다 따로 넘기면 헷갈리기 쉬우므로,
# 서로 관련 있는 값들을 클래스 하나로 묶어 관리한다.
# =====================================================

# 초보자 설명: 프로그램에서 사용하는 주요 파일 경로를 한 묶음으로 저장하는 데이터 클래스다.
@dataclass(frozen=True)
class AppPaths:
    base_dir: str
    cycle_time: str
    product: str
    plan: str
    result_png: str
    process_time: str

    ga_result_png: str
    ga_plan: str
    ga_process_time: str


# 초보자 설명: CycleTime.xlsx를 읽은 결과를 보관하는 데이터 클래스다. 원본 행과 검색용 인덱스를 함께 가진다.
@dataclass
class CycleTimeData:
    rows: List[Tuple[Any, ...]]
    index: Dict[Tuple[str, str, str, str], List[int]]
    sheet_name: Any


# 초보자 설명: Product.xlsx에서 파생된 계획용 중간 데이터들을 한 번에 묶어 전달하는 데이터 클래스다.
@dataclass
class ProductPlanData:
    product_rows: List[Tuple[Any, ...]]
    tmp_product_headers: List[str]
    tmp_product_rows: List[Tuple[Any, ...]]
    spool_fit_detail: Dict[str, Tuple[str, str]]
    srt_plan_rows: List[Tuple[Any, ...]]
    plan1_rows: List[Tuple[Any, ...]]
    plan2_rows: List[Tuple[Any, ...]]
    srt_plan_sim_rows: List[Tuple[Any, ...]]
    plan1_sim_rows: List[Tuple[Any, ...]]
    plan2_sim_rows: List[Tuple[Any, ...]]


SHORT_PRODUCT_TYPES = {"2DS", "2DL", "Short", "F2DS", "F2DL", "FF", "ff"}
CUT_BASED_SHORT_PRODUCT_TYPES = {"2DS", "2DL", "Short", "F2DS", "F2DL"}
LONG_3D_PRODUCT_TYPES = {"F2D", "H3D", "F3D"}
MULTI_SUB_PRODUCT_TYPES = {"F2D", "F3D", "F2DS", "F2DL", "FF"}
AUX_INSERT_PRODUCT_ORDER = ("F2D", "F3D", "H3D")
ELBOW_TEE_FITTING_TYPES = {"45el", "90el", "Tee", "ReTee"}
VALID_FITTING_TYPES = ELBOW_TEE_FITTING_TYPES | {"fl", "sShBevel"}
WELD_LABEL_NUMBERS = {"2D-1": "1", "2D-2": "2", "3D": "3"}


# =====================================================
# 실행 경로 설정 안내
# -----------------------------------------------------
# 현재 개발 환경은 Windows + PyCharm 기준으로 정리한다.
# - 우선순위 1: 환경변수 SPF_BASE_DIR
# - 우선순위 2: exe 실행이면 exe 폴더
# - 우선순위 3: 일반 Python 실행이면 현재 .py 파일 폴더
# - 우선순위 4: 현재 작업 폴더(cwd)
# =====================================================

# =====================================================
# 파일 경로 / 엑셀 입출력
# -----------------------------------------------------
# 이 구간은 입력 파일을 찾고, 읽고, 저장하는 역할을 한다.
# 시뮬레이션 계산 자체보다는 '데이터를 준비하는 단계'라고 보면 된다.
# =====================================================

# 초보자 설명: 입력 파일이 실제로 존재하는 기준 폴더를 우선순위에 따라 찾아낸다.
def pick_existing_base_dir() -> str:
    """입력 파일이 존재하는 최적의 기준 경로를 선택한다."""
    candidate_dirs: List[str] = []

    env_base = os.environ.get("SPF_BASE_DIR")
    if env_base:
        candidate_dirs.append(env_base)

    try:
        if getattr(sys, "frozen", False):
            candidate_dirs.append(os.path.dirname(sys.executable))
        else:
            candidate_dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass

    cwd = os.getcwd()
    if cwd:
        candidate_dirs.append(cwd)

    unique_dirs = []
    seen = set()
    for base in candidate_dirs:
        if not base:
            continue
        norm = os.path.abspath(base)
        if norm not in seen:
            seen.add(norm)
            unique_dirs.append(norm)

    required_names = ["CycleTime.xlsx", "Product.xlsx"]
    for base in unique_dirs:
        try:
            if all(os.path.exists(os.path.join(base, name)) for name in required_names):
                return base
        except Exception:
            continue

    for base in unique_dirs:
        if os.path.isdir(base):
            return base

    try:
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))
    except Exception:
        return os.getcwd()


# 초보자 설명: 선택한 파일이 진짜 .xlsx 형식인지 먼저 검사한다.
def validate_xlsx_file(path: str) -> None:
    """입력 파일이 일반 .xlsx(zip) 또는 Excel COM으로 열 수 있는 보호 파일인지 검사한다."""
    preferred_dir = os.path.dirname(path) or None
    debug_log(f"validate_xlsx_file 시작 | path={path}", preferred_dir=preferred_dir)

    if not os.path.exists(path):
        msg = f"엑셀 파일이 없습니다: {path}"
        debug_log(f"validate_xlsx_file 실패 | {msg}", preferred_dir=preferred_dir)
        raise FileNotFoundError(msg)

    mode = detect_excel_access_mode(path, preferred_dir=preferred_dir)
    debug_log(f"validate_xlsx_file access_mode={mode} | path={path}", preferred_dir=preferred_dir)

    if mode == "zip":
        return

    if mode == "com":
        try:
            rows, book, used_sheet_name = excel_read_via_com(path, sheet_name=0, skip_header=False)
            try:
                debug_log(
                    f"validate_xlsx_file 보호/보안 파일로 판단 | Excel COM 열기 성공 | rows={len(rows)} | used_sheet_name={used_sheet_name}",
                    preferred_dir=preferred_dir,
                )
                _set_cached_excel_access_mode(path, "com")
            finally:
                try:
                    book.close()
                except Exception:
                    pass
            return
        except Exception as exc:
            debug_log_exception(f"validate_xlsx_file COM 확인 실패: {path}", exc, preferred_dir=preferred_dir)
            inspect_excel_file(path, label="validate_xlsx_file COM 실패 후 재조사", preferred_dir=preferred_dir)
            probe_excel_read_methods(path, preferred_dir=preferred_dir)
            signature_hint = describe_file_signature(read_file_header_bytes(path, 64))
            raise ValueError(
                f"'{os.path.basename(path)}' 파일은 일반 ZIP 기반 .xlsx가 아니며, 현재 PC에서 Excel COM으로도 안정적으로 열리지 않습니다.\n"
                f"- 현재 파일 서명 추정: {signature_hint}\n"
                f"- 보안/DRM 프로그램이 Excel 자동화를 차단했거나, Excel COM 연결이 중간에 끊어진 상태일 수 있습니다.\n"
                f"- 열려 있는 Excel 창을 모두 닫고 다시 실행하거나, 파일을 다시 저장한 뒤 시도해 주세요.\n\n"
                f"디버그 로그 파일:\n{get_debug_log_path(preferred_dir=preferred_dir)}"
            )

    inspect_excel_file(path, label="validate_xlsx_file unsupported", preferred_dir=preferred_dir)
    signature_hint = describe_file_signature(read_file_header_bytes(path, 64))
    raise ValueError(
        f"'{os.path.basename(path)}' 파일을 현재 환경에서는 읽을 수 없습니다.\n"
        f"- 현재 파일 서명 추정: {signature_hint}\n"
        f"- 일반 ZIP 기반 .xlsx가 아니고, Windows Excel COM도 사용할 수 없는 환경입니다.\n\n"
        f"디버그 로그 파일:\n{get_debug_log_path(preferred_dir=preferred_dir)}"
    )


# 초보자 설명: 기준 폴더를 바탕으로 입력/출력 파일 경로를 한꺼번에 만든다.
def build_paths(base_dir: Optional[str] = None) -> AppPaths:
    base = base_dir or pick_existing_base_dir()
    cycle_time_path = os.path.join(base, "CycleTime.xlsx")
    product_path_pref = os.path.join(base, "Product.xlsx")
    return AppPaths(
        base_dir=base,
        cycle_time=cycle_time_path,
        product=product_path_pref,
        plan=os.path.join(base, "Plan_rule.xlsx"),
        result_png=os.path.join(base, "result_rule.png"),
        process_time=os.path.join(base, "ProcessTime_rule.xlsx"),
        ga_result_png=os.path.join(base, "result_GA.png"),
        ga_plan=os.path.join(base, "Plan_GA.xlsx"),
        ga_process_time=os.path.join(base, "ProcessTime_GA.xlsx"),
    )


# 초보자 설명: 만들어진 경로 정보를 기존 전역 변수 형식에 맞춰 반영한다.
def configure_paths(paths: AppPaths) -> None:
    """기존 전역 경로 사용 코드를 유지하기 위해 모듈 전역 변수에 반영한다."""
    global BASE_DIR, ct_path, product_path, tmp_product_path, tmp_plan_path, plan_xlsx_path
    BASE_DIR = paths.base_dir
    ct_path = paths.cycle_time
    product_path = paths.product
    tmp_product_path = None
    tmp_plan_path = None
    plan_xlsx_path = paths.plan
    set_debug_log_path(preferred_dir=paths.base_dir)
    print(f"[INFO] BASE_DIR = {BASE_DIR}")
    print(f"[INFO] DEBUG_LOG = {get_debug_log_path(preferred_dir=paths.base_dir)}")


# =====================================================
# 파일 읽기 디버그 로그
# -----------------------------------------------------
# 다른 PC에서만 .xlsx 판정이 실패하는 문제를 재현하기 위해
# 파일 헤더, zip 판정, COM/openpyxl 예외, traceback을
# .txt 파일로 남긴다.
# =====================================================

DEBUG_LOG_PATH = None


# 초보자 설명: 디버그 로그를 저장할 기본 폴더를 고른다.
def _pick_debug_log_base_dir(preferred_dir: Optional[str] = None) -> str:
    candidates: List[str] = []
    for cand in [
        preferred_dir,
        globals().get("BASE_DIR"),
        os.environ.get("SPF_BASE_DIR"),
        os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else None,
        os.path.dirname(os.path.abspath(__file__)) if "__file__" in globals() else None,
        os.getcwd(),
        tempfile.gettempdir(),
    ]:
        if not cand:
            continue
        try:
            abs_cand = os.path.abspath(cand)
        except Exception:
            continue
        if abs_cand not in candidates:
            candidates.append(abs_cand)

    for base in candidates:
        try:
            os.makedirs(base, exist_ok=True)
            probe_path = os.path.join(base, "__spf_write_test__.tmp")
            with open(probe_path, "w", encoding="utf-8") as fp:
                fp.write("test")
            os.remove(probe_path)
            return base
        except Exception:
            continue
    return tempfile.gettempdir()


# 초보자 설명: 디버그 로그 txt 파일 경로를 만든다.
def build_debug_log_path(preferred_dir: Optional[str] = None) -> str:
    base = _pick_debug_log_base_dir(preferred_dir)
    return os.path.join(base, "spf_file_read_debug.txt")


# 초보자 설명: 현재 세션에서 사용할 디버그 로그 파일을 설정한다.
def set_debug_log_path(path: Optional[str] = None, preferred_dir: Optional[str] = None) -> str:
    global DEBUG_LOG_PATH
    new_path = os.path.abspath(path) if path else build_debug_log_path(preferred_dir)
    changed = (DEBUG_LOG_PATH != new_path)
    DEBUG_LOG_PATH = new_path
    try:
        os.makedirs(os.path.dirname(DEBUG_LOG_PATH), exist_ok=True)
    except Exception:
        pass
    if changed:
        try:
            with open(DEBUG_LOG_PATH, "a", encoding="utf-8") as fp:
                fp.write("\n" + "=" * 78 + "\n")
                fp.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 새 디버그 세션 시작\n")
                fp.write(f"로그 파일: {DEBUG_LOG_PATH}\n")
        except Exception:
            pass
    return DEBUG_LOG_PATH


# 초보자 설명: 현재 디버그 로그 파일 경로를 돌려준다.
def get_debug_log_path(preferred_dir: Optional[str] = None) -> str:
    global DEBUG_LOG_PATH
    if not DEBUG_LOG_PATH:
        return set_debug_log_path(preferred_dir=preferred_dir)
    return DEBUG_LOG_PATH


# 초보자 설명: 문자열 한 줄을 디버그 로그 txt 파일에 남긴다.
def debug_log(message: str, preferred_dir: Optional[str] = None) -> None:
    path = get_debug_log_path(preferred_dir=preferred_dir)
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{stamp}] {message}\n"
    try:
        with open(path, "a", encoding="utf-8") as fp:
            fp.write(line)
    except Exception:
        pass


# 초보자 설명: 예외와 traceback을 디버그 로그에 자세히 남긴다.
def debug_log_exception(context: str, exc: Exception, preferred_dir: Optional[str] = None) -> None:
    debug_log(f"[EXCEPTION] {context} | {type(exc).__name__}: {exc}", preferred_dir=preferred_dir)
    tb = traceback.format_exc()
    if tb and tb.strip() != "NoneType: None":
        debug_log(tb.rstrip(), preferred_dir=preferred_dir)


# 초보자 설명: 파일 맨 앞 바이트를 읽어 형식 추정에 도움을 준다.
def read_file_header_bytes(path: str, size: int = 16) -> bytes:
    try:
        with open(path, "rb") as fp:
            return fp.read(size)
    except Exception:
        return b""


# 초보자 설명: 파일 헤더 바이트를 사람이 보기 쉬운 설명으로 바꾼다.
def describe_file_signature(header: bytes) -> str:
    if not header:
        return "헤더를 읽지 못함"
    if header.startswith(b"PK"):
        return "ZIP 계열 서명(PK) - 일반적인 .xlsx 후보"
    if header.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "OLE 서명 - 구형 Excel(.xls) 또는 다른 OLE 문서 후보"
    if header.startswith(b"<## NASCA DRM FILE"):
        return "NASCA DRM 보호 파일 서명"
    if header.startswith(b"<##"):
        return "텍스트 기반 보안/DRM 래퍼 헤더"
    return "알 수 없는 서명"


EXCEL_ACCESS_MODE_CACHE = {}


def _get_excel_file_cache_key(path: str):
    try:
        st = os.stat(path)
        return (os.path.abspath(path), int(st.st_size), int(st.st_mtime))
    except Exception:
        return (os.path.abspath(path), None, None)


def _get_cached_excel_access_mode(path: str):
    key = _get_excel_file_cache_key(path)
    return EXCEL_ACCESS_MODE_CACHE.get(key)


def _set_cached_excel_access_mode(path: str, mode: str) -> None:
    key = _get_excel_file_cache_key(path)
    EXCEL_ACCESS_MODE_CACHE[key] = mode


def detect_excel_access_mode(path: str, preferred_dir: Optional[str] = None) -> str:
    cached = _get_cached_excel_access_mode(path)
    if cached:
        debug_log(f"detect_excel_access_mode 캐시 사용 | path={path} | mode={cached}", preferred_dir=preferred_dir)
        return cached

    if not os.path.exists(path):
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {path}")

    try:
        is_zip = zipfile.is_zipfile(path)
    except Exception as exc:
        debug_log_exception(f"detect_excel_access_mode zip 판정 실패: {path}", exc, preferred_dir=preferred_dir)
        raise

    if is_zip:
        _set_cached_excel_access_mode(path, "zip")
        return "zip"

    header = read_file_header_bytes(path, 64)
    signature_hint = describe_file_signature(header)
    debug_log(
        f"detect_excel_access_mode non-zip 감지 | path={path} | signature_hint={signature_hint}",
        preferred_dir=preferred_dir,
    )

    if platform.system().lower() == "windows" and win32 is not None:
        _set_cached_excel_access_mode(path, "com")
        return "com"

    _set_cached_excel_access_mode(path, "unsupported")
    return "unsupported"


def _safe_set_excel_property(excel, attr_name, value, preferred_dir: Optional[str] = None) -> bool:
    try:
        setattr(excel, attr_name, value)
        return True
    except Exception as exc:
        debug_log_exception(f"Excel.Application 속성 설정 실패: {attr_name}={value}", exc, preferred_dir=preferred_dir)
        return False


def _safe_close_workbook(wb, path: str, preferred_dir: Optional[str] = None) -> None:
    if wb is None:
        return
    try:
        wb.Close(SaveChanges=False)
    except Exception as exc:
        debug_log_exception(f"excel_read_via_com wb.Close 실패: {path}", exc, preferred_dir=preferred_dir)


def _safe_quit_excel_application(excel, path: str, preferred_dir: Optional[str] = None) -> None:
    if excel is None:
        return
    try:
        excel.Quit()
    except Exception as exc:
        debug_log_exception(f"excel_read_via_com excel.Quit 실패: {path}", exc, preferred_dir=preferred_dir)



# 초보자 설명: 현재 파일을 zip/xlsx 관점에서 자세히 조사해 로그로 남긴다.
def inspect_excel_file(path: str, label: str = "", preferred_dir: Optional[str] = None) -> None:
    label_text = f" [{label}]" if label else ""
    debug_log(f"--- 파일 조사 시작{label_text} ---", preferred_dir=preferred_dir)
    debug_log(f"path = {path}", preferred_dir=preferred_dir)
    try:
        abs_path = os.path.abspath(path)
    except Exception:
        abs_path = path
    debug_log(f"abs_path = {abs_path}", preferred_dir=preferred_dir)
    debug_log(f"exists = {os.path.exists(path)} | isfile = {os.path.isfile(path)} | readable = {os.access(path, os.R_OK) if os.path.exists(path) else False}", preferred_dir=preferred_dir)

    if os.path.exists(path):
        try:
            st = os.stat(path)
            debug_log(
                f"size = {st.st_size} bytes | mtime = {datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')} | mode = {oct(st.st_mode)}",
                preferred_dir=preferred_dir,
            )
        except Exception as exc:
            debug_log_exception(f"os.stat 실패: {path}", exc, preferred_dir=preferred_dir)

    header = read_file_header_bytes(path, 32)
    debug_log(f"header_hex = {header.hex(' ') if header else '(읽기 실패)'}", preferred_dir=preferred_dir)
    debug_log(f"signature = {describe_file_signature(header)}", preferred_dir=preferred_dir)

    try:
        is_zip = zipfile.is_zipfile(path)
        debug_log(f"zipfile.is_zipfile = {is_zip}", preferred_dir=preferred_dir)
    except Exception as exc:
        debug_log_exception(f"zipfile.is_zipfile 실패: {path}", exc, preferred_dir=preferred_dir)
        is_zip = False

    if is_zip:
        try:
            with zipfile.ZipFile(path, "r") as zf:
                names = zf.namelist()
                debug_log(f"zip entry count = {len(names)}", preferred_dir=preferred_dir)
                debug_log(f"zip first entries = {names[:20]}", preferred_dir=preferred_dir)
                required = ["[Content_Types].xml", "xl/workbook.xml"]
                missing = [name for name in required if name not in names]
                debug_log(f"zip required missing = {missing if missing else '없음'}", preferred_dir=preferred_dir)
        except Exception as exc:
            debug_log_exception(f"ZipFile 열기 실패: {path}", exc, preferred_dir=preferred_dir)

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            debug_log(f"openpyxl.load_workbook 성공 | sheets = {list(wb.sheetnames)}", preferred_dir=preferred_dir)
        finally:
            wb.close()
    except Exception as exc:
        debug_log_exception(f"openpyxl.load_workbook 실패: {path}", exc, preferred_dir=preferred_dir)

    debug_log(f"--- 파일 조사 종료{label_text} ---", preferred_dir=preferred_dir)


# 초보자 설명: 엑셀 읽기 방법(COM/pandas/openpyxl)을 실제로 시험해 로그에 남긴다.
def probe_excel_read_methods(path: str, sheet_name=0, skip_header=False, preferred_dir: Optional[str] = None) -> None:
    debug_log(
        f"read probe 시작 | file={path} | sheet_name={sheet_name} | skip_header={skip_header} | platform={platform.platform()} | python={sys.version.split()[0]}",
        preferred_dir=preferred_dir,
    )

    try:
        rows, _book, used_sheet_name = excel_read_via_com(path, sheet_name=sheet_name, skip_header=skip_header)
        debug_log(f"COM probe 성공 | rows={len(rows)} | used_sheet_name={used_sheet_name}", preferred_dir=preferred_dir)
    except Exception as exc:
        debug_log_exception(f"COM probe 실패: {path}", exc, preferred_dir=preferred_dir)

    try:
        if skip_header:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        else:
            df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
        debug_log(f"pandas/openpyxl probe 성공 | shape={df.shape}", preferred_dir=preferred_dir)
    except Exception as exc:
        debug_log_exception(f"pandas/openpyxl probe 실패: {path}", exc, preferred_dir=preferred_dir)


# 초보자 설명: 오류 메시지에 디버그 로그 파일 위치를 함께 붙여 준다.
def format_exception_with_debug_log(exc: Exception, preferred_dir: Optional[str] = None) -> str:
    log_path = get_debug_log_path(preferred_dir=preferred_dir)
    return f"{exc}\n\n디버그 로그 파일:\n{log_path}"


AUTO_RECOVER_MISSING_STATION_TOKEN = True
# 입력/실행 상태 전역값
CT = []
cycle_index = {}
spool_fit_detail = {}
spool_length_pair = {}
plan = []
srtPlan_rows = []
plan1_rows = []
plan2_rows = []
srtPlan_sim_rows = []
plan1_sim_rows = []
plan2_sim_rows = []
product_input_rows = []
tmp_product_headers = []
tmp_product_rows = []
CYCLE_DICT_CACHE = {}


# 초보자 설명: 중간 repaired 파일을 만들지 않고 원본 xlsx만 검증한다.
def validate_input_xlsx(path):
    """중간 repaired 파일을 만들지 않고 원본 xlsx만 검증한다."""
    debug_log(f"validate_input_xlsx 호출 | path={path}", preferred_dir=os.path.dirname(path) or None)
    validate_xlsx_file(path)
    return path

# 초보자 설명: Windows Excel COM을 이용해 엑셀을 직접 읽는다. 서식이 복잡한 파일에서 보조적으로 사용된다.
def excel_read_via_com(path, sheet_name=0, skip_header=False, max_retries=3):
    preferred_dir = os.path.dirname(path) or None
    debug_log(
        f"excel_read_via_com 시작 | file={path} | sheet_name={sheet_name} | skip_header={skip_header} | max_retries={max_retries}",
        preferred_dir=preferred_dir,
    )
    if win32 is None:
        raise RuntimeError("win32com.client를 사용할 수 없는 환경입니다.")
    if platform.system().lower() != "windows":
        raise RuntimeError("win32com 방식은 Windows 환경에서만 사용할 수 있습니다.")

    abs_path = os.path.abspath(path)
    last_exc = None

    for attempt in range(1, max_retries + 1):
        excel = None
        wb = None
        used_sheet_name = sheet_name
        co_initialized = False
        dispatch_method = None
        try:
            if pythoncom is not None:
                pythoncom.CoInitialize()
                co_initialized = True

            dispatch_builders = [
                ("DispatchEx", lambda: win32.DispatchEx("Excel.Application")),
                ("EnsureDispatch", lambda: win32.gencache.EnsureDispatch("Excel.Application")),
                ("Dispatch", lambda: win32.Dispatch("Excel.Application")),
            ]

            for method_name, builder in dispatch_builders:
                try:
                    excel = builder()
                    dispatch_method = method_name
                    debug_log(
                        f"excel COM Application 생성 성공 | method={method_name} | attempt={attempt}",
                        preferred_dir=preferred_dir,
                    )
                    break
                except Exception as exc:
                    debug_log_exception(
                        f"excel COM Application 생성 실패 | method={method_name} | attempt={attempt}",
                        exc,
                        preferred_dir=preferred_dir,
                    )
                    excel = None

            if excel is None:
                raise RuntimeError("Excel.Application COM 객체를 생성하지 못했습니다.")

            _safe_set_excel_property(excel, "Visible", False, preferred_dir=preferred_dir)
            _safe_set_excel_property(excel, "DisplayAlerts", False, preferred_dir=preferred_dir)
            _safe_set_excel_property(excel, "ScreenUpdating", False, preferred_dir=preferred_dir)
            _safe_set_excel_property(excel, "EnableEvents", False, preferred_dir=preferred_dir)
            time.sleep(0.2)

            debug_log(
                f"excel COM Open 시도 | abs_path={abs_path} | method={dispatch_method} | attempt={attempt}",
                preferred_dir=preferred_dir,
            )
            wb = excel.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=True)
            if isinstance(sheet_name, int):
                ws = wb.Worksheets(sheet_name + 1)
                used_sheet_name = sheet_name
            else:
                ws = wb.Worksheets(sheet_name)
                used_sheet_name = sheet_name

            used_range = ws.UsedRange
            n_row = int(used_range.Rows.Count)
            n_col = int(used_range.Columns.Count)
            if n_row <= 0 or n_col <= 0:
                rows = []
            else:
                values = ws.Range(ws.Cells(1, 1), ws.Cells(n_row, n_col)).Value
                if values is None:
                    rows = []
                elif n_row == 1 and n_col == 1:
                    rows = [(values,)]
                elif n_row == 1:
                    rows = [tuple(values)]
                else:
                    rows = [tuple(r) if isinstance(r, tuple) else (r,) for r in values]

            if skip_header and rows:
                rows = rows[1:]

            debug_log(
                f"excel_read_via_com 성공 | rows={len(rows)} | used_sheet_name={used_sheet_name} | method={dispatch_method} | attempt={attempt}",
                preferred_dir=preferred_dir,
            )
            _set_cached_excel_access_mode(path, "com")

            class DummyBook:
                def close(self):
                    pass

            return rows, DummyBook(), used_sheet_name
        except Exception as exc:
            last_exc = exc
            debug_log_exception(
                f"excel_read_via_com 실패: {path} | attempt={attempt} | method={dispatch_method}",
                exc,
                preferred_dir=preferred_dir,
            )
        finally:
            _safe_close_workbook(wb, path, preferred_dir=preferred_dir)
            _safe_quit_excel_application(excel, path, preferred_dir=preferred_dir)
            if co_initialized and pythoncom is not None:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        if attempt < max_retries:
            wait_sec = min(1.5, 0.5 * attempt)
            debug_log(
                f"excel_read_via_com 재시도 대기 | file={path} | next_attempt={attempt + 1} | wait_sec={wait_sec}",
                preferred_dir=preferred_dir,
            )
            time.sleep(wait_sec)

    raise last_exc if last_exc is not None else RuntimeError("Excel COM으로 파일을 읽지 못했습니다.")


# 초보자 설명: 가능하면 COM으로 읽고, 실패하면 pandas/openpyxl로 읽는 자동 선택 함수다.
def open_xlsx_reader_auto(path, skip_header=False, sheet_name=0):
    preferred_dir = os.path.dirname(path) or None
    debug_log(
        f"open_xlsx_reader_auto 시작 | file={path} | sheet_name={sheet_name} | skip_header={skip_header}",
        preferred_dir=preferred_dir,
    )

    if not os.path.exists(path):
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {path}")

    access_mode = detect_excel_access_mode(path, preferred_dir=preferred_dir)
    debug_log(f"open_xlsx_reader_auto access_mode={access_mode} | file={path}", preferred_dir=preferred_dir)

    com_exc = None
    pandas_exc = None

    if access_mode == "zip":
        try:
            if skip_header:
                df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            else:
                df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
            debug_log(f"open_xlsx_reader_auto pandas/openpyxl 성공 | file={path} | shape={df.shape}", preferred_dir=preferred_dir)
            _set_cached_excel_access_mode(path, "zip")
            df = df.where(pd.notna(df), None)
            rows = [tuple(row) for row in df.itertuples(index=False, name=None)]

            class DummyBook:
                def close(self):
                    pass

            return rows, DummyBook(), sheet_name
        except Exception as e:
            pandas_exc = e
            debug_log_exception(f"open_xlsx_reader_auto pandas/openpyxl 실패: {path}", e, preferred_dir=preferred_dir)

        if platform.system().lower() == "windows" and win32 is not None:
            try:
                rows, book, used_sheet_name = excel_read_via_com(path, sheet_name=sheet_name, skip_header=skip_header)
                print(f"[INFO] Excel COM으로 읽기 성공: {os.path.basename(path)}")
                debug_log(
                    f"open_xlsx_reader_auto COM fallback 성공 | file={path} | rows={len(rows)} | used_sheet_name={used_sheet_name}",
                    preferred_dir=preferred_dir,
                )
                return rows, book, used_sheet_name
            except Exception as e:
                com_exc = e
                debug_log_exception(f"open_xlsx_reader_auto COM fallback 실패: {path}", e, preferred_dir=preferred_dir)
    elif access_mode == "com":
        try:
            rows, book, used_sheet_name = excel_read_via_com(path, sheet_name=sheet_name, skip_header=skip_header)
            print(f"[INFO] Excel COM으로 읽기 성공: {os.path.basename(path)}")
            debug_log(
                f"open_xlsx_reader_auto COM 성공 | file={path} | rows={len(rows)} | used_sheet_name={used_sheet_name}",
                preferred_dir=preferred_dir,
            )
            return rows, book, used_sheet_name
        except Exception as e:
            com_exc = e
            print(f"[WARN] Excel COM 읽기 실패: {os.path.basename(path)} -> {e}")
            debug_log_exception(f"open_xlsx_reader_auto COM 실패: {path}", e, preferred_dir=preferred_dir)

        try:
            if skip_header:
                df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            else:
                df = pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")
            debug_log(f"open_xlsx_reader_auto pandas/openpyxl fallback 성공 | file={path} | shape={df.shape}", preferred_dir=preferred_dir)
            df = df.where(pd.notna(df), None)
            rows = [tuple(row) for row in df.itertuples(index=False, name=None)]

            class DummyBook:
                def close(self):
                    pass

            return rows, DummyBook(), sheet_name
        except Exception as e:
            pandas_exc = e
            debug_log_exception(f"open_xlsx_reader_auto pandas/openpyxl fallback 실패: {path}", e, preferred_dir=preferred_dir)
    else:
        inspect_excel_file(path, label="open_xlsx_reader_auto unsupported", preferred_dir=preferred_dir)
        raise RuntimeError(
            f"'{os.path.basename(path)}' 파일을 현재 환경에서는 읽을 수 없습니다.\n\n"
            f"디버그 로그 파일:\n{get_debug_log_path(preferred_dir=preferred_dir)}"
        )

    inspect_excel_file(path, label="open_xlsx_reader_auto 실패 후 재조사", preferred_dir=preferred_dir)
    signature_hint = describe_file_signature(read_file_header_bytes(path, 64))
    raise RuntimeError(
        f"'{os.path.basename(path)}' 파일을 읽는 중 오류가 발생했습니다.\n"
        f"- 현재 파일 서명 추정: {signature_hint}\n"
        f"- access_mode: {access_mode}\n"
        f"- COM 오류: {com_exc}\n"
        f"- pandas/openpyxl 오류: {pandas_exc}\n"
        f"- 보안 파일이라면 Excel COM 연결이 일시적으로 끊어졌을 수 있습니다. 열려 있는 Excel 창을 모두 닫고 다시 시도해 주세요.\n\n"
        f"디버그 로그 파일:\n{get_debug_log_path(preferred_dir=preferred_dir)}"
    )



# =====================================================
# Excel COM 상세 디버깅 보강
# -----------------------------------------------------
# 보안/DRM 환경에서 COM이 어느 단계에서 끊기는지 더 명확히 확인하기 위해
# Application 생성 -> 속성 설정 -> Workbook Open -> Worksheet 접근 -> UsedRange 읽기
# -> 값 추출 -> Close/Quit 전 과정을 단계별로 로그에 남긴다.
# =====================================================

EXCEL_COM_CALL_COUNTER = 0


# 초보자 설명: Excel COM 읽기 호출마다 고유 번호를 붙여 로그를 구분한다.
def _next_excel_com_call_id():
    global EXCEL_COM_CALL_COUNTER
    EXCEL_COM_CALL_COUNTER += 1
    return EXCEL_COM_CALL_COUNTER


# 초보자 설명: 현재 프로세스/스레드 정보를 디버그 문자열로 만든다.
def _excel_debug_thread_tag():
    try:
        import threading
        return f"pid={os.getpid()} | tid={threading.get_ident()}"
    except Exception:
        return f"pid={os.getpid()}"


# 초보자 설명: COM 객체의 속성을 안전하게 읽는다.
def _safe_com_attr(obj, attr_name, default=None):
    try:
        return getattr(obj, attr_name)
    except Exception:
        return default


# 초보자 설명: 값이 너무 길면 로그에 적당한 길이만 남긴다.
def _trim_debug_value(value, max_len=240):
    try:
        text = repr(value)
    except Exception:
        text = str(type(value))
    if len(text) > max_len:
        return text[:max_len] + " ..."
    return text


# 초보자 설명: 현재 Windows의 EXCEL.EXE 프로세스 상태를 로그로 남긴다.
def _log_excel_process_snapshot(label, preferred_dir: Optional[str] = None):
    if platform.system().lower() != "windows":
        debug_log(f"Excel 프로세스 스냅샷 생략 | label={label} | non-windows", preferred_dir=preferred_dir)
        return
    try:
        import subprocess
        result = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq EXCEL.EXE", "/FO", "CSV", "/NH"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        stdout = (result.stdout or "").strip().splitlines()
        stderr = (result.stderr or "").strip()
        preview = stdout[:5]
        debug_log(
            f"Excel 프로세스 스냅샷 | label={label} | returncode={result.returncode} | rows={len(stdout)} | preview={preview} | stderr={stderr}",
            preferred_dir=preferred_dir,
        )
    except Exception as exc:
        debug_log_exception(f"Excel 프로세스 스냅샷 실패 | label={label}", exc, preferred_dir=preferred_dir)


# 초보자 설명: Excel Application 객체의 핵심 상태를 단계별로 로그에 남긴다.
def _log_excel_application_state(excel, stage, call_id, attempt, preferred_dir: Optional[str] = None):
    if excel is None:
        debug_log(f"Excel.Application 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | excel=None", preferred_dir=preferred_dir)
        return
    try:
        wb_count = None
        try:
            workbooks = getattr(excel, "Workbooks")
            wb_count = getattr(workbooks, "Count")
        except Exception:
            wb_count = None
        state = {
            "Version": _safe_com_attr(excel, "Version"),
            "Hwnd": _safe_com_attr(excel, "Hwnd"),
            "Visible": _safe_com_attr(excel, "Visible"),
            "DisplayAlerts": _safe_com_attr(excel, "DisplayAlerts"),
            "ScreenUpdating": _safe_com_attr(excel, "ScreenUpdating"),
            "EnableEvents": _safe_com_attr(excel, "EnableEvents"),
            "Ready": _safe_com_attr(excel, "Ready"),
            "Interactive": _safe_com_attr(excel, "Interactive"),
            "Workbooks.Count": wb_count,
        }
        debug_log(
            f"Excel.Application 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | state={state}",
            preferred_dir=preferred_dir,
        )
    except Exception as exc:
        debug_log_exception(
            f"Excel.Application 상태 조회 실패 | call_id={call_id} | attempt={attempt} | stage={stage}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: Workbook 객체 상태를 로그에 남긴다.
def _log_excel_workbook_state(wb, stage, call_id, attempt, preferred_dir: Optional[str] = None):
    if wb is None:
        debug_log(f"Workbook 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | wb=None", preferred_dir=preferred_dir)
        return
    try:
        state = {
            "Name": _safe_com_attr(wb, "Name"),
            "FullName": _safe_com_attr(wb, "FullName"),
            "ReadOnly": _safe_com_attr(wb, "ReadOnly"),
            "Saved": _safe_com_attr(wb, "Saved"),
        }
        try:
            sheets = getattr(wb, "Worksheets")
            state["Worksheets.Count"] = getattr(sheets, "Count")
        except Exception:
            state["Worksheets.Count"] = None
        debug_log(
            f"Workbook 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | state={state}",
            preferred_dir=preferred_dir,
        )
    except Exception as exc:
        debug_log_exception(
            f"Workbook 상태 조회 실패 | call_id={call_id} | attempt={attempt} | stage={stage}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: Worksheet 객체 상태를 로그에 남긴다.
def _log_excel_worksheet_state(ws, stage, call_id, attempt, preferred_dir: Optional[str] = None):
    if ws is None:
        debug_log(f"Worksheet 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | ws=None", preferred_dir=preferred_dir)
        return
    try:
        state = {
            "Name": _safe_com_attr(ws, "Name"),
            "Index": _safe_com_attr(ws, "Index"),
            "Visible": _safe_com_attr(ws, "Visible"),
        }
        debug_log(
            f"Worksheet 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | state={state}",
            preferred_dir=preferred_dir,
        )
    except Exception as exc:
        debug_log_exception(
            f"Worksheet 상태 조회 실패 | call_id={call_id} | attempt={attempt} | stage={stage}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: UsedRange 정보(주소/행/열 수)를 로그에 남긴다.
def _log_excel_usedrange_state(used_range, stage, call_id, attempt, preferred_dir: Optional[str] = None):
    if used_range is None:
        debug_log(f"UsedRange 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | used_range=None", preferred_dir=preferred_dir)
        return
    try:
        state = {
            "Address": _safe_com_attr(used_range, "Address"),
        }
        try:
            state["Rows.Count"] = getattr(getattr(used_range, "Rows"), "Count")
        except Exception:
            state["Rows.Count"] = None
        try:
            state["Columns.Count"] = getattr(getattr(used_range, "Columns"), "Count")
        except Exception:
            state["Columns.Count"] = None
        debug_log(
            f"UsedRange 상태 | call_id={call_id} | attempt={attempt} | stage={stage} | state={state}",
            preferred_dir=preferred_dir,
        )
    except Exception as exc:
        debug_log_exception(
            f"UsedRange 상태 조회 실패 | call_id={call_id} | attempt={attempt} | stage={stage}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: Excel 속성 설정 성공/실패와 실제 반영값을 로그에 남긴다.
def _safe_set_excel_property(excel, attr_name, value, preferred_dir: Optional[str] = None, call_id=None, attempt=None) -> bool:
    try:
        setattr(excel, attr_name, value)
        actual = _safe_com_attr(excel, attr_name, default="(확인 실패)")
        debug_log(
            f"Excel.Application 속성 설정 성공 | call_id={call_id} | attempt={attempt} | {attr_name}={value} | actual={actual}",
            preferred_dir=preferred_dir,
        )
        return True
    except Exception as exc:
        debug_log_exception(
            f"Excel.Application 속성 설정 실패 | call_id={call_id} | attempt={attempt} | {attr_name}={value}",
            exc,
            preferred_dir=preferred_dir,
        )
        return False


# 초보자 설명: Workbook Close 성공/실패를 로그에 남긴다.
def _safe_close_workbook(wb, path: str, preferred_dir: Optional[str] = None, call_id=None, attempt=None) -> None:
    if wb is None:
        debug_log(f"Workbook Close 생략 | call_id={call_id} | attempt={attempt} | wb=None | path={path}", preferred_dir=preferred_dir)
        return
    try:
        wb.Close(SaveChanges=False)
        debug_log(f"Workbook Close 성공 | call_id={call_id} | attempt={attempt} | path={path}", preferred_dir=preferred_dir)
    except Exception as exc:
        debug_log_exception(
            f"excel_read_via_com wb.Close 실패 | call_id={call_id} | attempt={attempt} | path={path}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: Excel Application Quit 성공/실패를 로그에 남긴다.
def _safe_quit_excel_application(excel, path: str, preferred_dir: Optional[str] = None, call_id=None, attempt=None) -> None:
    if excel is None:
        debug_log(f"Excel Quit 생략 | call_id={call_id} | attempt={attempt} | excel=None | path={path}", preferred_dir=preferred_dir)
        return
    try:
        excel.Quit()
        debug_log(f"Excel Quit 성공 | call_id={call_id} | attempt={attempt} | path={path}", preferred_dir=preferred_dir)
    except Exception as exc:
        debug_log_exception(
            f"excel_read_via_com excel.Quit 실패 | call_id={call_id} | attempt={attempt} | path={path}",
            exc,
            preferred_dir=preferred_dir,
        )


# 초보자 설명: 경로 비교용으로 Windows 경로 문자열을 정규화한다.
def _normalize_excel_path_for_compare(path_value):
    try:
        s = os.path.abspath(str(path_value))
    except Exception:
        s = str(path_value)
    return os.path.normcase(os.path.normpath(s)).strip()


# 초보자 설명: 현재 Excel 인스턴스에서 대상 파일이 이미 열려 있는지 찾는다.
def _find_open_workbook_in_excel_app(excel, abs_path, preferred_dir: Optional[str] = None, call_id=None, attempt=None):
    target = _normalize_excel_path_for_compare(abs_path)
    try:
        workbooks = excel.Workbooks
        count = int(getattr(workbooks, 'Count'))
    except Exception as exc:
        debug_log_exception(
            f"열린 Workbook 목록 조회 실패 | call_id={call_id} | attempt={attempt} | abs_path={abs_path}",
            exc,
            preferred_dir=preferred_dir,
        )
        return None

    debug_log(
        f"열린 Workbook 검사 시작 | call_id={call_id} | attempt={attempt} | workbook_count={count} | abs_path={abs_path}",
        preferred_dir=preferred_dir,
    )
    for idx in range(1, count + 1):
        try:
            wb = workbooks(idx)
            full_name = _safe_com_attr(wb, 'FullName', default='')
            wb_name = _safe_com_attr(wb, 'Name', default='')
            full_name_norm = _normalize_excel_path_for_compare(full_name) if full_name else ''
            name_match = bool(wb_name and norm_text(wb_name).lower() == norm_text(os.path.basename(abs_path)).lower())
            full_match = bool(full_name_norm and full_name_norm == target)
            debug_log(
                f"열린 Workbook 후보 | call_id={call_id} | attempt={attempt} | idx={idx} | Name={wb_name} | FullName={full_name} | full_match={full_match} | name_match={name_match}",
                preferred_dir=preferred_dir,
            )
            if full_match or name_match:
                debug_log(
                    f"열린 Workbook 일치 발견 | call_id={call_id} | attempt={attempt} | idx={idx} | Name={wb_name} | FullName={full_name}",
                    preferred_dir=preferred_dir,
                )
                return wb
        except Exception as exc:
            debug_log_exception(
                f"열린 Workbook 후보 검사 실패 | call_id={call_id} | attempt={attempt} | idx={idx}",
                exc,
                preferred_dir=preferred_dir,
            )
    return None


# 초보자 설명: 사용자가 이미 열어 둔 Excel/Workbook에 붙을 수 있으면 붙어서 재사용한다.
def _try_attach_open_workbook(abs_path, preferred_dir: Optional[str] = None, call_id=None, attempt=None):
    attached_excel = None
    attached_wb = None

    get_object = getattr(win32, 'GetObject', None) if win32 is not None else None
    if get_object is None:
        debug_log(
            f"GetObject 미지원 환경 | call_id={call_id} | attempt={attempt} | abs_path={abs_path}",
            preferred_dir=preferred_dir,
        )
        return None, None, None

    # 1) 실행 중인 Excel.Application에 먼저 붙어서 열린 Workbook 목록에서 찾는다.
    try:
        debug_log(
            f"실행 중 Excel.Application attach 시도 | call_id={call_id} | attempt={attempt}",
            preferred_dir=preferred_dir,
        )
        attached_excel = get_object(Class='Excel.Application')
        debug_log(
            f"실행 중 Excel.Application attach 성공 | call_id={call_id} | attempt={attempt} | excel_type={type(attached_excel)}",
            preferred_dir=preferred_dir,
        )
        _log_excel_application_state(attached_excel, 'attached-existing-app', call_id, attempt, preferred_dir=preferred_dir)
        attached_wb = _find_open_workbook_in_excel_app(attached_excel, abs_path, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
        if attached_wb is not None:
            _log_excel_workbook_state(attached_wb, 'attached-existing-workbook', call_id, attempt, preferred_dir=preferred_dir)
            return 'existing_excel_application', attached_excel, attached_wb
    except Exception as exc:
        debug_log_exception(
            f"실행 중 Excel.Application attach 실패 | call_id={call_id} | attempt={attempt}",
            exc,
            preferred_dir=preferred_dir,
        )

    # 2) 경로로 직접 Workbook 객체 attach를 시도한다.
    try:
        debug_log(
            f"Workbook 직접 attach 시도 | call_id={call_id} | attempt={attempt} | abs_path={abs_path}",
            preferred_dir=preferred_dir,
        )
        attached_wb = get_object(abs_path)
        attached_excel = _safe_com_attr(attached_wb, 'Application', default=None)
        debug_log(
            f"Workbook 직접 attach 성공 | call_id={call_id} | attempt={attempt} | wb_type={type(attached_wb)} | excel_type={type(attached_excel)}",
            preferred_dir=preferred_dir,
        )
        _log_excel_workbook_state(attached_wb, 'attached-direct-workbook', call_id, attempt, preferred_dir=preferred_dir)
        if attached_excel is not None:
            _log_excel_application_state(attached_excel, 'attached-direct-application', call_id, attempt, preferred_dir=preferred_dir)
        return 'direct_workbook', attached_excel, attached_wb
    except Exception as exc:
        debug_log_exception(
            f"Workbook 직접 attach 실패 | call_id={call_id} | attempt={attempt} | abs_path={abs_path}",
            exc,
            preferred_dir=preferred_dir,
        )

    return None, None, None


# 초보자 설명: Workbook/Worksheet/UsedRange에서 실제 셀 값을 읽어 rows로 변환한다.
def _read_rows_from_workbook(wb, sheet_name=0, skip_header=False, preferred_dir: Optional[str] = None, call_id=None, attempt=None):
    used_sheet_name = sheet_name
    ws = None
    used_range = None

    debug_log(
        f"Worksheet 접근 시도 | call_id={call_id} | attempt={attempt} | sheet_name={sheet_name} | sheet_name_type={type(sheet_name).__name__}",
        preferred_dir=preferred_dir,
    )
    if isinstance(sheet_name, int):
        ws = wb.Worksheets(sheet_name + 1)
        used_sheet_name = sheet_name
    else:
        ws = wb.Worksheets(sheet_name)
        used_sheet_name = sheet_name
    _log_excel_worksheet_state(ws, 'after-worksheet-select', call_id, attempt, preferred_dir=preferred_dir)

    debug_log(
        f"UsedRange 접근 시도 | call_id={call_id} | attempt={attempt}",
        preferred_dir=preferred_dir,
    )
    used_range = ws.UsedRange
    _log_excel_usedrange_state(used_range, 'after-usedrange', call_id, attempt, preferred_dir=preferred_dir)

    n_row = int(getattr(getattr(used_range, 'Rows'), 'Count'))
    n_col = int(getattr(getattr(used_range, 'Columns'), 'Count'))
    debug_log(
        f"UsedRange 크기 확인 | call_id={call_id} | attempt={attempt} | rows={n_row} | cols={n_col}",
        preferred_dir=preferred_dir,
    )

    if n_row <= 0 or n_col <= 0:
        rows = []
    else:
        debug_log(
            f"Range.Value 추출 시도 | call_id={call_id} | attempt={attempt} | top_left=(1,1) | bottom_right=({n_row},{n_col})",
            preferred_dir=preferred_dir,
        )
        values = ws.Range(ws.Cells(1, 1), ws.Cells(n_row, n_col)).Value
        debug_log(
            f"Range.Value 추출 성공 | call_id={call_id} | attempt={attempt} | value_type={type(values).__name__} | preview={_trim_debug_value(values)}",
            preferred_dir=preferred_dir,
        )
        if values is None:
            rows = []
        elif n_row == 1 and n_col == 1:
            rows = [(values,)]
        elif n_row == 1:
            rows = [tuple(values)]
        else:
            rows = [tuple(r) if isinstance(r, tuple) else (r,) for r in values]

    debug_log(
        f"행 변환 완료 | call_id={call_id} | attempt={attempt} | rows_before_skip={len(rows)} | skip_header={skip_header}",
        preferred_dir=preferred_dir,
    )
    if skip_header and rows:
        rows = rows[1:]
        debug_log(
            f"헤더 건너뜀 적용 | call_id={call_id} | attempt={attempt} | rows_after_skip={len(rows)}",
            preferred_dir=preferred_dir,
        )

    return rows, used_sheet_name


# 초보자 설명: Windows Excel COM을 이용해 엑셀을 직접 읽는다. 단계별 상세 디버그 로그를 남긴다.
def excel_read_via_com(path, sheet_name=0, skip_header=False, max_retries=3):
    preferred_dir = os.path.dirname(path) or None
    call_id = _next_excel_com_call_id()
    debug_log(
        f"excel_read_via_com 시작 | call_id={call_id} | file={path} | sheet_name={sheet_name} | skip_header={skip_header} | max_retries={max_retries} | {_excel_debug_thread_tag()}",
        preferred_dir=preferred_dir,
    )
    _log_excel_process_snapshot(f"call_id={call_id} | before-start", preferred_dir=preferred_dir)

    if win32 is None:
        raise RuntimeError("win32com.client를 사용할 수 없는 환경입니다.")
    if platform.system().lower() != "windows":
        raise RuntimeError("win32com 방식은 Windows 환경에서만 사용할 수 있습니다.")

    abs_path = os.path.abspath(path)
    last_exc = None

    for attempt in range(1, max_retries + 1):
        excel = None
        wb = None
        ws = None
        used_range = None
        used_sheet_name = sheet_name
        co_initialized = False
        dispatch_method = None
        opened_by_us = False
        attached_existing_workbook = False
        debug_log(
            f"excel_read_via_com attempt 시작 | call_id={call_id} | attempt={attempt} | file={path}",
            preferred_dir=preferred_dir,
        )
        _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | pre-dispatch", preferred_dir=preferred_dir)
        try:
            if pythoncom is not None:
                pythoncom.CoInitialize()
                co_initialized = True
                debug_log(
                    f"pythoncom.CoInitialize 성공 | call_id={call_id} | attempt={attempt}",
                    preferred_dir=preferred_dir,
                )
            else:
                debug_log(
                    f"pythoncom 미사용 | call_id={call_id} | attempt={attempt}",
                    preferred_dir=preferred_dir,
                )

            attached_mode, attached_excel, attached_wb = _try_attach_open_workbook(
                abs_path,
                preferred_dir=preferred_dir,
                call_id=call_id,
                attempt=attempt,
            )
            opened_by_us = False
            attached_existing_workbook = False

            if attached_wb is not None:
                wb = attached_wb
                excel = attached_excel
                dispatch_method = f"attach:{attached_mode}"
                attached_existing_workbook = True
                debug_log(
                    f"이미 열려 있는 Workbook 재사용 | call_id={call_id} | attempt={attempt} | attach_mode={attached_mode} | abs_path={abs_path}",
                    preferred_dir=preferred_dir,
                )
                _log_excel_application_state(excel, "after-attach-open-workbook", call_id, attempt, preferred_dir=preferred_dir)
                _log_excel_workbook_state(wb, "after-attach-open-workbook", call_id, attempt, preferred_dir=preferred_dir)
                _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | attached-open-workbook", preferred_dir=preferred_dir)
            else:
                dispatch_builders = [
                    ("DispatchEx", lambda: win32.DispatchEx("Excel.Application")),
                    ("EnsureDispatch", lambda: win32.gencache.EnsureDispatch("Excel.Application")),
                    ("Dispatch", lambda: win32.Dispatch("Excel.Application")),
                ]

                for method_name, builder in dispatch_builders:
                    try:
                        debug_log(
                            f"Excel.Application 생성 시도 | call_id={call_id} | attempt={attempt} | method={method_name}",
                            preferred_dir=preferred_dir,
                        )
                        excel = builder()
                        dispatch_method = method_name
                        debug_log(
                            f"Excel.Application 생성 성공 | call_id={call_id} | attempt={attempt} | method={method_name} | excel_type={type(excel)}",
                            preferred_dir=preferred_dir,
                        )
                        break
                    except Exception as exc:
                        debug_log_exception(
                            f"Excel.Application 생성 실패 | call_id={call_id} | attempt={attempt} | method={method_name}",
                            exc,
                            preferred_dir=preferred_dir,
                        )
                        excel = None

                if excel is None:
                    raise RuntimeError("Excel.Application COM 객체를 생성하지 못했습니다.")

                _log_excel_application_state(excel, "after-dispatch", call_id, attempt, preferred_dir=preferred_dir)
                _safe_set_excel_property(excel, "Visible", False, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
                _safe_set_excel_property(excel, "DisplayAlerts", False, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
                _safe_set_excel_property(excel, "ScreenUpdating", False, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
                _safe_set_excel_property(excel, "EnableEvents", False, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
                time.sleep(0.2)
                _log_excel_application_state(excel, "after-properties", call_id, attempt, preferred_dir=preferred_dir)
                _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | before-open", preferred_dir=preferred_dir)

                debug_log(
                    f"Workbook.Open 시도 | call_id={call_id} | attempt={attempt} | abs_path={abs_path} | method={dispatch_method}",
                    preferred_dir=preferred_dir,
                )
                wb = excel.Workbooks.Open(
                    abs_path,
                    UpdateLinks=0,
                    ReadOnly=True,
                    Notify=False,
                    IgnoreReadOnlyRecommended=True,
                    AddToMru=False,
                )
                opened_by_us = True
                debug_log(
                    f"Workbook.Open 성공 | call_id={call_id} | attempt={attempt} | abs_path={abs_path}",
                    preferred_dir=preferred_dir,
                )
                _log_excel_application_state(excel, "after-open", call_id, attempt, preferred_dir=preferred_dir)
                _log_excel_workbook_state(wb, "after-open", call_id, attempt, preferred_dir=preferred_dir)
                _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | after-open", preferred_dir=preferred_dir)

            rows, used_sheet_name = _read_rows_from_workbook(
                wb,
                sheet_name=sheet_name,
                skip_header=skip_header,
                preferred_dir=preferred_dir,
                call_id=call_id,
                attempt=attempt,
            )

            debug_log(
                f"excel_read_via_com 성공 | call_id={call_id} | attempt={attempt} | rows={len(rows)} | used_sheet_name={used_sheet_name} | method={dispatch_method}",
                preferred_dir=preferred_dir,
            )
            _set_cached_excel_access_mode(path, "com")
            _log_excel_application_state(excel, "before-return", call_id, attempt, preferred_dir=preferred_dir)

            class DummyBook:
                def close(self):
                    pass

            return rows, DummyBook(), used_sheet_name
        except Exception as exc:
            last_exc = exc
            debug_log_exception(
                f"excel_read_via_com 실패 | call_id={call_id} | attempt={attempt} | method={dispatch_method} | file={path}",
                exc,
                preferred_dir=preferred_dir,
            )
            _log_excel_application_state(excel, "on-exception", call_id, attempt, preferred_dir=preferred_dir)
            _log_excel_workbook_state(wb, "on-exception", call_id, attempt, preferred_dir=preferred_dir)
            _log_excel_worksheet_state(ws, "on-exception", call_id, attempt, preferred_dir=preferred_dir)
            _log_excel_usedrange_state(used_range, "on-exception", call_id, attempt, preferred_dir=preferred_dir)
            _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | on-exception", preferred_dir=preferred_dir)
        finally:
            _log_excel_application_state(excel, "before-close-finally", call_id, attempt, preferred_dir=preferred_dir)
            if attached_existing_workbook:
                debug_log(
                    f"열려 있던 Workbook 재사용으로 Close/Quit 생략 | call_id={call_id} | attempt={attempt} | path={path}",
                    preferred_dir=preferred_dir,
                )
            else:
                _safe_close_workbook(wb, path, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
                _safe_quit_excel_application(excel, path, preferred_dir=preferred_dir, call_id=call_id, attempt=attempt)
            if co_initialized and pythoncom is not None:
                try:
                    pythoncom.CoUninitialize()
                    debug_log(
                        f"pythoncom.CoUninitialize 성공 | call_id={call_id} | attempt={attempt}",
                        preferred_dir=preferred_dir,
                    )
                except Exception as exc:
                    debug_log_exception(
                        f"pythoncom.CoUninitialize 실패 | call_id={call_id} | attempt={attempt}",
                        exc,
                        preferred_dir=preferred_dir,
                    )
            _log_excel_process_snapshot(f"call_id={call_id} | attempt={attempt} | after-finally", preferred_dir=preferred_dir)

        if attempt < max_retries:
            wait_sec = min(1.5, 0.5 * attempt)
            debug_log(
                f"excel_read_via_com 재시도 대기 | call_id={call_id} | file={path} | next_attempt={attempt + 1} | wait_sec={wait_sec}",
                preferred_dir=preferred_dir,
            )
            time.sleep(wait_sec)

    raise last_exc if last_exc is not None else RuntimeError("Excel COM으로 파일을 읽지 못했습니다.")


# 초보자 설명: 헤더와 행 데이터를 단일 시트 엑셀 파일로 저장한다.
def save_rows_to_xlsx(path, headers, rows):
    recreate_output_file(path)
    df = pd.DataFrame(rows, columns=headers)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)


# 초보자 설명: 여러 시트를 가진 엑셀 파일을 저장한다.
def save_multi_sheet_xlsx(path, sheet_dict):
    recreate_output_file(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, (headers, rows) in sheet_dict.items():
            df = pd.DataFrame(rows, columns=headers)
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# 초보자 설명: 같은 이름의 출력 파일이 있으면 먼저 지워 저장 오류를 막는다.
def recreate_output_file(path):
    """기존 파일이 있으면 먼저 삭제한 뒤 새 파일을 생성할 수 있게 준비한다."""
    if not path:
        return
    try:
        if os.path.exists(path):
            os.remove(path)
            print(f"[INFO] 기존 출력 파일 삭제: {path}")
    except Exception as exc:
        raise RuntimeError(f"기존 출력 파일 삭제 실패: {path}\n상세 오류: {exc}")


# =====================================================
# 값 정리용 보조 함수
# -----------------------------------------------------
# 엑셀에서 읽은 값은 문자열, 빈칸, 숫자가 섞여 있기 때문에
# 먼저 비교하기 쉬운 형태로 정리하는 작은 함수들이 필요하다.
# =====================================================

# 초보자 설명: 공정 시간이 비어 있거나 이상한 값일 때도 안전한 숫자로 바꿔 준다.
def safe_ct(value, default=1.0):
    try:
        v = float(value)
        if math.isnan(v) or math.isinf(v):
            return float(default)
        return v
    except:
        return float(default)


# 초보자 설명: 문자열 숫자를 int/float/None 등 적절한 파이썬 값으로 자동 변환한다.
def auto_convert(value):
    value = str(value).replace("\ufeff", "").strip()
    if value == "" or value == "None" or value.lower() == "nan":
        return None
    try:
        f = float(value)
        return int(f) if f.is_integer() else f
    except:
        return value


# 초보자 설명: 비교하기 쉽도록 공백과 BOM을 제거한 문자열로 정리한다.
def norm_text(v):
    if v is None:
        return ""
    return str(v).replace("\ufeff", "").strip()


# 초보자 설명: 숫자처럼 보이는 값을 문자열 기준 키로 맞추기 좋게 정리한다.
def norm_num_text(v):
    if v is None:
        return ""
    s = str(v).replace("\ufeff", "").strip()
    if s == "":
        return ""
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else str(f)
    except:
        return s


# =====================================================
# 제품 타입 / CycleTime 열 이름 정의
# -----------------------------------------------------
# 아래 상수들은 엑셀 열 위치와 내부 표준 이름을 연결해 준다.
# 나중에 숫자 인덱스를 직접 외우지 않게 해 주는 중요한 부분이다.
# =====================================================

PRODUCT_TYPE_ALIASES = {
    "2DS": "2DS",
    "2DL": "2DL",
    "F2DS": "F2DS",
    "F2DL": "F2DL",
    "2D": "2D",
    "F2D": "F2D",
    "H3D": "H3D",
    "F3D": "F3D",
    "FF": "FF",
    "SHORT": "Short",
    "Short": "Short",
    "short": "Short",
    "ff": "FF",
}

DISPLAY_PROC_ALIASES = {
    "Con1": "Conv1",
    "Con2": "Conv2",
    "Bevel1": "Bev1",
    "Bevel2": "Bev2",
}

CT_COL = {
    "inch": 0,
    "sch": 1,
    "prod": 2,
    "weld": 3,
    "length": 4,
    "input": 5,
    "first_cut": 6,
    "cut": 7,
    "conv1": 8,
    "sen1_1": 9,
    "bev1_j": 10,
    "bev1_v": 11,
    "sen2_1": 12,
    "conv2": 13,
    "sen1_2": 14,
    "bev2_j": 15,
    "bev2_v": 16,
    "sen2_2": 17,
    "conv3": 18,
    "gan1_2d1": 19,
    "r2_2d1": 20,
    "gan1_2d2": 21,
    "r2_2d2": 22,
    "gan1_3d": 23,
    "r2_short_long": 24,
    "r2_short_short": 25,
    "gan4": 26,
    "r3_2d2_to_3d": 27,
    "r2_fit_2d1": 28,
    "r2_fit_2d2": 29,
    "r3_fit_3d": 30,
    "r2_fit_short_1": 31,
    "r2_fit_short_2": 32,
    "fitup_spf": 33,
    "fitup_3d": 34,
    "tig_spf": 35,
    "tig2_spf": 36,
    "mag_spf": 37,
    "r14_pick_1": 38,
    "r14_pick_2": 39,
    "fitup_short_1": 40,
    "fitup3d_short_1": 41,
    "tig_short_1": 42,
    "tig2_short_1": 43,
    "mag_short_1": 44,
    "turn_short_2": 45,
    "r14_pick_2b": 46,
    "fitup3d_short_2": 47,
    "tig_short_2": 48,
    "tig2_short_2": 49,
    "mag_short_2": 50,
    "out_2d1": 51,
    "out_2d2": 52,
    "out_3d": 53,
    "out_short": 54,
    "fit_45el": 55,
    "fit_90el": 56,
    "fit_90el_both": 57,
    "fit_fl": 58,
    "fit_tee": 59,
    "shrt_bev1_j": 60,
    "shrt_bev1_v": 61,
    "shrt_bev2_j": 62,
    "shrt_bev2_v": 63,
    "fitmov_short": 64,
    "r1_short": 65,
    "short_single": 60,
    "short_double": 62,
}


# 초보자 설명: 스풀 타입 표기를 내부에서 쓰는 표준 이름으로 통일한다.
def normalize_product_type_name(v):
    s = norm_text(v)
    return PRODUCT_TYPE_ALIASES.get(s, PRODUCT_TYPE_ALIASES.get(s.upper(), s))


# 초보자 설명: 길이 구간 값을 CycleTime 검색에 맞는 문자열로 정리한다.
def normalize_length_bucket(v):
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    s = norm_text(v)
    if s == "":
        return ""
    if s.upper() in {"N/A", "NA", "NONE", "NULL", "-"}:
        return ""
    return s


# 초보자 설명: 제품 타입과 용접장, 길이를 보고 CycleTime의 길이 구간 키를 추론한다.
def infer_length_bucket(prod, weld, length):
    prod = normalize_product_type_name(prod)
    weld = norm_text(weld)
    l = to_num(length, 0.0)

    if prod in ["2DS", "F2DS"]:
        return "300mm이상 ~ 500mm 미만"
    if prod in ["2DL", "F2DL"]:
        return "500mm이상 ~ 600mm 미만"
    if prod == "FF":
        return ""
    if prod == "2D" and weld in ["2D-1", "2D-2"]:
        if 600 <= l < 1700:
            return "600mm이상~1,700mm미만"
        if 1700 <= l <= 10500:
            return "1,700mm이상~10,500mm이하"
    if prod == "2D" and weld == "3D":
        if 600 <= l <= 10500:
            return "600mm이상~10,500mm이하"
    if prod in ["F2D", "F3D"] and weld == "2D-1":
        if 1100 <= l < 1700:
            return "1,100mm이상~1,700mm미만"
        if 1700 <= l <= 10500:
            return "1,700mm이상~10,500mm이하"
    if prod in ["F3D", "H3D"] and weld == "2D-2":
        if 700 <= l < 1700:
            return "700mm이상~1,700mm미만"
    if prod == "H3D" and weld == "3D":
        if 1100 <= l <= 10500:
            return "1,100mm이상~10,500mm이하"
    if prod in ["F2D", "F3D"] and weld == "3D":
        return ""
    return ""


# 초보자 설명: CycleTime 검색에 쓰는 복합 키를 만든다.
def make_cycle_key(inch, sch, prod, weld, length_bucket=""):
    return (
        norm_num_text(inch),
        norm_text(sch),
        normalize_product_type_name(prod),
        norm_text(weld),
        normalize_length_bucket(length_bucket),
    )


# 초보자 설명: CycleTime 행들을 빠르게 찾기 위한 인덱스 사전으로 바꾼다.
def build_cycle_index_from_rows(ct_rows):
    cycle_index_local = {}
    for row_num, row in enumerate(ct_rows, start=1):
        if len(row) < 10:
            continue
        converted_row = tuple(auto_convert(cell) for cell in row)
        key = make_cycle_key(
            converted_row[CT_COL["inch"]],
            converted_row[CT_COL["sch"]],
            converted_row[CT_COL["prod"]],
            converted_row[CT_COL["weld"]],
            converted_row[CT_COL["length"]],
        )
        cycle_index_local.setdefault(key, []).append(row_num)
    return cycle_index_local


# 초보자 설명: CycleTime.xlsx를 읽어서 시뮬레이션에 바로 쓸 수 있는 형태로 준비한다.
def load_cycle_time_data(path):
    preferred_dir = os.path.dirname(path) or None
    debug_log(f"load_cycle_time_data 시작 | file={path}", preferred_dir=preferred_dir)
    try:
        rows, workbook, sheet_name = open_xlsx_reader_auto(path, skip_header=False, sheet_name=0)
        try:
            ct_rows = [tuple(row) for row in rows[3:]]
        finally:
            try:
                workbook.close()
            except Exception:
                pass
        debug_log(
            f"load_cycle_time_data reader 성공 | raw_rows={len(rows)} | data_rows={len(ct_rows)} | sheet_name={sheet_name}",
            preferred_dir=preferred_dir,
        )
    except Exception as e:
        debug_log_exception(f"load_cycle_time_data 실패: {path}", e, preferred_dir=preferred_dir)
        inspect_excel_file(path, label="load_cycle_time_data 실패 후 재조사", preferred_dir=preferred_dir)
        probe_excel_read_methods(path, sheet_name=0, skip_header=False, preferred_dir=preferred_dir)
        raise RuntimeError(
            f"'{os.path.basename(path)}' 파일을 읽는 중 오류가 발생했습니다.\n"
            f"일반 xlsx 또는 Excel COM으로 열 수 있는 보안 파일만 지원합니다.\n"
            f"상세 오류: {e}\n\n"
            f"디버그 로그 파일:\n{get_debug_log_path(preferred_dir=preferred_dir)}"
        )
    cycle_index_local = build_cycle_index_from_rows(ct_rows)
    print(f"[INFO] CycleTime.xlsx sheet = {sheet_name}")
    print(f"[INFO] CycleTime key 개수: {len(cycle_index_local)}")
    debug_log(
        f"load_cycle_time_data 완료 | data_rows={len(ct_rows)} | key_count={len(cycle_index_local)} | sheet_name={sheet_name}",
        preferred_dir=preferred_dir,
    )
    return CycleTimeData(rows=list(ct_rows), index=cycle_index_local, sheet_name=sheet_name)

# 초보자 설명: 0, 빈칸, None처럼 “없음”으로 취급할 값을 판정한다.
def is_zero_mark(v):
    if v is None:
        return True
    s = norm_text(v)
    if s == "":
        return True
    try:
        return float(s) == 0.0
    except:
        return s == "0"


# 초보자 설명: 값을 가능한 한 안전하게 숫자로 변환한다.
def to_num(v, default=0.0):
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return default
        return x
    except:
        return default


# 초보자 설명: Length와 피팅 조합을 보고 스풀 타입을 판정한다.
def decide_spool_type(length1, length2, fit1, fit2):
    l1_zero = is_zero_mark(length1)
    l2_zero = is_zero_mark(length2)
    f1_zero = is_zero_mark(fit1)
    f2_zero = is_zero_mark(fit2)
    l1_val = to_num(length1, 0.0)
    if (not l1_zero) and l2_zero and (not f1_zero) and (not f2_zero):
        if 300 <= l1_val < 500:
            return "F2DS"
        if 500 <= l1_val < 600:
            return "F2DL"
    if (not l1_zero) and l2_zero and (not f1_zero) and f2_zero:
        return "2D" if l1_val >= 600 else "Short"
    if (not l1_zero) and l2_zero and (not f1_zero) and (not f2_zero):
        return "F2D"
    if (not l1_zero) and (not l2_zero) and (not f1_zero) and f2_zero:
        return "H3D"
    if l1_zero and l2_zero and (not f1_zero) and (not f2_zero):
        return "FF"
    if (not l1_zero) and (not l2_zero) and (not f1_zero) and (not f2_zero):
        return "F3D"
    return "F3D"


# 초보자 설명: 대표 피팅 종류를 하나 뽑아 계획 데이터에 넣는다.
def decide_rep_fitting(fit1, fit2):
    if not is_zero_mark(fit1):
        return norm_text(fit1)
    if not is_zero_mark(fit2):
        return norm_text(fit2)
    return None


# 초보자 설명: 임시 스풀 타입 값을 내부 제품 타입 표기로 넘겨준다.
def decide_internal_prod(tmp_spool_type):
    return norm_text(tmp_spool_type)


# 초보자 설명: 시뮬레이션에서 쓸 대표 길이 값을 결정한다.
def decide_internal_length(tmp_spool_type, length1, length2):
    s = norm_text(tmp_spool_type)
    l1 = to_num(length1, 0.0)
    l2 = to_num(length2, 0.0)
    if s in ["2D", "Short", "F2D", "H3D", "F3D", "F2DS", "F2DL", "2DS", "2DL"]:
        return max(l1, l2)
    elif s == "FF":
        return 0.0
    return max(l1, l2)


# 초보자 설명: Product 파일에 이미 적힌 타입 힌트를 내부 표준 이름으로 바꾼다.
def normalize_product_type_hint(v):
    s = norm_text(v)
    if s in ["F2DS", "F2DS", "FSHORT", "FSSHORT"]:
        return "F2DS"
    if s in ["F2DL", "F2DL", "FLSHORT", "FLSHRT"]:
        return "F2DL"
    return s


# 초보자 설명: 피팅 종류 문자열을 실제 공정 이름으로 바꾼다.
def get_fit_proc_from_ft(ft):
    ft_str = norm_text(ft)
    if ft_str == "45el":
        return "45el"
    elif ft_str == "90el":
        return "90el"
    elif ft_str == "fl":
        return "fl"
    elif ft_str == "Tee":
        return "Tee"
    elif ft_str == "ReTee":
        return "ReTee"
    elif ft_str in ["2DS", "sShBevel"]:
        return "sShBevel"
    return "FitMov"


# 초보자 설명: 다중 피팅 스풀의 1차/2차 피팅 공정을 각각 구한다.
def get_two_fit_procs_for_spool(spool_name, fallback_ft=None):
    fit1_ft, fit2_ft = spool_fit_detail.get(spool_name, ("", ""))
    fit1_ft = norm_text(fit1_ft)
    fit2_ft = norm_text(fit2_ft)
    fallback_ft = norm_text(fallback_ft)
    fit1_proc = get_fit_proc_from_ft(fit1_ft) if fit1_ft != "" else get_fit_proc_from_ft(fallback_ft)
    fit2_proc = get_fit_proc_from_ft(fit2_ft) if fit2_ft != "" else fit1_proc
    return fit1_proc, fit2_proc


# 초보자 설명: 인치 값을 정렬용 숫자로 변환한다.
def _inch_to_float(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return float("-inf")
    try:
        return float(str(v).strip())
    except:
        return float("-inf")


# 초보자 설명: 계획 행의 번호를 1번부터 다시 매긴다.
def renumber_plan_rows(rows, keep_forced=True):
    renum = []
    for idx, r in enumerate(rows, start=1):
        rr = list(r)
        rr[0] = idx
        if keep_forced and len(rr) < 9:
            rr.append(None)
        renum.append(tuple(rr))
    return renum


# 초보자 설명: 임시 계획 행의 번호를 다시 정리한다.
def renumber_tmp_rows(rows):
    return [tuple([idx] + list(r)[1:]) for idx, r in enumerate(rows, start=1)]



# 초보자 설명: 주어진 Job 조건에서 CycleTime 상 가능한 2D 용접장을 찾는다.
def get_valid_2d_welds_for_job_spec(inch, sch, length):
    valid = []
    for weld in ["2D-1", "2D-2", "3D"]:
        key = make_cycle_key(inch, sch, "2D", weld, infer_length_bucket("2D", weld, length))
        if cycle_index.get(key):
            valid.append(weld)
            continue
        fallback_key = make_cycle_key(inch, sch, "2D", weld, "")
        if cycle_index.get(fallback_key):
            valid.append(weld)
    return valid


# 초보자 설명: 임시 행 한 줄을 기준으로 가능한 2D 용접장을 찾는다.
def get_valid_2d_welds_for_tmp_row(tmp_row):
    converted_row = tuple(auto_convert(cell) for cell in tmp_row)
    internal_prod = decide_internal_prod(converted_row[4])
    if internal_prod != "2D":
        return []
    length = decide_internal_length(internal_prod, converted_row[6], converted_row[7])
    return get_valid_2d_welds_for_job_spec(converted_row[2], converted_row[3], length)


# 초보자 설명: 가능 후보 중 우선순위에 맞는 2D 용접장을 선택한다.
def choose_valid_2d_weld_for_tmp_row(tmp_row, preferred=None):
    valid = get_valid_2d_welds_for_tmp_row(tmp_row)
    if not valid:
        return preferred or "2D-1"

    preferred_order = []
    if preferred:
        preferred_order.append(preferred)
    for cand in ["2D-1", "2D-2", "3D"]:
        if cand not in preferred_order:
            preferred_order.append(cand)

    for cand in preferred_order:
        if cand in valid:
            return cand
    return valid[0]


# 초보자 설명: 규칙 기반 모드의 2D 용접장 배정 규칙을 적용한다.
def assign_srtplan_weld(rows):
    """규칙 기반 2D 용접장 할당.

    CycleTime 존재 여부로 규칙을 바꾸지 않는다.
    첫 2D는 3D, 이후 2D는 2D-1 / 2D-2를 전역 순서로 번갈아 배정한다.
    실제 CycleTime 매칭 실패 여부는 이후 시뮬레이션 단계에서 명확한 오류로 처리한다.
    """
    out = []
    first_done = False
    alt_idx = 0
    for row in rows:
        rr = list(row)
        desired_weld = "3D" if not first_done else ("2D-1" if alt_idx % 2 == 0 else "2D-2")
        rr[5] = desired_weld
        first_done = True
        if desired_weld != "3D":
            alt_idx += 1
        out.append(tuple(rr))
    return out


# 초보자 설명: 같은 Spool No.가 여러 번 있으면 마지막 행 기준으로 중복을 정리한다.
def dedup_product_rows_by_spool(product_rows):
    last_row_by_spool = {}
    spool_order = []
    for row in product_rows:
        converted_row = tuple(auto_convert(cell) for cell in row)
        if len(converted_row) < 2:
            continue
        spool_no = norm_text(converted_row[1])
        if spool_no == "":
            continue
        if spool_no not in last_row_by_spool:
            spool_order.append(spool_no)
        last_row_by_spool[spool_no] = converted_row
    return [last_row_by_spool[spool_no] for spool_no in spool_order]


# =====================================================
# Product -> 계획 데이터 변환
# -----------------------------------------------------
# Product.xlsx의 원본 행을 읽어 스풀 타입을 판정하고,
# 규칙 기반 생산계획(srtPlan, plan1, plan2) 형태로 바꾸는 구간이다.
# =====================================================

# 초보자 설명: Product 원본을 임시 계획용 표 형태로 바꾼다.
def build_tmp_product_rows(product_rows):
    tmp_headers = ["번호", "Spool No.", "Inch", "Sch.", "스풀타입", "용접장", "Length1", "Length2", "피팅1", "피팅2"]
    tmp_rows = []
    spool_fit_detail_local = {}
    dedup_rows = dedup_product_rows_by_spool(product_rows)
    for new_no, converted_row in enumerate(dedup_rows, start=1):
        if len(converted_row) < 8:
            continue
        no = new_no
        spool = converted_row[1]
        inch = converted_row[2]
        sch = converted_row[3]
        length1 = converted_row[4]
        length2 = converted_row[5]
        fit1 = converted_row[6]
        fit2 = converted_row[7]
        l1_val_for_short = to_num(length1, 0.0)
        product_type_hint = normalize_product_type_hint(converted_row[10]) if len(converted_row) >= 11 else None
        l2_zero = is_zero_mark(length2)
        f1_zero = is_zero_mark(fit1)
        f2_zero = is_zero_mark(fit2)
        if product_type_hint == "F2DS":
            spool_type = "F2DS"
        elif product_type_hint == "F2DL":
            spool_type = "F2DL"
        elif l2_zero and (not f1_zero) and (not f2_zero) and 300 <= l1_val_for_short < 500:
            spool_type = "F2DS"
        elif l2_zero and (not f1_zero) and (not f2_zero) and 500 <= l1_val_for_short < 600:
            spool_type = "F2DL"
        elif 300 <= l1_val_for_short < 500:
            spool_type = "2DS"
        elif 500 <= l1_val_for_short < 600:
            spool_type = "2DL"
        else:
            spool_type = decide_spool_type(length1, length2, fit1, fit2)
        tmp_rows.append((
            no, spool, inch, sch, spool_type, None,
            auto_convert(length1), auto_convert(length2),
            norm_text(fit1) if not is_zero_mark(fit1) else 0,
            norm_text(fit2) if not is_zero_mark(fit2) else 0,
        ))
        spool_fit_detail_local[spool] = (
            norm_text(fit1) if not is_zero_mark(fit1) else "",
            norm_text(fit2) if not is_zero_mark(fit2) else "",
        )
    return tmp_headers, tmp_rows, spool_fit_detail_local


# 초보자 설명: 임시 계획 데이터를 실제 시뮬레이션용 계획 형식으로 변환한다.
def convert_tmp_to_plan_rows(tmp_rows):
    plan_rows = []
    for row in tmp_rows:
        converted_row = tuple(auto_convert(cell) for cell in row)
        if len(converted_row) < 10:
            continue
        internal_prod = decide_internal_prod(converted_row[4])
        rep_fit = decide_rep_fitting(converted_row[8], converted_row[9])
        plan_weld = converted_row[5]
        if norm_text(internal_prod) in SHORT_PRODUCT_TYPES:
            plan_weld = "Short"
        if norm_text(internal_prod) in LONG_3D_PRODUCT_TYPES and norm_text(plan_weld) == "":
            plan_weld = "3D"
        length = decide_internal_length(internal_prod, converted_row[6], converted_row[7])
        plan_rows.append(
            (0, converted_row[1], converted_row[2], converted_row[3], internal_prod, plan_weld, length, rep_fit, None))
    return plan_rows


# 초보자 설명: 긴 제품군과 Short 제품군을 분리해 1차 계획/2차 계획으로 나눈다.
def split_plan1_plan2(sorted_tmp_rows):
    plan1, plan2 = [], []
    for row in sorted_tmp_rows:
        rr = list(row)
        prod = norm_text(rr[4])
        if prod in SHORT_PRODUCT_TYPES:
            rr[5] = "Short"
            plan2.append(tuple(rr))
        else:
            if prod in LONG_3D_PRODUCT_TYPES:
                rr[5] = "3D"
            plan1.append(tuple(rr))
    return renumber_tmp_rows(plan1), renumber_tmp_rows(plan2)


# 초보자 설명: plan1에서 2D 대상 행을 뽑아 srtPlan의 기본 골격을 만든다.
def build_srtplan_from_plan1(plan1_rows):
    srt_2d = [tuple(row) for row in plan1_rows if norm_text(row[4]) == "2D"]
    srt_2d = assign_srtplan_weld(srt_2d)
    srt_2d = renumber_tmp_rows(srt_2d)
    plan1_keep = renumber_tmp_rows([tuple(r) for r in plan1_rows])
    return srt_2d, plan1_keep


# 초보자 설명: 보조 삽입용 Job을 인치와 타입별로 묶는다.
def _group_aux_rows_by_inch_and_type(rows):
    grouped = defaultdict(lambda: {tp: deque() for tp in AUX_INSERT_PRODUCT_ORDER})
    for row in rows:
        prod = norm_text(row[4])
        if prod not in AUX_INSERT_PRODUCT_ORDER:
            continue
        inch = _inch_to_float(row[2])
        grouped[inch][prod].append(tuple(row))
    return grouped


# 초보자 설명: 같은 인치 그룹 안에서 F2D/F3D/H3D 순환 삽입 순서를 만든다.
def _build_aux_sequence_for_inch(type_map):
    order = list(AUX_INSERT_PRODUCT_ORDER)
    out = []
    while True:
        pushed = False
        for tp in order:
            if type_map[tp]:
                out.append(type_map[tp].popleft())
                pushed = True
        if not pushed:
            break
    return deque(out)


# 초보자 설명: 2D 사이사이에 보조 Job을 규칙대로 끼워 넣는다.
def insert_aux_jobs_into_srtplan(srt_2d_rows, plan1_rows):
    only_2d = [tuple(r) for r in srt_2d_rows]
    grouped_aux = _group_aux_rows_by_inch_and_type(plan1_rows)
    aux_by_inch = {}
    for inch, type_map in grouped_aux.items():
        copied = {"F2D": deque(type_map["F2D"]), "F3D": deque(type_map["F3D"]), "H3D": deque(type_map["H3D"])}
        aux_by_inch[inch] = _build_aux_sequence_for_inch(copied)
    result = []
    used_inch = set()
    if not only_2d:
        leftover = []
        for inch in sorted(aux_by_inch.keys(), reverse=True):
            while aux_by_inch[inch]:
                leftover.append(aux_by_inch[inch].popleft())
        return renumber_tmp_rows(leftover), renumber_tmp_rows([tuple(r) for r in plan1_rows])
    for i, row_2d in enumerate(only_2d):
        result.append(tuple(row_2d))
        cur_inch = _inch_to_float(row_2d[2])
        used_inch.add(cur_inch)
        next_row = only_2d[i + 1] if i + 1 < len(only_2d) else None
        cur_weld = norm_text(row_2d[5])
        if next_row is not None:
            next_inch = _inch_to_float(next_row[2])
            next_weld = norm_text(next_row[5])
            if cur_inch == next_inch and cur_weld == "2D-2" and next_weld == "2D-1":
                if cur_inch in aux_by_inch and aux_by_inch[cur_inch]:
                    result.append(aux_by_inch[cur_inch].popleft())
        if next_row is None or _inch_to_float(next_row[2]) != cur_inch:
            if cur_inch in aux_by_inch:
                while aux_by_inch[cur_inch]:
                    result.append(aux_by_inch[cur_inch].popleft())
    leftover_inches = sorted([inch for inch in aux_by_inch.keys() if inch not in used_inch], reverse=True)
    for inch in leftover_inches:
        while aux_by_inch[inch]:
            result.append(aux_by_inch[inch].popleft())
    return renumber_tmp_rows(result), renumber_tmp_rows([tuple(r) for r in plan1_rows])


# 초보자 설명: 디버깅용 임시 계획 시트 구조를 메모리용 딕셔너리로 만든다.
def save_tmp_plan_xlsx(srtPlan_rows, plan1_rows, plan2_rows):
    tmp_headers = ["번호", "Spool No.", "Inch", "Sch.", "스풀타입", "용접장", "Length1", "Length2", "피팅1", "피팅2"]
    return {
        "srtPlan": (tmp_headers, list(srtPlan_rows)),
        "plan1": (tmp_headers, list(plan1_rows)),
        "plan2": (tmp_headers, list(plan2_rows)),
    }


# 초보자 설명: 2D가 아닌 제품군은 강제 용접장 규칙을 다시 맞춘다.
def force_weld_for_non_2d(rows):
    out = []
    for row in rows:
        rr = list(row)
        prod = norm_text(rr[4])
        if prod in LONG_3D_PRODUCT_TYPES:
            rr[5] = "3D"
        elif prod in SHORT_PRODUCT_TYPES:
            rr[5] = "Short"
        out.append(tuple(rr))
    return renumber_tmp_rows(out)


# 초보자 설명: Product 데이터를 받아 규칙 기반 계획 전체를 단계별로 만들어 낸다.
def build_plan_pipeline(product_rows):
    global spool_fit_detail
    tmp_headers, tmp_rows, spool_fit_detail_local = build_tmp_product_rows(product_rows)
    spool_fit_detail = spool_fit_detail_local
    # tmpProduct.xlsx는 디버깅용 임시 산출물로 더 이상 파일 저장하지 않고 변수로만 유지
    tmp_rows_sorted = sorted(tmp_rows, key=lambda r: _inch_to_float(r[2]), reverse=True)
    plan1_raw, plan2_raw = split_plan1_plan2(tmp_rows_sorted)
    srt_2d, plan1_keep = build_srtplan_from_plan1(plan1_raw)
    srtPlan_rows, _unused_plan1 = insert_aux_jobs_into_srtplan(srt_2d, plan1_keep)
    srtPlan_rows = force_weld_for_non_2d(srtPlan_rows)
    plan1_keep = force_weld_for_non_2d(plan1_keep)
    plan2_raw = force_weld_for_non_2d(plan2_raw)
    _tmp_plan_book = save_tmp_plan_xlsx(srtPlan_rows, plan1_keep, plan2_raw)
    srtPlan_sim_rows = renumber_plan_rows(convert_tmp_to_plan_rows(srtPlan_rows))
    plan1_sim_rows = renumber_plan_rows(convert_tmp_to_plan_rows(plan1_keep))
    plan2_sim_rows = renumber_plan_rows(convert_tmp_to_plan_rows(plan2_raw))
    return srtPlan_rows, plan1_keep, plan2_raw, srtPlan_sim_rows, plan1_sim_rows, plan2_sim_rows


# 초보자 설명: Product 파일에서 파생되는 여러 중간 결과를 한 객체로 묶는다.
def prepare_product_plan_data(product_rows):
    tmp_headers, tmp_rows, spool_fit_detail_local = build_tmp_product_rows(product_rows)
    (
        srt_plan_rows,
        plan1_rows_local,
        plan2_rows_local,
        srt_plan_sim_rows,
        plan1_sim_rows_local,
        plan2_sim_rows_local,
    ) = build_plan_pipeline(product_rows)
    return ProductPlanData(
        product_rows=list(product_rows),
        tmp_product_headers=list(tmp_headers),
        tmp_product_rows=list(tmp_rows),
        spool_fit_detail=dict(spool_fit_detail_local),
        srt_plan_rows=list(srt_plan_rows),
        plan1_rows=list(plan1_rows_local),
        plan2_rows=list(plan2_rows_local),
        srt_plan_sim_rows=list(srt_plan_sim_rows),
        plan1_sim_rows=list(plan1_sim_rows_local),
        plan2_sim_rows=list(plan2_sim_rows_local),
    )


# 초보자 설명: Product.xlsx를 읽어 계획용 데이터 세트를 만든다.
def load_product_plan_data(path):
    product_rows, workbook, sheet_name = open_xlsx_reader_auto(path, skip_header=True)
    workbook.close()
    print(f"[INFO] Product.xlsx sheet = {sheet_name}")
    return prepare_product_plan_data(product_rows)


# 초보자 설명: 읽어 둔 CycleTime 데이터를 전역 실행 상태에 반영한다.
def apply_cycle_time_data(cycle_data: CycleTimeData) -> None:
    global CT, cycle_index, CYCLE_DICT_CACHE
    CT = list(cycle_data.rows)
    cycle_index = dict(cycle_data.index)
    CYCLE_DICT_CACHE.clear()


# 초보자 설명: 읽어 둔 Product/Plan 데이터를 전역 실행 상태에 반영한다.
def apply_product_plan_data(plan_data: ProductPlanData) -> None:
    global spool_fit_detail, spool_length_pair, CYCLE_DICT_CACHE
    global srtPlan_rows, plan1_rows, plan2_rows
    global srtPlan_sim_rows, plan1_sim_rows, plan2_sim_rows
    global plan, product_input_rows, tmp_product_headers, tmp_product_rows

    product_input_rows = list(plan_data.product_rows)
    tmp_product_headers = list(plan_data.tmp_product_headers)
    tmp_product_rows = list(plan_data.tmp_product_rows)
    spool_fit_detail = dict(plan_data.spool_fit_detail)
    spool_length_pair = {
        norm_text(row[1]): (to_num(row[6], 0.0), to_num(row[7], 0.0))
        for row in tmp_product_rows
        if len(row) >= 8
    }
    srtPlan_rows = list(plan_data.srt_plan_rows)
    plan1_rows = list(plan_data.plan1_rows)
    plan2_rows = list(plan_data.plan2_rows)
    srtPlan_sim_rows = list(plan_data.srt_plan_sim_rows)
    plan1_sim_rows = list(plan_data.plan1_sim_rows)
    plan2_sim_rows = list(plan_data.plan2_sim_rows)
    plan = list(srtPlan_sim_rows)
    CYCLE_DICT_CACHE.clear()


# =====================================================
# CycleTime 매칭 / 문맥 사전(cycle_dict) 만들기
# -----------------------------------------------------
# 같은 제품이라도 용접장과 길이 구간이 다르면 시간이 달라지므로,
# 현재 Job 문맥에 맞는 CycleTime 행을 정확히 찾는 과정이 필요하다.
# =====================================================

# 초보자 설명: 계획 행 한 줄이 CycleTime의 몇 번째 행과 맞는지 찾는다.
def find_match_row_by_plan(plan_rows, pNum):
    if pNum < 1 or pNum > len(plan_rows):
        return None
    plan_row = plan_rows[pNum - 1]
    if len(plan_row) < 7:
        return None
    prod_val = normalize_product_type_name(plan_row[4])
    weld_val = norm_text(plan_row[5])
    if weld_val == "":
        if prod_val == "2D":
            weld_val = "2D-1"
        elif prod_val in ["F2D", "H3D", "F3D"]:
            weld_val = "3D"
        elif prod_val in ["2DS", "2DL", "Short", "FF", "F2DS", "F2DL", "ff"]:
            weld_val = "Short"
    length_bucket = infer_length_bucket(prod_val, weld_val, plan_row[6])
    key = make_cycle_key(plan_row[2], plan_row[3], prod_val, weld_val, length_bucket)
    rows = cycle_index.get(key, None)
    if rows:
        return rows[0]
    fallback_key = make_cycle_key(plan_row[2], plan_row[3], prod_val, weld_val, "")
    rows = cycle_index.get(fallback_key, None)
    return rows[0] if rows else None


# 초보자 설명: 현재 전역 plan 기준으로 CycleTime 매칭 행을 찾는다.
def find_match_row(pNum):
    return find_match_row_by_plan(plan, pNum)


# 초보자 설명: CycleTime 행에서 원본 값을 그대로 꺼낸다.
def _ct_raw_value(ct_row, col_name):
    idx = CT_COL[col_name]
    if idx >= len(ct_row):
        return None
    return ct_row[idx]


# 초보자 설명: CycleTime 행에서 숫자 값을 안전하게 꺼낸다.
def _ct_value(ct_row, col_name, default=None):
    v = _ct_raw_value(ct_row, col_name)
    if v is None:
        return default
    s = norm_text(v)
    if s == "" or s.upper() == "N/A":
        return default
    try:
        x = float(v)
        if math.isnan(x) or math.isinf(x):
            return default
        return float(x)
    except Exception:
        return default


# 초보자 설명: 스풀 이름으로 Length1/Length2 값을 다시 찾는다.
def get_spool_length_pair(spool_name):
    spool_text = norm_text(spool_name)
    pair = spool_length_pair.get(spool_text)
    if pair is not None:
        return pair
    for row in tmp_product_rows:
        if len(row) >= 8 and norm_text(row[1]) == spool_text:
            pair = (to_num(row[6], 0.0), to_num(row[7], 0.0))
            spool_length_pair[spool_text] = pair
            return pair
    pair = (0.0, 0.0)
    spool_length_pair[spool_text] = pair
    return pair


# 초보자 설명: 어느 용접 단계의 길이를 써야 하는지 계산한다.
def _stage_length_for_weld(prod, target_weld, len1, len2):
    prod = normalize_product_type_name(prod)
    target_weld = norm_text(target_weld)
    if prod in ["2DS", "2DL", "F2DS", "F2DL", "Short"]:
        return len1
    if prod == "FF":
        return None
    if prod == "2D":
        return len1
    if prod == "F2D":
        if target_weld == "2D-1":
            return len1
        return None
    if prod == "H3D":
        if target_weld == "2D-2":
            return len1
        if target_weld == "3D":
            return len2
    if prod == "F3D":
        if target_weld == "2D-1":
            return len1
        if target_weld == "2D-2":
            return len2
        return None
    return len1


# 초보자 설명: 현재 제품/용접 단계 문맥에 맞는 CycleTime 행을 다시 찾는다.
def _find_cycle_row_for_context(plan_row, target_weld, stage_length):
    prod = normalize_product_type_name(plan_row[4])
    key = make_cycle_key(plan_row[2], plan_row[3], prod, target_weld,
                         infer_length_bucket(prod, target_weld, stage_length) if stage_length is not None else "")
    rows = cycle_index.get(key, None)
    if rows:
        return CT[rows[0] - 1]
    key2 = make_cycle_key(plan_row[2], plan_row[3], prod, target_weld, "")
    rows = cycle_index.get(key2, None)
    if rows:
        return CT[rows[0] - 1]
    return None


# 초보자 설명: 같은 계획 행을 다시 계산하지 않도록 캐시용 키를 만든다.
def _make_cycle_dict_cache_key(plan_row):
    if plan_row is None:
        return None
    spool_name = plan_row[1] if len(plan_row) > 1 else ""
    return (
        norm_text(spool_name),
        norm_num_text(plan_row[2]) if len(plan_row) > 2 else "",
        norm_text(plan_row[3]) if len(plan_row) > 3 else "",
        normalize_product_type_name(plan_row[4]) if len(plan_row) > 4 else "",
        norm_text(plan_row[5]) if len(plan_row) > 5 else "",
        to_num(plan_row[6], 0.0) if len(plan_row) > 6 else 0.0,
        norm_text(plan_row[7]) if len(plan_row) > 7 else "",
    )


# 초보자 설명: 한 Job이 공정 시간을 찾을 때 필요한 문맥 정보를 사전 형태로 만든다.
def build_cycle_dict(ct_row, plan_row=None):
    cache_key = _make_cycle_dict_cache_key(plan_row)
    if cache_key is not None:
        cached = CYCLE_DICT_CACHE.get(cache_key)
        if cached is not None:
            return cached

    if plan_row is None:
        prod = normalize_product_type_name(_ct_raw_value(ct_row, "prod"))
        weld = norm_text(_ct_raw_value(ct_row, "weld"))
        length = 0.0
        rows_by_stage = {weld: ct_row}
        spool_name = ""
        len1 = len2 = 0.0
    else:
        prod = normalize_product_type_name(plan_row[4])
        weld = norm_text(plan_row[5])
        spool_name = plan_row[1] if len(plan_row) > 1 else ""
        len1, len2 = get_spool_length_pair(spool_name)
        length = to_num(plan_row[6], 0.0)
        stage_welds = ["2D-1", "2D-2", "3D", "Short"]
        rows_by_stage = {}
        for stage in stage_welds:
            stage_len = _stage_length_for_weld(prod, stage, len1, len2)
            rows_by_stage[stage] = _find_cycle_row_for_context(plan_row, stage, stage_len)

    cycle_dict = {
        "_ct_row": ct_row,
        "_rows": rows_by_stage,
        "_prod": prod,
        "_weld": weld,
        "_length": length,
        "_length1": len1 if plan_row is not None else length,
        "_length2": len2 if plan_row is not None else 0.0,
        "_inch": plan_row[2] if plan_row is not None and len(plan_row) > 2 else _ct_raw_value(ct_row, "inch"),
        "_sch": plan_row[3] if plan_row is not None and len(plan_row) > 3 else _ct_raw_value(ct_row, "sch"),
        "_spool": spool_name,
        "_fit": norm_text(plan_row[7]) if plan_row is not None and len(plan_row) > 7 else "",
    }
    if cache_key is not None:
        CYCLE_DICT_CACHE[cache_key] = cycle_dict
    return cycle_dict


# 초보자 설명: 우선 단계와 대체 단계를 고려해 사용할 CycleTime 행을 고른다.
def _ctx_row(cycle_dict, preferred_stage, fallback_stages=()):
    rows = cycle_dict.get("_rows", {}) or {}
    if preferred_stage in rows and rows.get(preferred_stage) is not None:
        return rows.get(preferred_stage)
    for st in fallback_stages:
        if st in rows and rows.get(st) is not None:
            return rows.get(st)
    # 마지막 fallback: 원래 매칭 행
    return cycle_dict.get("_ct_row")


# 초보자 설명: 현재 문맥에 맞는 CycleTime 값을 한 칸 꺼낸다.
def _ctx_value(cycle_dict, col_name, preferred_stage, fallback_stages=(), default=None):
    row = _ctx_row(cycle_dict, preferred_stage, fallback_stages)
    if row is None:
        return default
    return _ct_value(row, col_name, default)


# 초보자 설명: Short 준비 공정에서 대체 타입의 CycleTime 행을 찾는다.
def _find_alt_short_prep_row(cycle_dict, alt_prod):
    inch = cycle_dict.get("_inch")
    sch = cycle_dict.get("_sch")
    length = cycle_dict.get("_length")
    if inch is None or sch is None:
        return None
    key = make_cycle_key(inch, sch, alt_prod, "Short", infer_length_bucket(alt_prod, "Short", length))
    rows = cycle_index.get(key, None)
    if rows:
        return CT[rows[0] - 1]
    key2 = make_cycle_key(inch, sch, alt_prod, "Short", "")
    rows = cycle_index.get(key2, None)
    if rows:
        return CT[rows[0] - 1]
    return None


# 초보자 설명: Short 준비 공정용 시간을 우선/대체 규칙으로 찾는다.
def _ctx_short_prep_value(cycle_dict, col_names, default=None):
    if isinstance(col_names, str):
        names = [col_names]
    else:
        names = list(col_names)

    for name in names:
        val = _ctx_value(cycle_dict, name, "Short", ("Short",))
        if val is not None:
            return val

    prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
    alt_prod = {"2DS": "F2DS", "2DL": "F2DL"}.get(prod)
    if alt_prod:
        alt_row = _find_alt_short_prep_row(cycle_dict, alt_prod)
        if alt_row is not None:
            for name in names:
                val = _ct_value(alt_row, name, None)
                if val is not None:
                    return val

    return default


# =====================================================
# 시뮬레이션 전역 상태와 표시 설정
# -----------------------------------------------------
# SimPy 자원, 간트 차트 색상, 라인 상태, FIFO 토큰 같은
# 실행 중 계속 참조하는 공통 변수들을 여기서 정의한다.
# =====================================================

DEFAULT_CYCLE = {k: 0 for k in [

    "Cut", "Con1", "Conv1", "Sen1", "Sen1_2", "Bevel1", "Bev1", "Bevel2", "Bev2", "Sen2", "Sen2_2", "Con2", "Conv2",
    "Conv3",
    "Gan", "Gan1", "Gan2", "Gan3", "R1", "R2", "R3", "R14",
    "R_WELD", "R2tig", "R3tig", "R14tig", "Fit", "Angle", "FitMov", "tig", "Tig2", "Mag", "2D-1", "2D-2", "Short", "3D",
    "45el", "90el", "fl", "Tee", "ReTee", "sShBevel"
]}

COLOR = {

    "Cut": "#005BBB", "Con1": "#00A6A6", "Conv1": "#00A6A6", "Con2": "#2ECC40", "Conv2": "#2ECC40", "Conv3": "#20B2AA",
    "Sen1": "#7FDBFF", "Sen2": "#F7A8FF",
    "Bevel1": "#FF6A00", "Bev1": "#FF6A00", "Bevel2": "#C6E300", "Bev2": "#C6E300",
    "Gan": "#6A00FF",
    "R1": "#7A3E00", "R2": "#C68600", "R3": "#808000", "R14": "#C0C0C0",
    "R2tig": "#FF8C00", "R3tig": "#6B8E23", "R14tig": "#8FBC8F",
    "Pick1": "#708090", "Pick2": "#778899",
    "ShrtBev1": "#FFD400", "ShrtBev2": "#E6C200",
    "Fit": "#FF1493", "Angle": "#C71585", "FitMov": "#FF1493", "Tig2": "#FF5F1F", "Mag": "#8B4513",
    "2D-1": "#FF00B3", "2D-2": "#003CFF", "3D": "#5A5A5A", "Short": "#000000",
    "45el": "#FFB000", "90el": "#00D5FF", "fl": "#7CFC00", "Tee": "#8A2BE2", "ReTee": "#FF5A8A", "sShBevel": "#FFD400",
}

env = None
resources = {}
CutBuf = None
gantt = []
TRANSPORT_PROCS = {"R1", "R2", "R3", "R14", "Con1", "Con2", "Conv3", "Sen1", "Sen2", "Gan", "FitMov"}
WORKSTATION_CAPACITY = {"Cut": 6, "2D-1": 1, "2D-2": 1, "3D": 1, "Short": 1, "Bevel1": 1, "Bevel2": 1, "sShBevel": 1,
                        "45el": 1, "90el": 1, "Tee": 1, "ReTee": 1, "fl": 1, "FIT_ELBOW_TEE": 1, "FIT_FL": 1}
MAIN_BEVEL_GROUP_RESOURCE = {1: "MainBevelM1", 2: "MainBevelM2"}
# NOTE:
# FIT_ELBOW_TEE / FIT_FL 는 Pipe와 Sub가 같은 자원을 공유하면서도
# 한 Job 안에서 여러 번 반복 사용될 수 있다.
# 여기에 별도의 station token FIFO를 강제로 적용하면
# GA가 만든 자유로운 Job 순서에서 token 개수/실행 순서가 어긋나면서
# 일부 Job이 영원히 대기하는 현상이 생길 수 있다.
#
# 실제 capacity=1 은 SimPy Resource 자체가 이미 보장하므로
# fitting 공용 설비는 Resource FIFO에 맡기고,
# station token FIFO 는 Cut / Bevel / 최종 용접장에만 적용한다.
STRICT_STATION_FIFO = {"Cut", "Bevel1", "Bevel2", "2D-1", "2D-2", "3D"}
line_step_done = {}
station_job_tokens = defaultdict(list)
station_done_events = {}
station_token_index = {}
station_runtime_occ_counter = defaultdict(int)
FORCED_CUT_STARTS = {}
idleTimeR2 = []
pending_weld_holds = defaultdict(deque)
pending_continuous_short_blocks = {}
line_active_holds = {}
short_station_busy = False
short_station_pending = []
short_station_dispatch_scheduled = False
pending_main_weld_arrival_counts = defaultdict(int)
waiting_main_weld_arrivals = defaultdict(deque)
sub_start_job_tokens = defaultdict(list)
sub_start_done_events = {}
sub_start_token_index = {}
COMBINED_WELD_HOLD_ENABLED = True
PARALLEL_SUB_MULTI_PRODUCTS_ENABLED = True


# 초보자 설명: 간트 차트 작성을 위해 공정 시작/종료 기록을 한 줄씩 남긴다.
def record(job, line, proc, start, end):
    proc = DISPLAY_PROC_ALIASES.get(proc, proc)
    gantt.append((job, line, proc, start, end))


MAX_PIPE_STEPS = 60
MAX_SUB_STEPS = 60
line_prev_job = {}
SHORT_S_FAMILY_PRODS = {"2DS", "F2DS"}
SHORT_WELD_JOB_PRODS = set(SHORT_S_FAMILY_PRODS)  # Cut->FitMov->sShBevel 계열의 main weld 직렬화 전용
short_s_prev_job = {}
short_s_first_short_done = {}
short_s_prep_prev_job = {}
short_s_prep_done = {}
short_s_weld_prev_job = {}
short_s_weld_done = {}


# 초보자 설명: Sub 라인의 시작 공정을 계획 행 기준으로 구한다.
def get_sub_product_start_proc_from_plan_row(row, product_no=1):
    """Sub flow의 각 제품(1번/2번)의 시작 투입 공정을 구한다.

    - product_no=1 : 첫 번째 sub 제품
    - product_no=2 : 두 번째 sub 제품(존재하지 않으면 None)
    """
    try:
        spool_name = row[1] if len(row) > 1 else None
        prod = norm_text(row[4]) if len(row) > 4 else ""
        fallback_ft = row[7] if len(row) > 7 else None
    except Exception:
        return "FitMov" if product_no == 1 else None

    if prod in MULTI_SUB_PRODUCT_TYPES:
        fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)
        if product_no == 1:
            return norm_text(fit1_proc) or "FitMov"
        if product_no == 2:
            return norm_text(fit2_proc) or norm_text(fit1_proc) or "FitMov"
        return None

    if product_no == 1:
        return norm_text(get_fit_proc_from_ft(fallback_ft)) or "FitMov"
    return None


# 초보자 설명: 첫 번째 Sub 라인의 시작 공정을 구한다.
def get_sub_start_proc_from_plan_row(row):
    return get_sub_product_start_proc_from_plan_row(row, product_no=1)


# 초보자 설명: 두 번째 Sub 라인의 시작 공정을 구한다.
def get_sub2_start_proc_from_plan_row(row):
    return get_sub_product_start_proc_from_plan_row(row, product_no=2)


# 초보자 설명: 제품 타입에 따라 동기화에 사용할 실제 라인 이름을 정한다.
def resolve_line_sync_name(line_name, job_no):
    if line_name == "Pipe":
        try:
            row = plan[job_no - 1]
            prod = norm_text(row[4]) if len(row) > 4 else ""
        except Exception:
            return line_name
        if prod == "FF":
            return "PipeFF"
        if prod in CUT_BASED_SHORT_PRODUCT_TYPES:
            return "PipeShort"
        return "Pipe"

    if line_name == "Sub":
        try:
            row = plan[job_no - 1]
            first_sub_proc = get_sub_start_proc_from_plan_row(row)
        except Exception:
            return line_name
        return "SubFL" if first_sub_proc == "fl" else "SubCommon"

    if line_name == "Sub2":
        try:
            row = plan[job_no - 1]
            second_sub_proc = get_sub2_start_proc_from_plan_row(row)
        except Exception:
            return line_name
        if not second_sub_proc:
            return "Sub2Common"
        return "Sub2FL" if second_sub_proc == "fl" else "Sub2Common"

    return line_name


# =====================================================
# 시뮬레이션 초기화 / 순서 제어 / 자원 점유 관리
# -----------------------------------------------------
# 이 구간은 '누가 먼저 들어가야 하는가', '어떤 설비를 언제 놓아주는가'를
# 이벤트와 토큰으로 관리하는 핵심 제어 로직이다.
# =====================================================

# 초보자 설명: 시뮬레이션을 시작하기 전에 자원, 이벤트, 기록 상태를 모두 초기화한다.
def reset_runtime_state(current_plan):
    global env, resources, CutBuf, gantt
    global line_step_done, line_prev_job
    global short_s_prev_job, short_s_first_short_done
    global short_s_prep_prev_job, short_s_prep_done, short_s_weld_prev_job, short_s_weld_done
    global station_job_tokens, station_done_events, station_token_index, station_runtime_occ_counter, pending_weld_holds, pending_continuous_short_blocks, line_active_holds
    global short_station_busy, short_station_pending, short_station_dispatch_scheduled
    global pending_main_weld_arrival_counts, waiting_main_weld_arrivals
    global sub_start_job_tokens, sub_start_done_events, sub_start_token_index
    global proc_runtime_call_counter
    env = simpy.Environment()
    resources = {}
    for k in TRANSPORT_PROCS:
        resources[k] = simpy.Resource(env, capacity=1)
    for k, v in WORKSTATION_CAPACITY.items():
        resources[k] = simpy.Resource(env, capacity=v)
    for _grp_no, _grp_res_name in MAIN_BEVEL_GROUP_RESOURCE.items():
        resources[_grp_res_name] = simpy.PriorityResource(env, capacity=1)
    CutBuf = simpy.Container(env, capacity=6, init=0)
    gantt = []
    line_keys = ["Pipe", "PipeShort", "PipeFF", "Pipe2", "SubCommon", "SubFL", "Sub2Common", "Sub2FL"]
    line_step_done = {
        key: {j: [None] + [env.event() for _ in range(MAX_SUB_STEPS if key.startswith("Sub") else MAX_PIPE_STEPS)] for j
              in range(1, len(current_plan) + 1)}
        for key in line_keys
    }
    line_prev_job = {key: {} for key in line_keys}
    last_seen = {key: None for key in line_keys}
    for row in current_plan:
        job_no = row[0]
        prod = norm_text(row[4]) if len(row) > 4 else ""
        if prod == "FF":
            pipe_key = "PipeFF"
        elif prod in CUT_BASED_SHORT_PRODUCT_TYPES:
            pipe_key = "PipeShort"
        else:
            pipe_key = "Pipe"

        first_sub_proc = get_sub_start_proc_from_plan_row(row)
        second_sub_proc = get_sub2_start_proc_from_plan_row(row)
        sub_key = "SubFL" if first_sub_proc == "fl" else "SubCommon"
        sub2_key = None
        if second_sub_proc:
            sub2_key = "Sub2FL" if second_sub_proc == "fl" else "Sub2Common"

        line_prev_job[pipe_key][job_no] = last_seen[pipe_key]
        line_prev_job[sub_key][job_no] = last_seen[sub_key]
        line_prev_job["Pipe2"][job_no] = last_seen["Pipe2"]
        if sub2_key is not None:
            line_prev_job[sub2_key][job_no] = last_seen[sub2_key]
            last_seen[sub2_key] = job_no
        last_seen[pipe_key] = job_no
        last_seen[sub_key] = job_no
        last_seen["Pipe2"] = job_no
    station_job_tokens = defaultdict(list)
    station_done_events = {}
    station_token_index = {}
    station_runtime_occ_counter = defaultdict(int)
    pending_weld_holds = defaultdict(deque)
    pending_continuous_short_blocks = {}
    line_active_holds = {}
    short_station_busy = False
    short_station_pending = []
    short_station_dispatch_scheduled = False
    pending_main_weld_arrival_counts = defaultdict(int)
    waiting_main_weld_arrivals = defaultdict(deque)
    sub_start_job_tokens = defaultdict(list)
    sub_start_done_events = {}
    sub_start_token_index = {}
    reset_proc_runtime_call_counter()

    short_s_prev_job = {}
    short_s_first_short_done = {}
    short_s_prep_prev_job = {}
    short_s_prep_done = {}
    short_s_weld_prev_job = {}
    short_s_weld_done = {}
    last_short_s_job = None
    last_short_s_prep_job = None
    last_short_s_weld_job = None
    for row in current_plan:
        job_no = row[0]
        prod = norm_text(row[4]) if len(row) > 4 else ""
        if prod in SHORT_S_FAMILY_PRODS:
            short_s_prev_job[job_no] = last_short_s_job
            short_s_first_short_done[job_no] = env.event()
            short_s_prep_prev_job[job_no] = last_short_s_prep_job
            short_s_prep_done[job_no] = env.event()
            last_short_s_job = job_no
            last_short_s_prep_job = job_no
        if prod in SHORT_WELD_JOB_PRODS:
            short_s_weld_prev_job[job_no] = last_short_s_weld_job
            short_s_weld_done[job_no] = env.event()
            last_short_s_weld_job = job_no


# 초보자 설명: 같은 라인의 이전 Job이 같은 단계까지 끝날 때까지 기다린다.
def wait_prev_line_step(env, line_name, job_no, step_no):
    sync_line = resolve_line_sync_name(line_name, job_no)
    prev_job = line_prev_job.get(sync_line, {}).get(job_no, None)
    if prev_job is None:
        return
    prev_evt = line_step_done[sync_line][prev_job][step_no]
    if not prev_evt.triggered:
        yield prev_evt


# 초보자 설명: 한 라인 단계가 끝났음을 이벤트로 표시한다.
def mark_line_step_done(line_name, job_no, step_no):
    sync_line = resolve_line_sync_name(line_name, job_no)
    evt = line_step_done[sync_line][job_no][step_no]
    if not evt.triggered:
        evt.succeed()


# 초보자 설명: Short S 계열의 앞선 Job이 첫 Short를 끝낼 때까지 기다린다.
def wait_prev_short_s_first_short(env, job_no):
    prev_job = short_s_prev_job.get(job_no, None)
    if prev_job is None:
        return
    evt = short_s_first_short_done.get(prev_job, None)
    if evt is not None and not evt.triggered:
        yield evt


# 초보자 설명: Short S 계열 첫 Short 완료를 표시한다.
def mark_short_s_first_short_done(job_no):
    evt = short_s_first_short_done.get(job_no, None)
    if evt is not None and not evt.triggered:
        evt.succeed()


# 초보자 설명: Short S 계열의 준비 공정 순서를 맞추기 위해 기다린다.
def wait_prev_short_s_prep(env, job_no):
    prev_job = short_s_prep_prev_job.get(job_no, None)
    if prev_job is None:
        return
    evt = short_s_prep_done.get(prev_job, None)
    if evt is not None and not evt.triggered:
        yield evt


# 초보자 설명: Short S 계열 준비 공정 완료를 표시한다.
def mark_short_s_prep_done(job_no):
    evt = short_s_prep_done.get(job_no, None)
    if evt is not None and not evt.triggered:
        evt.succeed()


# 초보자 설명: Short S 계열 용접 순서를 맞추기 위해 기다린다.
def wait_prev_short_s_weld(env, job_no):
    prev_job = short_s_weld_prev_job.get(job_no, None)
    if prev_job is None:
        return
    evt = short_s_weld_done.get(prev_job, None)
    if evt is not None and not evt.triggered:
        yield evt


# 초보자 설명: Short S 계열 용접 완료를 표시한다.
def mark_short_s_weld_done(job_no):
    evt = short_s_weld_done.get(job_no, None)
    if evt is not None and not evt.triggered:
        evt.succeed()


# 초보자 설명: 한 라인에서 “기다림 + 실제 공정 실행 + 완료 표시”를 한 번에 처리한다.
def run_line_step(env, line_name, job_no, step_no, proc_gen, enforce_prev_step=True):
    if enforce_prev_step:
        yield env.process(wait_prev_line_step(env, line_name, job_no, step_no))
    yield env.process(proc_gen)
    mark_line_step_done(line_name, job_no, step_no)


# 초보자 설명: 더 이상 쓰지 않는 뒷단 단계 이벤트를 모두 종료 처리한다.
def release_remaining_line_steps(line_name, job_no, last_step):
    sync_line = resolve_line_sync_name(line_name, job_no)
    max_steps = MAX_SUB_STEPS if sync_line.startswith("Sub") else MAX_PIPE_STEPS
    for s in range(last_step + 1, max_steps + 1):
        evt = line_step_done[sync_line][job_no][s]
        if not evt.triggered:
            evt.succeed()
    # 해당 line의 더 이상 다음 공정이 없다면 마지막 공정 점유는 여기서 해제한다.
    release_line_active_hold(job_no, line_name)


# 초보자 설명: 제품 타입과 피팅 정보를 보고 실제 사용 설비 순서를 계산한다.
def get_actual_fitting_station_sequence(prod, spool_name, rep_ft):
    seq = []
    rep_proc = get_fit_proc_from_ft(rep_ft)
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, rep_ft)
    if prod == "FF":
        if rep_proc in VALID_FITTING_TYPES:
            seq += [rep_proc, rep_proc]
    elif prod in ["F2D", "F3D", "F2DS", "F2DL"]:
        if fit1_proc in VALID_FITTING_TYPES:
            seq.append(fit1_proc)
        if fit2_proc in VALID_FITTING_TYPES:
            seq.append(fit2_proc)
    elif prod in ["2D", "H3D", "2DS", "2DL", "Short"]:
        if rep_proc in VALID_FITTING_TYPES:
            seq.append(rep_proc)
    return seq


# 초보자 설명: 각 Job이 설비를 몇 번 사용하는지 미리 계산해 FIFO 토큰을 만든다.
def count_station_uses(row):
    no, name, inch, Sch, prod, weld, length, ft = row[:8]
    counts = defaultdict(int)
    if prod == "FF":
        pass
    elif prod in ["H3D", "F3D"]:
        counts["Cut"] += 2
        counts["Bevel1"] += 2
        counts["Bevel2"] += 2
    elif prod in ["2DS", "F2DS"]:
        counts["Cut"] += 1
        counts["FIT_ELBOW_TEE"] += 2
    elif prod == "F2DL":
        counts["Cut"] += 1
        counts["Bevel1"] += 1
        counts["Bevel2"] += 1
    else:
        counts["Cut"] += 1
        counts["Bevel1"] += 1
        counts["Bevel2"] += 1
    for proc in get_actual_fitting_station_sequence(prod, name, ft):
        counts[get_sub_fitting_resource_name(proc)] += 1
    if prod == "2D":
        counts[weld if weld in ["2D-1", "2D-2"] else "3D"] += 1
    elif prod == "F2D":
        counts["2D-1"] += 1
        counts["3D"] += 1
    elif prod == "H3D":
        counts["2D-2"] += 1
        counts["3D"] += 1
    elif prod == "F3D":
        # F3D는 2D-1에서 1차 조립, 2D-2에서 2차 조립을 각각 수행한 뒤
        # 최종적으로 3D에서 두 조립품을 합류시킨다.
        counts["2D-1"] += 1
        counts["2D-2"] += 1
        counts["3D"] += 1
    elif prod in ["F2DS", "F2DL"]:
        counts["Short"] += 2
    elif prod == "FF":
        counts["Short"] += 1
    elif prod in ["2DS", "2DL", "Short"]:
        counts["Short"] += 1
    return counts


# 초보자 설명: 간트 차트 라벨에 들어갈 용접장 번호 표기를 만든다.
def get_weld_station_label(prod, weld):
    prod = normalize_product_type_name(prod)
    weld = norm_text(weld)

    if prod == "2D":
        nums = [WELD_LABEL_NUMBERS[weld]] if weld in WELD_LABEL_NUMBERS else []
        return f"({','.join(nums)})" if nums else ""

    if prod == "F2D":
        return "(1,3)"

    if prod == "H3D":
        return "(2,3)"

    if prod == "F3D":
        return "(1,2,3)"

    if prod in ["2DS", "2DL", "F2DS", "F2DL", "FF", "Short"]:
        return "(4)"

    return ""


# 초보자 설명: 설비별 강제 FIFO 토큰 목록을 현재 계획 기준으로 다시 만든다.
def rebuild_station_fifo_tokens(current_plan):
    global station_job_tokens, station_done_events, station_token_index, station_runtime_occ_counter
    station_job_tokens = defaultdict(list)
    station_done_events = {}
    station_token_index = {}
    station_runtime_occ_counter = defaultdict(int)
    for row in current_plan:
        job_no = row[0]
        counts = count_station_uses(row)
        for proc, cnt in counts.items():
            for occ in range(1, cnt + 1):
                station_job_tokens[proc].append((job_no, occ))
    for proc, tokens in station_job_tokens.items():
        station_done_events[proc] = {}
        station_token_index[proc] = {}
        for idx, token in enumerate(tokens):
            station_token_index[proc][token] = idx
            station_done_events[proc][token] = env.event()


# 초보자 설명: Sub 라인 시작 공정용 FIFO 토큰을 다시 만든다.
def rebuild_sub_start_fifo_tokens(current_plan):
    global sub_start_job_tokens, sub_start_done_events, sub_start_token_index
    sub_start_job_tokens = defaultdict(list)
    sub_start_done_events = {}
    sub_start_token_index = {}
    reset_proc_runtime_call_counter()
    for row in current_plan:
        job_no = row[0]
        first_proc = get_sub_start_proc_from_plan_row(row)
        second_proc = get_sub2_start_proc_from_plan_row(row)
        if first_proc:
            grp = "fl" if norm_text(first_proc) == "fl" else "common"
            sub_start_job_tokens[grp].append((job_no, 1))
        if second_proc:
            grp = "fl" if norm_text(second_proc) == "fl" else "common"
            sub_start_job_tokens[grp].append((job_no, 2))
    for grp, tokens in sub_start_job_tokens.items():
        sub_start_done_events[grp] = {}
        sub_start_token_index[grp] = {}
        for idx, token in enumerate(tokens):
            sub_start_token_index[grp][token] = idx
            sub_start_done_events[grp][token] = env.event()


# 초보자 설명: 같은 Sub 시작 설비를 쓰는 이전 Job 순서를 기다린다.
def wait_prev_sub_start_fifo(env, fit_proc, job_no, slot_no):
    grp = "fl" if norm_text(fit_proc) == "fl" else "common"
    token = (job_no, slot_no)
    if grp not in sub_start_token_index or token not in sub_start_token_index[grp]:
        return grp, None
    pos = sub_start_token_index[grp][token]
    if pos > 0:
        prev_tok = sub_start_job_tokens[grp][pos - 1]
        prev_evt = sub_start_done_events[grp][prev_tok]
        if not prev_evt.triggered:
            yield prev_evt
    return grp, pos


# 초보자 설명: Sub 시작 FIFO 토큰 완료를 표시한다.
def mark_sub_start_fifo_done(grp, pos):
    if pos is None:
        return
    tok = sub_start_job_tokens[grp][pos]
    evt = sub_start_done_events[grp][tok]
    if not evt.triggered:
        evt.succeed()


# 초보자 설명: 대부분의 일반 공정을 실행하는 공통 SimPy 공정 함수다.
def process(env, job, line, name, cycle_dict):
    if name == "Gan" and "Gan" in dyn_ct._override:
        ct = safe_ct(dyn_ct._override.get("Gan"), 1.0)
    else:
        ct = safe_ct(get_process_cycle_time(job, line, name, cycle_dict), 1.0)

    pending_key = _pending_weld_key(job, name)
    if name in WELD_PROCS and pending_weld_holds.get(pending_key):
        release_line_active_hold(job, line)
        yield env.process(finish_combined_weld_hold(env, job, line, name, ct))
        return

    resource_name = get_runtime_resource_name(name)
    fifo_key = resource_name if resource_name in STRICT_STATION_FIFO else name
    pos = None

    if name in WELD_PROCS:
        if fifo_key in STRICT_STATION_FIFO:
            _occ_no, pos = yield from wait_prev_station_fifo(env, fifo_key, job)

        if resource_name == "Short":
            yield from acquire_short_station(env, job)
            release_line_active_hold(job, line)
            start = env.now
            yield env.timeout(ct)
            end = env.now
            record(job, line, name, start, end)
            store_line_active_hold(job, line, kind="short", fifo_key=fifo_key, pos=pos)
            return

        req = resources[resource_name].request()
        yield req
        release_line_active_hold(job, line)
        start = env.now
        yield env.timeout(ct)
        end = env.now
        record(job, line, name, start, end)
        store_line_active_hold(job, line, kind="simpy", resource_name=resource_name, req=req, fifo_key=fifo_key,
                               pos=pos)
        return

    if fifo_key in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, fifo_key, job)

    req = resources[resource_name].request()
    yield req
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, name, start, end)
    if fifo_key in STRICT_STATION_FIFO:
        mark_station_fifo_done(fifo_key, pos)
    resources[resource_name].release(req)


# 초보자 설명: CutBuf에서 재고 1개를 꺼낸 뒤 다음 공정을 수행한다.
def process_after_cutbuf(env, job, line, name, cycle_dict):
    ct = safe_ct(get_process_cycle_time(job, line, name, cycle_dict), 1.0)
    resource_name = get_runtime_resource_name(name)
    fifo_key = resource_name if resource_name in STRICT_STATION_FIFO else name
    pos = None

    if fifo_key in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, fifo_key, job)

    req = resources[resource_name].request()
    yield req
    yield CutBuf.get(1)
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, name, start, end)
    if fifo_key in STRICT_STATION_FIFO:
        mark_station_fifo_done(fifo_key, pos)
    resources[resource_name].release(req)


# 초보자 설명: 피팅 공정을 실제 공유 설비 이름으로 변환한다.
def get_sub_fitting_resource_name(name):
    fit_name = norm_text(name)
    if fit_name in ELBOW_TEE_FITTING_TYPES:
        return "FIT_ELBOW_TEE"
    if fit_name == "fl":
        return "FIT_FL"
    return name


# 초보자 설명: Sub 피팅 공정을 제품 특성에 맞게 실행한다.
def process_sub_fitting(env, job, line, name, cycle_dict):
    if name in ["2DS", "sShrt", "sShBevel"]:
        yield env.process(process(env, job, line, "sShBevel", cycle_dict))
    else:
        yield env.process(process(env, job, line, name, cycle_dict))


# 초보자 설명: Short 계열이 공용 피팅 설비를 사용할 때의 실행 함수다.
def process_short_shared_fit(env, job, line, display_name, cycle_dict):
    ct = safe_ct(get_process_cycle_time(job, line, display_name, cycle_dict), 1.0)
    resource_name = "FIT_ELBOW_TEE"
    req = resources[resource_name].request()
    yield req
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, display_name, start, end)
    resources[resource_name].release(req)


# 초보자 설명: Sub 라인의 첫 피팅 공정을 FIFO 규칙까지 포함해 실행한다.
def process_sub_fitting_first(env, job, line, name, cycle_dict, slot_no):
    if not PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
        yield env.process(process_sub_fitting(env, job, line, name, cycle_dict))
        return
    grp, pos = yield from wait_prev_sub_start_fifo(env, name, job, slot_no)
    yield env.process(process_sub_fitting(env, job, line, name, cycle_dict))
    mark_sub_start_fifo_done(grp, pos)


# 초보자 설명: 설비 FIFO 토큰의 현재 위치를 찾는다.
def get_station_token_pos(proc, job_no, occ_no):
    token = (job_no, occ_no)
    if proc not in station_token_index:
        return None
    return station_token_index[proc].get(token, None)


# 초보자 설명: FIFO 토큰을 Job 번호 순으로 정렬할 때 쓰는 기준이다.
def _token_sort_key(token):
    return (token[0], token[1])


# 초보자 설명: 예상치 못한 토큰 누락이 생기면 자동 복구용 토큰을 만든다.
def ensure_station_token(proc, job_no, occ_no):
    global station_job_tokens, station_done_events, station_token_index
    token = (job_no, occ_no)
    if proc not in station_job_tokens:
        station_job_tokens[proc] = []
    if proc not in station_done_events:
        station_done_events[proc] = {}
    if proc not in station_token_index:
        station_token_index[proc] = {}
    if token in station_token_index[proc]:
        return station_token_index[proc][token]
    old_tokens = list(station_job_tokens[proc])
    new_tokens = list(old_tokens)
    new_tokens.append(token)
    new_tokens.sort(key=_token_sort_key)
    old_event_map = station_done_events.get(proc, {})
    new_event_map = {tok: old_event_map[tok] if tok in old_event_map else env.event() for tok in new_tokens}
    new_index_map = {tok: idx for idx, tok in enumerate(new_tokens)}
    station_job_tokens[proc] = new_tokens
    station_done_events[proc] = new_event_map
    station_token_index[proc] = new_index_map
    return station_token_index[proc][token]


# 초보자 설명: 현재 공정이 참조하는 FIFO 토큰이 정상인지 검사한다.
def validate_station_token(proc, job_no, occ_no, pos):
    token = (job_no, occ_no)
    if proc not in station_job_tokens:
        if AUTO_RECOVER_MISSING_STATION_TOKEN:
            return ensure_station_token(proc, job_no, occ_no)
        raise RuntimeError(f"station FIFO proc 누락: {proc}")
    if pos is None:
        if AUTO_RECOVER_MISSING_STATION_TOKEN:
            return ensure_station_token(proc, job_no, occ_no)
        raise RuntimeError(f"현재 시뮬레이션 PLAN 기준 token이 아닌 값 참조: proc={proc}, token={token}")
    if proc not in station_done_events:
        if AUTO_RECOVER_MISSING_STATION_TOKEN:
            ensure_station_token(proc, job_no, occ_no)
            return get_station_token_pos(proc, job_no, occ_no)
        raise RuntimeError(f"station_done_events proc 누락: {proc}")
    if token not in station_done_events[proc]:
        if AUTO_RECOVER_MISSING_STATION_TOKEN:
            return ensure_station_token(proc, job_no, occ_no)
        raise RuntimeError(f"station_done_events 누락: proc={proc}, token={token}")
    return pos


# 초보자 설명: 같은 설비를 쓰는 이전 토큰이 끝날 때까지 기다린다.
def wait_prev_station_fifo(env, proc, job_no):
    station_runtime_occ_counter[(proc, job_no)] += 1
    occ_no = station_runtime_occ_counter[(proc, job_no)]
    pos = get_station_token_pos(proc, job_no, occ_no)
    pos = validate_station_token(proc, job_no, occ_no, pos)
    if pos > 0:
        prev_tok = station_job_tokens[proc][pos - 1]
        prev_evt = station_done_events[proc].get(prev_tok, None)
        if prev_evt is None:
            if AUTO_RECOVER_MISSING_STATION_TOKEN:
                ensure_station_token(proc, prev_tok[0], prev_tok[1])
                prev_evt = station_done_events[proc].get(prev_tok, None)
            if prev_evt is None:
                raise RuntimeError(f"이전 token 완료 이벤트가 없음: proc={proc}, prev_token={prev_tok}")
        if not prev_evt.triggered:
            yield prev_evt
    return occ_no, pos


# 초보자 설명: 설비 FIFO 토큰 완료를 표시한다.
def mark_station_fifo_done(proc, pos):
    if pos is None:
        return
    tok = station_job_tokens[proc][pos]
    evt = station_done_events[proc].get(tok, None)
    if evt is None:
        if AUTO_RECOVER_MISSING_STATION_TOKEN:
            ensure_station_token(proc, tok[0], tok[1])
            evt = station_done_events[proc].get(tok, None)
        if evt is None:
            raise RuntimeError(f"완료 처리할 station_done_event가 없습니다: proc={proc}, token={tok}")
    if not evt.triggered:
        evt.succeed()


# 초보자 설명: 라인 점유 정보를 저장할 때 쓰는 키를 만든다.
def _line_hold_key(job, line):
    return (int(job), norm_text(line))


# 초보자 설명: 보관 중인 설비 점유를 즉시 해제한다.
def _release_line_active_hold_now(hold_info):
    hold_kind = hold_info.get("kind")
    fifo_key = hold_info.get("fifo_key")
    pos = hold_info.get("pos")

    if hold_kind == "short":
        release_short_station(env)
    elif hold_kind == "simpy":
        resource_name = hold_info.get("resource_name")
        req = hold_info.get("req")
        if resource_name is not None and req is not None:
            resources[resource_name].release(req)
    else:
        raise RuntimeError(f"알 수 없는 line hold kind: {hold_kind}")

    if fifo_key in STRICT_STATION_FIFO:
        mark_station_fifo_done(fifo_key, pos)


# 초보자 설명: 특정 Job/라인이 들고 있는 설비 점유를 해제한다.
def release_line_active_hold(job, line):
    key = _line_hold_key(job, line)
    hold_info = line_active_holds.pop(key, None)
    if hold_info is None:
        return
    _release_line_active_hold_now(hold_info)


# 초보자 설명: 나중에 이어서 쓰기 위해 현재 설비 점유 상태를 저장한다.
def store_line_active_hold(job, line, *, kind, resource_name=None, req=None, fifo_key=None, pos=None):
    key = _line_hold_key(job, line)
    prev_hold = line_active_holds.pop(key, None)
    if prev_hold is not None:
        _release_line_active_hold_now(prev_hold)
    line_active_holds[key] = {
        "kind": kind,
        "resource_name": resource_name,
        "req": req,
        "fifo_key": fifo_key,
        "pos": pos,
    }


# 초보자 설명: 한 Job이 들고 있는 모든 설비 점유를 정리한다.
def release_all_job_line_active_holds(job):
    release_keys = [key for key in list(line_active_holds.keys()) if key[0] == int(job)]
    for key in release_keys:
        hold_info = line_active_holds.pop(key, None)
        if hold_info is not None:
            _release_line_active_hold_now(hold_info)


# 초보자 설명: Short 설비 대기열에서 다음 작업을 시작할지 판단한다.
def _try_schedule_short_station_dispatch(env):
    global short_station_dispatch_scheduled
    if (not short_station_busy) and short_station_pending and (not short_station_dispatch_scheduled):
        short_station_dispatch_scheduled = True
        env.process(_dispatch_short_station(env))


# 초보자 설명: Short 설비 대기열에서 실제 다음 작업을 꺼내 실행시킨다.
def _dispatch_short_station(env):
    global short_station_dispatch_scheduled, short_station_busy
    yield env.timeout(0)
    yield env.timeout(0)
    short_station_dispatch_scheduled = False
    if short_station_busy or not short_station_pending:
        return
    best_idx = min(
        range(len(short_station_pending)),
        key=lambda i: (short_station_pending[i][1], short_station_pending[i][0]),
    )
    _ready_time, _job_no, evt = short_station_pending.pop(best_idx)
    short_station_busy = True
    if not evt.triggered:
        evt.succeed()


# 초보자 설명: Short 설비 사용 권한을 얻을 때까지 기다린다.
def acquire_short_station(env, job_no):
    evt = env.event()
    short_station_pending.append((float(env.now), int(job_no), evt))
    _try_schedule_short_station_dispatch(env)
    yield evt


# 초보자 설명: Short 설비 사용을 끝내고 다음 Job에게 넘긴다.
def release_short_station(env):
    global short_station_busy
    short_station_busy = False
    _try_schedule_short_station_dispatch(env)


# 초보자 설명: 결합 용접 hold 상태를 저장할 때 쓰는 키다.
def _pending_weld_key(job, weld_proc):
    return (int(job), norm_text(weld_proc))


# 초보자 설명: Main 용접 도착 신호를 저장할 때 쓰는 키다.
def _main_weld_arrival_key(job, weld_proc):
    return (int(job), norm_text(weld_proc))


# 초보자 설명: Main 라인이 용접 위치에 도착했음을 다른 라인에 알린다.
def mark_main_weld_arrival(job, weld_proc):
    key = _main_weld_arrival_key(job, weld_proc)
    waiters = waiting_main_weld_arrivals.get(key)
    if waiters:
        evt = waiters.popleft()
        if not evt.triggered:
            evt.succeed()
    else:
        pending_main_weld_arrival_counts[key] += 1


# 초보자 설명: Sub 라인이 Main 라인 도착 신호를 기다린다.
def wait_for_main_weld_arrival(env, job, weld_proc):
    key = _main_weld_arrival_key(job, weld_proc)
    if pending_main_weld_arrival_counts.get(key, 0) > 0:
        pending_main_weld_arrival_counts[key] -= 1
        return
    evt = env.event()
    waiting_main_weld_arrivals[key].append(evt)
    yield evt


# 초보자 설명: 로봇+tig가 이어지는 결합 용접 시작 구간을 처리한다.
def start_combined_weld_hold(env, job, line, display_name, weld_proc, hold_ct):
    hold_ct = safe_ct(hold_ct, 1.0)
    weld_proc = norm_text(weld_proc)
    resource_name = get_runtime_resource_name(weld_proc)
    fifo_key = resource_name if resource_name in STRICT_STATION_FIFO else weld_proc

    pos = None
    if fifo_key in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, fifo_key, job)

    req = None
    uses_short_station = (resource_name == "Short")
    if uses_short_station:
        yield from acquire_short_station(env, job)
    else:
        req = resources[resource_name].request()
        yield req

    release_line_active_hold(job, line)
    start = env.now
    yield env.timeout(hold_ct)
    end = env.now
    record(job, line, display_name, start, end)

    pending_weld_holds[_pending_weld_key(job, weld_proc)].append({
        'req': req,
        'resource_name': resource_name,
        'fifo_key': fifo_key,
        'pos': pos,
        'uses_short_station': uses_short_station,
    })


# 초보자 설명: 앞에서 잡아 둔 결합 용접 hold를 실제 용접 완료까지 이어서 마무리한다.
def finish_combined_weld_hold(env, job, line, weld_proc, weld_ct):
    weld_ct = safe_ct(weld_ct, 1.0)
    weld_proc = norm_text(weld_proc)
    key = _pending_weld_key(job, weld_proc)
    if not pending_weld_holds.get(key):
        raise RuntimeError(f'결합 용접 hold가 없습니다: job={job}, proc={weld_proc}')

    hold_info = pending_weld_holds[key].popleft()
    req = hold_info['req']
    fifo_key = hold_info['fifo_key']
    pos = hold_info['pos']

    start = env.now
    yield env.timeout(weld_ct)
    end = env.now
    record(job, line, weld_proc, start, end)

    store_line_active_hold(
        job,
        line,
        kind="short" if hold_info.get('uses_short_station') else "simpy",
        resource_name=None if hold_info.get('uses_short_station') else hold_info['resource_name'],
        req=None if hold_info.get('uses_short_station') else req,
        fifo_key=fifo_key,
        pos=pos,
    )


# 초보자 설명: 실행 중 특정 공정의 시간을 임시로 덮어쓸 때 사용하는 작은 보조 클래스다.
class DynamicCycleTime:
    # 초보자 설명: '__init__' 단계의 처리를 맡는 보조 함수다.
    def __init__(self):
        self._override = {}

    # 초보자 설명: 'set' 단계의 처리를 맡는 보조 함수다.
    def set(self, proc_name, value):
        if value is None:
            self._override.pop(proc_name, None)
        else:
            self._override[proc_name] = float(value)

    # 초보자 설명: 'get' 단계의 처리를 맡는 보조 함수다.
    def get(self, proc_name, fallback):
        return self._override.get(proc_name, fallback)


dyn_ct = DynamicCycleTime()
proc_runtime_call_counter = defaultdict(int)


# 초보자 설명: 같은 공정이 몇 번째 호출인지 세는 카운터를 초기화한다.
def reset_proc_runtime_call_counter():
    global proc_runtime_call_counter
    proc_runtime_call_counter = defaultdict(int)


# 초보자 설명: 현재 공정이 몇 번째 실행인지 번호를 하나 올려 반환한다.
def _next_proc_call_occ(job, line, name):
    proc_runtime_call_counter[(int(job), norm_text(line), norm_text(name))] += 1
    return proc_runtime_call_counter[(int(job), norm_text(line), norm_text(name))]


# 초보자 설명: 현재 또는 다음 호출 번호를 미리 확인한다.
def _current_or_next_occ(job, line, name):
    return proc_runtime_call_counter.get((int(job), norm_text(line), norm_text(name)), 0) + 1


# 초보자 설명: 정의되지 않은 공정이 나오면 디버깅에 필요한 정보를 담아 오류를 낸다.
def _raise_undefined_process(job, line, proc, cycle_dict, detail=""):
    prod = cycle_dict.get("_prod", "")
    weld = cycle_dict.get("_weld", "")
    length = cycle_dict.get("_length", "")
    extra = f"\\n{detail}" if detail else ""
    raise ValueError(
        f"정의되지 않은 공정입니다.\\n"
        f"Job No: {job}\\n"
        f"Line: {line}\\n"
        f"Process: {proc}\\n"
        f"Prod: {prod}\\n"
        f"Weld: {weld}\\n"
        f"Length: {length}{extra}"
    )


# 초보자 설명: 공정 시간이 반드시 있어야 할 때 없으면 바로 오류를 낸다.
def _require_defined_ct(job, line, proc, cycle_dict, value, detail=""):
    if value is None:
        _raise_undefined_process(job, line, proc, cycle_dict, detail=detail)
    return float(value)


# 초보자 설명: 현재 라인 문맥에서 우선 참조할 용접 단계를 정한다.
def _preferred_stage_for_line(cycle_dict, line):
    prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
    line = norm_text(line)
    if prod in ["2DS", "2DL", "F2DS", "F2DL", "FF", "Short"]:
        return "Short"
    if prod == "2D":
        return norm_text(cycle_dict.get("_weld", ""))
    if prod == "F2D":
        if line in ["Pipe", "Sub"]:
            return "2D-1"
        return "3D"
    if prod == "H3D":
        if line in ["Pipe", "Sub"]:
            return "2D-2"
        return "3D"
    if prod == "F3D":
        if line in ["Pipe", "Sub"]:
            return "2D-1"
        if line in ["Pipe2", "Sub2"]:
            return "2D-2"
        return "3D"
    return norm_text(cycle_dict.get("_weld", "")) or "3D"


# 초보자 설명: 준비 공정이 어느 용접 단계의 CycleTime을 봐야 하는지 정한다.
def _prep_stage_for_proc(cycle_dict, line, proc):
    prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
    line = norm_text(line)
    proc = norm_text(proc)
    prep_procs = {"Cut", "Con1", "Conv1", "Sen1", "Bevel1", "Bev1", "Con2", "Conv2", "Bevel2", "Bev2", "Sen2", "Conv3"}
    if proc not in prep_procs:
        return _preferred_stage_for_line(cycle_dict, line)
    if line == "Pipe2" and prod in ["H3D", "F3D"]:
        return "2D-2"
    if line == "Pipe" and prod in ["F2D", "F3D"]:
        return "2D-1"
    if line == "Pipe" and prod == "H3D":
        return "2D-2"
    return _preferred_stage_for_line(cycle_dict, line)


# =====================================================
# 공정 시간 계산 로직
# -----------------------------------------------------
# 공정 이름 하나만으로는 시간을 알 수 없고, 제품 타입/용접장/라인 문맥을
# 함께 봐야 하므로 여기서 실제 CycleTime 값을 최종 결정한다.
# =====================================================

# 초보자 설명: 공정 이름과 제품 문맥을 바탕으로 실제 CycleTime 값을 찾아준다.
def get_process_cycle_time(job, line, name, cycle_dict):
    proc = norm_text(name)
    line = norm_text(line)
    prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
    occ = _next_proc_call_occ(job, line, proc)
    stage = _preferred_stage_for_line(cycle_dict, line)
    prep_stage = _prep_stage_for_proc(cycle_dict, line, proc)

    if proc == "Cut":
        val = _ctx_value(cycle_dict, "cut", prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc in ["Con1", "Conv1"]:
        val = _ctx_value(cycle_dict, "conv1", prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "Sen1":
        col = "sen1_1" if occ == 1 else "sen1_2"
        val = _ctx_value(cycle_dict, col, prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val, detail=f"occurrence={occ}")
    if proc in ["Bevel1", "Bev1"]:
        if prep_stage == "2D-1":
            primary, secondary = "bev1_j", "bev1_v"
        else:
            primary, secondary = "bev1_v", "bev1_j"
        val = _ctx_value(cycle_dict, primary, prep_stage, ("2D-1", "2D-2", "3D", "Short"))
        if val is None:
            val = _ctx_value(cycle_dict, secondary, prep_stage, ("2D-1", "2D-2", "3D", "Short"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc in ["Con2", "Conv2"]:
        val = _ctx_value(cycle_dict, "conv2", prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "Conv3":
        val = _ctx_value(cycle_dict, "conv3", prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc in ["Bevel2", "Bev2"]:
        if prep_stage == "2D-1":
            primary, secondary = "bev2_v", "bev2_j"
        else:
            primary, secondary = "bev2_j", "bev2_v"
        val = _ctx_value(cycle_dict, primary, prep_stage, ("2D-1", "2D-2", "3D", "Short"))
        if val is None:
            val = _ctx_value(cycle_dict, secondary, prep_stage, ("2D-1", "2D-2", "3D", "Short"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "Sen2":
        primary = "sen2_1" if occ == 1 else "sen2_2"
        secondary = "sen2_2" if primary == "sen2_1" else "sen2_1"
        val = _ctx_value(cycle_dict, primary, prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        if val is None:
            val = _ctx_value(cycle_dict, secondary, prep_stage, ("2D-1", "2D-2", "Short", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val, detail=f"occurrence={occ}")
    if proc == "Gan":
        # Gan1 in legacy flow
        if stage == "2D-1":
            val = _ctx_value(cycle_dict, "gan1_2d1", "2D-1", ("3D",))
        elif stage == "2D-2":
            val = _ctx_value(cycle_dict, "gan1_2d2", "2D-2", ("3D",))
        else:
            val = _ctx_value(cycle_dict, "gan1_3d", "3D", ("2D-1", "2D-2"))
            if val is None:
                val = _ctx_value(cycle_dict, "gan4", "3D", ("2D-1", "2D-2"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "Gan2":
        col = {"2D-1": "out_2d1", "2D-2": "out_2d2", "3D": "out_3d", "Short": "out_short"}.get(stage, "out_3d")
        val = _ctx_value(cycle_dict, col, stage, ("3D", "2D-1", "2D-2", "Short"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "R1":
        if stage == "Short" and prod in ["2DS", "F2DS"]:
            val = _ctx_short_prep_value(cycle_dict, "r1_short")
            return _require_defined_ct(job, line, proc, cycle_dict, val)
        # 새 포맷에 R1 전용 컬럼이 없으므로 기존 generic robot time 유지
        val = _ctx_value(cycle_dict, "r2_2d1", stage, ("2D-1", "2D-2", "Short", "3D"), default=1.0)
        if val is None:
            val = 1.0
        return float(val)
    if proc == "R2":
        if line in ["Sub", "Sub2"] and prod in ["2DS", "2DL", "F2DS", "F2DL", "FF", "Short"]:
            col = "r2_fit_short_1" if occ == 1 else "r2_fit_short_2"
            val = _ctx_value(cycle_dict, col, "Short", ("Short",))
            return _require_defined_ct(job, line, proc, cycle_dict, val, detail=f"occurrence={occ}")
        if line in ["Pipe", "Pipe2"] and prod in ["2DS", "2DL", "F2DS", "F2DL", "Short"]:
            col = "r2_short_long" if prod in ["2DL", "F2DL"] else "r2_short_short"
            val = _ctx_value(cycle_dict, col, "Short", ("Short",))
            return _require_defined_ct(job, line, proc, cycle_dict, val)
        val = _ctx_value(cycle_dict, "r2_2d1" if stage == "2D-1" else "r2_2d2", stage, ("2D-1", "2D-2", "3D"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "R3":
        val = _ctx_value(cycle_dict, "r3_2d2_to_3d", "3D", ("3D", "2D-2"))
        if val is None:
            val = _ctx_value(cycle_dict, "r3_fit_3d", "3D", ("3D",))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "R14":
        # legacy plain R14는 Short 후 배출로 해석
        val = _ctx_value(cycle_dict, "out_short", "Short", ("Short",))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "FitMov":
        if stage == "Short" and prod in ["2DS", "F2DS"]:
            val = _ctx_short_prep_value(cycle_dict, "fitmov_short")
            return _require_defined_ct(job, line, proc, cycle_dict, val)
        # 새 포맷에 직접 컬럼이 없어서 기존 short prep 이동시간과 동일하게 유지
        val = _ctx_value(cycle_dict, "r14_pick_1", "Short", ("Short",), default=3.0)
        if val is None:
            val = 3.0
        return float(val)
    if proc == "sShBevel":
        if stage == "Short" and prod in ["2DS", "F2DS"]:
            if occ == 1:
                val = _ctx_short_prep_value(cycle_dict, ("shrt_bev1_j", "shrt_bev1_v"))
            else:
                val = _ctx_short_prep_value(cycle_dict, ("shrt_bev2_j", "shrt_bev2_v"))
            return _require_defined_ct(job, line, proc, cycle_dict, val, detail=f"occurrence={occ}")
        col = "short_double" if prod in ["F2DS", "F2DL"] else "short_single"
        val = _ctx_value(cycle_dict, col, "Short", ("Short",))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "45el":
        return _require_defined_ct(job, line, proc, cycle_dict,
                                   _ctx_value(cycle_dict, "fit_45el", stage, ("2D-1", "2D-2", "3D", "Short")))
    if proc == "90el":
        val = _ctx_value(cycle_dict, "fit_90el", stage, ("2D-1", "2D-2", "3D", "Short"))
        if val is None:
            val = _ctx_value(cycle_dict, "fit_90el_both", stage, ("2D-1", "2D-2", "3D", "Short"))
        return _require_defined_ct(job, line, proc, cycle_dict, val)
    if proc == "fl":
        return _require_defined_ct(job, line, proc, cycle_dict,
                                   _ctx_value(cycle_dict, "fit_fl", stage, ("2D-1", "2D-2", "3D", "Short")))
    if proc in ["Tee", "ReTee"]:
        return _require_defined_ct(job, line, proc, cycle_dict,
                                   _ctx_value(cycle_dict, "fit_tee", stage, ("2D-1", "2D-2", "3D", "Short")))
    if proc == "2D-1":
        return _require_defined_ct(job, line, proc, cycle_dict, _ctx_value(cycle_dict, "tig_spf", "2D-1", ("2D-1",)))
    if proc == "2D-2":
        return _require_defined_ct(job, line, proc, cycle_dict, _ctx_value(cycle_dict, "tig_spf", "2D-2", ("2D-2",)))
    if proc == "3D":
        return _require_defined_ct(job, line, proc, cycle_dict, _ctx_value(cycle_dict, "tig_spf", "3D", ("3D",)))
    if proc == "Short":
        return _require_defined_ct(job, line, proc, cycle_dict,
                                   _ctx_value(cycle_dict, "tig_short_1", "Short", ("Short",)))
    if proc == "tig":
        return _require_defined_ct(job, line, proc, cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage, ("2D-1", "2D-2", "3D", "Short")))
    val = _ctx_value(cycle_dict, proc, stage, ("2D-1", "2D-2", "3D", "Short"))
    return _require_defined_ct(job, line, proc, cycle_dict, val)


# 초보자 설명: R2/R3/R14 같은 로봇의 이동+tig 시간을 상황별 규칙으로 계산한다.
def get_robot_step_timing(job, line_name, robot_name, next_proc, from_gantry, cycle_dict, use_tig_display):
    line_name = norm_text(line_name)
    next_proc = norm_text(next_proc)
    occ = _current_or_next_occ(job, line_name, robot_name)
    display_name = robot_name

    if robot_name == "R2":
        if next_proc == "2D-1":
            if use_tig_display:
                val = _ctx_value(cycle_dict, "r2_fit_2d1", "2D-1", ("2D-1",), 0.0)
                tig = _ctx_value(cycle_dict, "tig_spf", "2D-1", ("2D-1",), 0.0)
                ct = (val or 0.0) + (tig or 0.0)
                display_name = "R2tig"
            elif from_gantry:
                ct = _ctx_value(cycle_dict, "tig_spf", "2D-1", ("2D-1",))
            else:
                ct = _ctx_value(cycle_dict, "r2_2d1", "2D-1", ("2D-1",))
        elif next_proc == "2D-2":
            if use_tig_display:
                val = _ctx_value(cycle_dict, "r2_fit_2d2", "2D-2", ("2D-2",), 0.0)
                tig = _ctx_value(cycle_dict, "tig_spf", "2D-2", ("2D-2",), 0.0)
                ct = (val or 0.0) + (tig or 0.0)
                display_name = "R2tig"
            elif from_gantry:
                ct = _ctx_value(cycle_dict, "tig_spf", "2D-2", ("2D-2",))
            else:
                ct = _ctx_value(cycle_dict, "r2_2d2", "2D-2", ("2D-2",))
        elif next_proc == "Short":
            if use_tig_display:
                # sub line에서는 R14가 tig를 수행하고 R2는 handoff만 수행한다.
                ct = _ctx_value(cycle_dict, "r2_fit_short_1" if occ == 1 else "r2_fit_short_2", "Short", ("Short",))
            elif line_name in ["Sub", "Sub2"]:
                ct = _ctx_value(cycle_dict, "r2_fit_short_1" if occ == 1 else "r2_fit_short_2", "Short", ("Short",))
            else:
                prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
                col = "r2_short_long" if prod in ["2DL", "F2DL"] else "r2_short_short"
                ct = _ctx_value(cycle_dict, col, "Short", ("Short",))
        else:
            ct = _ctx_value(cycle_dict, "r2_2d1", _preferred_stage_for_line(cycle_dict, line_name),
                            ("2D-1", "2D-2", "3D"))
    elif robot_name == "R3":
        if next_proc == "3D" and use_tig_display:
            val = _ctx_value(cycle_dict, "r3_fit_3d", "3D", ("3D",), 0.0)
            tig = _ctx_value(cycle_dict, "tig_spf", "3D", ("3D",), 0.0)
            ct = (val or 0.0) + (tig or 0.0)
            display_name = "R3tig"
        elif next_proc == "3D" and from_gantry:
            ct = _ctx_value(cycle_dict, "tig_spf", "3D", ("3D",))
        elif next_proc == "3D":
            ct = _ctx_value(cycle_dict, "r3_2d2_to_3d", "3D", ("3D",))
            if ct is None:
                ct = _ctx_value(cycle_dict, "r3_fit_3d", "3D", ("3D",))
        else:
            ct = _ctx_value(cycle_dict, "r3_fit_3d", "3D", ("3D",))
    elif robot_name == "R14":
        if next_proc == "Short":
            if use_tig_display:
                if occ == 1:
                    base = _ctx_value(cycle_dict, "r14_pick_2", "Short", ("Short",), 0.0)
                    tig = _ctx_value(cycle_dict, "tig_short_1", "Short", ("Short",), 0.0)
                else:
                    base = _ctx_value(cycle_dict, "r14_pick_2b", "Short", ("Short",), 0.0)
                    tig = _ctx_value(cycle_dict, "tig_short_2", "Short", ("Short",), 0.0)
                    if base is None:
                        base = _ctx_value(cycle_dict, "r14_pick_2", "Short", ("Short",), 0.0)
                    if tig is None:
                        tig = _ctx_value(cycle_dict, "tig_short_1", "Short", ("Short",), 0.0)
                ct = (base or 0.0) + (tig or 0.0)
                display_name = "R14tig"
            else:
                if occ == 1:
                    ct = _ctx_value(cycle_dict, "r14_pick_1", "Short", ("Short",))
                else:
                    ct = _ctx_value(cycle_dict, "r14_pick_2b", "Short", ("Short",))
                    if ct is None:
                        ct = _ctx_value(cycle_dict, "r14_pick_1", "Short", ("Short",))
        else:
            ct = _ctx_value(cycle_dict, "out_short", "Short", ("Short",))
    else:
        ct = 1.0

    ct = _require_defined_ct(job, line_name, robot_name, cycle_dict, ct,
                             detail=f"next_proc={next_proc}, from_gantry={from_gantry}, tig={use_tig_display}")
    return display_name, ct


# 초보자 설명: Gan 공정이 어느 단계 이동인지에 따라 시간을 계산한다.
def get_gan_step_cycle_time(job, line_name, gan_key, cycle_dict):
    prod = normalize_product_type_name(cycle_dict.get("_prod", ""))
    line_name = norm_text(line_name)
    gan_key = norm_text(gan_key)
    occ = _next_proc_call_occ(job, line_name, f"_{gan_key}")

    if gan_key == "Gan2":
        if prod == "2D":
            final_stage = norm_text(cycle_dict.get("_weld", "")) or "3D"
        elif prod in ["F2D", "H3D", "F3D"]:
            final_stage = "3D"
        elif prod in ["2DS", "2DL", "F2DS", "F2DL", "FF", "Short"]:
            final_stage = "Short"
        else:
            final_stage = norm_text(cycle_dict.get("_weld", "")) or "3D"
        col = {"2D-1": "out_2d1", "2D-2": "out_2d2", "3D": "out_3d", "Short": "out_short"}.get(final_stage, "out_3d")
        return _ctx_value(cycle_dict, col, final_stage, ("2D-1", "2D-2", "3D", "Short"))

    if gan_key == "Gan1":
        if prod == "2D":
            target = norm_text(cycle_dict.get("_weld", "")) or "3D"
            if target == "2D-1":
                return _ctx_value(cycle_dict, "gan1_2d1", "2D-1", ("2D-1",))
            if target == "2D-2":
                return _ctx_value(cycle_dict, "gan1_2d2", "2D-2", ("2D-2",))
            return _ctx_value(cycle_dict, "gan1_3d", "3D", ("3D",))
        if prod == "F2D":
            return _ctx_value(cycle_dict, "gan1_2d1", "2D-1", ("2D-1",)) if occ == 1 else _ctx_value(cycle_dict, "gan4",
                                                                                                     "3D", ("3D",))
        if prod == "F3D":
            if line_name == "Pipe2":
                return _ctx_value(cycle_dict, "gan1_2d2", "2D-2", ("2D-2",))
            return _ctx_value(cycle_dict, "gan1_2d1", "2D-1", ("2D-1",)) if occ == 1 else _ctx_value(cycle_dict, "gan4",
                                                                                                     "3D", ("3D",))
        if prod == "H3D":
            if line_name == "Pipe2":
                return _ctx_value(cycle_dict, "gan1_3d", "3D", ("3D",))
            return _ctx_value(cycle_dict, "gan1_2d2", "2D-2", ("2D-2",))

    return _ctx_value(cycle_dict, "gan1_3d", "3D", ("3D", "2D-1", "2D-2"))


# 초보자 설명: Gan 공정을 한 단계 실행한다.
def run_gan_step(env, jNo, line_name, step_no, cycle_dict, gan_key):
    prev_gan_override = dyn_ct._override.get("Gan", None)
    dyn_ct.set("Gan", get_gan_step_cycle_time(jNo, line_name, gan_key, cycle_dict))
    yield env.process(run_line_step(env, line_name, jNo, step_no, process(env, jNo, line_name, "Gan", cycle_dict)))
    if prev_gan_override is not None:
        dyn_ct.set("Gan", prev_gan_override)
    else:
        dyn_ct.set("Gan", None)


# 초보자 설명: 현재 공정명이 실제 SimPy 자원 이름과 다를 때 맞춰 준다.
def get_runtime_resource_name(name):
    resource_name = get_sub_fitting_resource_name(name)
    return resource_name if resource_name in resources else name


# 초보자 설명: Short S 계열의 첫 Short 용접을 수행하는 얇은 래퍼 함수다.
def process_short_s_first_short(env, job, line, cycle_dict):
    yield env.process(process(env, job, line, "Short", cycle_dict))


# 초보자 설명: 현재 라인에서 Short 공정을 바로 수행하는 얇은 래퍼 함수다.
def process_short_s_local_short(env, job, line, cycle_dict):
    yield env.process(process(env, job, line, "Short", cycle_dict))


# 초보자 설명: Main 라인에서 로봇+tig 표시용 공정을 수행한다.
def process_main_robot_tig(env, job, line, robot_name, cycle_dict, display_name=None):
    display_name = display_name or ROBOT_TIG_DISPLAY.get(robot_name, robot_name)
    _d, ct = get_robot_step_timing(job, line, robot_name, cycle_dict.get("_weld", ""), False, cycle_dict, True)
    yield env.process(process_robot_with_display(env, job, line, robot_name, display_name, ct))


# 초보자 설명: 로봇+tig 뒤에 Short가 이어지는 결합 공정을 시작한다.
def process_main_robot_tig_and_short(env, job, line, robot_name, weld_proc, cycle_dict, display_name=None):
    display_name = display_name or ROBOT_TIG_DISPLAY.get(robot_name, robot_name)
    _d, ct = get_robot_step_timing(job, line, robot_name, weld_proc, False, cycle_dict, True)
    yield env.process(start_combined_weld_hold(env, job, line, display_name, weld_proc, ct))


# 초보자 설명: 연속 Short 블록 상태를 구분할 키를 만든다.
def _continuous_short_block_key(job, line):
    return (int(job), norm_text(line))


# 초보자 설명: Short 설비를 연속으로 쓰는 블록의 시작을 처리한다.
def begin_continuous_short_block(env, job, line, cycle_dict, robot_name="R14", display_name="R14tig"):
    key = _continuous_short_block_key(job, line)
    if key in pending_continuous_short_blocks:
        raise RuntimeError(f"이미 연속 Short 블록이 시작되었습니다: job={job}, line={line}")

    yield from acquire_short_station(env, job)
    pending_continuous_short_blocks[key] = True

    _d, tig_ct = get_robot_step_timing(job, line, robot_name, "Short", False, cycle_dict, True)
    req = resources[robot_name].request()
    yield req
    start = env.now
    yield env.timeout(tig_ct)
    end = env.now
    record(job, line, display_name, start, end)
    resources[robot_name].release(req)


# 초보자 설명: 연속 Short 블록 안에서 다음 R14tig를 이어서 수행한다.
def continue_continuous_short_block_r14tig(env, job, line, cycle_dict, robot_name="R14", display_name="R14tig"):
    key = _continuous_short_block_key(job, line)
    if key not in pending_continuous_short_blocks:
        raise RuntimeError(f"연속 Short 블록이 없습니다: job={job}, line={line}")

    _d, tig_ct = get_robot_step_timing(job, line, robot_name, "Short", False, cycle_dict, True)
    req = resources[robot_name].request()
    yield req
    start = env.now
    yield env.timeout(tig_ct)
    end = env.now
    record(job, line, display_name, start, end)
    resources[robot_name].release(req)


# 초보자 설명: 연속 Short 블록 안에서 Short 본공정을 이어서 수행한다.
def continue_continuous_short_block_short(env, job, line, cycle_dict, release_after=False):
    key = _continuous_short_block_key(job, line)
    req = pending_continuous_short_blocks.get(key)
    if req is None:
        raise RuntimeError(f"연속 Short 블록이 없습니다: job={job}, line={line}")

    release_line_active_hold(job, line)
    short_ct = safe_ct(get_process_cycle_time(job, line, "Short", cycle_dict), 1.0)
    start = env.now
    yield env.timeout(short_ct)
    end = env.now
    record(job, line, "Short", start, end)

    if release_after:
        pending_continuous_short_blocks.pop(key, None)
        store_line_active_hold(job, line, kind="short", fifo_key="Short", pos=None)


# 초보자 설명: FF 타입의 main 쪽 R14tig와 Short 연속 구간을 처리한다.
def process_ff_main_r14tig_and_short(env, job, line, cycle_dict):
    _d, tig_ct = get_robot_step_timing(job, line, "R14", "Short", False, cycle_dict, True)
    yield env.process(process_robot_with_display(env, job, line, "R14", "R14tig", tig_ct))
    yield env.process(process(env, job, line, "Short", cycle_dict))


# 초보자 설명: FF 타입의 최종 M1 공정 시퀀스를 순서대로 실행한다.
def run_ff_main_final_sequence(env, jNo, pipe_step, cycle_dict, ff_fit1_ready_evt=None, ff_fit2_ready_evt=None):
    """FF M1 최종 공정.

    [S1,S2 합류] -> Pick1(AM) -> Pick2(AN) -> Angle(AP) -> R14Tig1(AQ)
    -> Tig2(AR) -> Mag(AS) -> R14(BC)

    자원 점유 규칙
    - R14: Pick1 ~ Angle 동안 점유, 이후 해제
    - Short: Pick1 ~ Mag 동안 점유, 이후 해제
    """
    line_name = "Pipe"
    stage = "Short"

    if ff_fit1_ready_evt is not None and not ff_fit1_ready_evt.triggered:
        yield ff_fit1_ready_evt
    if ff_fit2_ready_evt is not None and not ff_fit2_ready_evt.triggered:
        yield ff_fit2_ready_evt

    pick1_ct = _require_defined_ct(
        jNo, line_name, "Pick1", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_1", stage, (stage,)),
    )
    pick2_ct = _require_defined_ct(
        jNo, line_name, "Pick2", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_2", stage, (stage,)),
    )
    angle_ct = _require_defined_ct(
        jNo, line_name, "Angle", cycle_dict,
        _ctx_value(cycle_dict, "fitup3d_short_1", stage, (stage,)),
    )
    r14tig_ct = _require_defined_ct(
        jNo, line_name, "R14tig", cycle_dict,
        _ctx_value(cycle_dict, "tig_short_1", stage, (stage,)),
    )
    tig2_ct = _require_defined_ct(
        jNo, line_name, "Tig2", cycle_dict,
        _ctx_value(cycle_dict, "tig2_short_1", stage, (stage,)),
    )
    mag_ct = _require_defined_ct(
        jNo, line_name, "Mag", cycle_dict,
        _ctx_value(cycle_dict, "mag_short_1", stage, (stage,)),
    )
    r14_out_ct = _require_defined_ct(
        jNo, line_name, "R14", cycle_dict,
        _ctx_value(cycle_dict, "out_short", stage, (stage,)),
    )

    yield from acquire_short_station(env, jNo)

    r14_req = resources["R14"].request()
    yield r14_req

    for display_name, ct in [("Pick1", pick1_ct), ("Pick2", pick2_ct), ("Angle", angle_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    resources["R14"].release(r14_req)

    for display_name, ct in [("R14tig", r14tig_ct), ("Tig2", tig2_ct), ("Mag", mag_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    release_short_station(env)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "R14", "R14", r14_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step


# 초보자 설명: 치명적인 규칙 오류가 발견되면 메시지를 출력하고 즉시 중단한다.
def alarm(msg):
    print(msg)
    raise SystemExit(msg)


# 초보자 설명: 용접장에 따라 Bevel1/Bevel2가 J/V 중 무엇인지 정한다.
def get_2d_bevel_types(weld):
    w = norm_text(weld)
    if w == "2D-1":
        return "J", "V"
    elif w == "2D-2":
        return "V", "J"
    elif w == "3D":
        return "V", "J"
    else:
        alarm("ALARM 3 : 용접장 미지정")


# 초보자 설명: 아직 실제 공정을 넣지 않은 자리에서 흐름만 유지하기 위한 빈 단계다.
def process_optional_future_step(env, job, line, logical_name, cycle_dict):
    yield env.timeout(0)


WELD_PROCS = {"2D-1", "2D-2", "3D", "Short"}
ROBOT_TIG_DISPLAY = {"R2": "R2tig", "R3": "R3tig", "R14": "R14tig"}


# 초보자 설명: 지정한 자원을 점유한 채 원하는 표시 이름으로 간트 기록을 남긴다.
def process_robot_with_display(env, job, line, resource_name, display_name, ct):
    ct = safe_ct(ct, 1.0)
    req = resources[resource_name].request()
    yield req
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, display_name, start, end)
    resources[resource_name].release(req)


# 초보자 설명: 특정 자원을 쓰는 임의 이름 공정을 기록용으로 실행한다.
def process_named_with_resource(env, job, line, resource_name, display_name, ct):
    ct = safe_ct(ct, 1.0)
    req = resources[resource_name].request()
    yield req
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, display_name, start, end)
    resources[resource_name].release(req)


# 초보자 설명: 이미 점유 중인 자원 위에서 이름만 바꿔 추가 공정을 기록한다.
def process_named_on_held_resource(env, job, line, display_name, ct):
    ct = safe_ct(ct, 1.0)
    start = env.now
    yield env.timeout(ct)
    end = env.now
    record(job, line, display_name, start, end)


# 초보자 설명: Main 라인의 Sen1-Bevel-Sen2 묶음을 한 설비 점유 안에서 연속 실행한다.
def run_main_bevel_group(env, jNo, line_name, step_no, cycle_dict, group_no, include_sen2=True, step_runner=None):
    """Main flow bevel machine group.

    group 1: Sen1 -> Bevel1 -> Sen2
    group 2: Sen1 -> Bevel2 -> Sen2

    한 번 group이 시작되면 해당 머신을 계속 점유한 채 내부 공정들을 바로 이어서 수행한다.
    따라서 다른 Job이 Sen1/Bevel/Sen2 사이에 끼어들 수 없다.
    """
    if group_no not in MAIN_BEVEL_GROUP_RESOURCE:
        raise ValueError(f"알 수 없는 main bevel group 번호: {group_no}")

    resource_name = MAIN_BEVEL_GROUP_RESOURCE[group_no]
    proc_names = ["Sen1", f"Bevel{group_no}"]
    if include_sen2:
        proc_names.append("Sen2")

    req = resources[resource_name].request(priority=int(jNo))
    yield req
    try:
        for idx, proc_name in enumerate(proc_names):
            ct = _require_defined_ct(jNo, line_name, proc_name, cycle_dict,
                                     get_process_cycle_time(jNo, line_name, proc_name, cycle_dict))
            enforce_prev_step = (idx == 0)
            if step_runner is None:
                yield env.process(run_line_step(
                    env, line_name, jNo, step_no,
                    process_named_on_held_resource(env, jNo, line_name, proc_name, ct),
                    enforce_prev_step=enforce_prev_step,
                ))
            else:
                yield env.process(step_runner(step_no, proc_name, ct, enforce_prev_step))
            step_no += 1
    finally:
        resources[resource_name].release(req)

    return step_no


# 초보자 설명: 로봇의 순수 이동/인계 단계만 실행한다.
def run_robot_transport_only(env, jNo, line_name, step_no, cycle_dict, robot_name, next_proc=None, from_gantry=False,
                             enforce_prev_step=True):
    next_proc = norm_text(next_proc)
    display_name = robot_name
    ct = None
    if robot_name == "R2" and next_proc == "2D-1":
        ct = _ctx_value(cycle_dict, "r2_fit_2d1", "2D-1", ("2D-1",))
    elif robot_name == "R2" and next_proc == "2D-2":
        ct = _ctx_value(cycle_dict, "r2_fit_2d2", "2D-2", ("2D-2",))
    elif robot_name == "R3" and next_proc == "3D":
        ct = _ctx_value(cycle_dict, "r3_fit_3d", "3D", ("3D",))
    if ct is None:
        display_name, ct = get_robot_step_timing(jNo, line_name, robot_name, next_proc, from_gantry, cycle_dict, False)
    ct = _require_defined_ct(jNo, line_name, robot_name, cycle_dict, ct,
                             detail=f"next_proc={next_proc}, from_gantry={from_gantry}, tig=False")
    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        step_no,
        process_robot_with_display(env, jNo, line_name, robot_name, display_name, ct),
        enforce_prev_step=enforce_prev_step,
    ))


# 초보자 설명: Gan 자원을 사용하지만 표시 이름을 따로 주고 싶은 경우에 쓰는 래퍼다.
def run_named_gan_step(env, jNo, line_name, step_no, cycle_dict, display_name, ct, enforce_prev_step=True):
    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        step_no,
        process_named_with_resource(env, jNo, line_name, "Gan", display_name, ct),
        enforce_prev_step=enforce_prev_step,
    ))


# 초보자 설명: 2D main 진입 시 필요한 Gan 시간을 구한다.
def _get_2d_main_entry_gan(job, line, cycle_dict, weld):
    stage = norm_text(weld)
    if stage == "2D-1":
        val = _ctx_value(cycle_dict, "gan1_2d1", "2D-1", ("2D-1",))
        return _require_defined_ct(job, line, "Gan", cycle_dict, val)
    if stage == "2D-2":
        val = _ctx_value(cycle_dict, "gan1_2d2", "2D-2", ("2D-2",))
        return _require_defined_ct(job, line, "Gan", cycle_dict, val)
    val = _ctx_value(cycle_dict, "gan1_3d", "3D", ("3D",))
    return _require_defined_ct(job, line, "Gan", cycle_dict, val)


# 초보자 설명: 2D main 종료 배출 시 필요한 Gan 시간을 구한다.
def _get_2d_main_final_gan(job, line, cycle_dict, weld):
    stage = norm_text(weld)
    if stage == "2D-1":
        val = _ctx_value(cycle_dict, "out_2d1", "2D-1", ("2D-1",))
    elif stage == "2D-2":
        val = _ctx_value(cycle_dict, "out_2d2", "2D-2", ("2D-2",))
    else:
        val = _ctx_value(cycle_dict, "out_3d", "3D", ("3D",))
    return _require_defined_ct(job, line, "Gan", cycle_dict, val)


# 초보자 설명: 2D 계열의 main 용접 구간 전체를 표준 순서로 실행한다.
def run_2d_main_weld_sequence(env, jNo, pipe_step, cycle_dict, weld, fit_ready_evt, entry_mode):
    stage = norm_text(weld)
    if stage not in ["2D-1", "2D-2", "3D"]:
        alarm("ALARM 3 : 용접장 미지정")

    line_name = "Pipe"
    robot_name = "R2" if stage in ["2D-1", "2D-2"] else "R3"
    robot_display = "R2tig" if robot_name == "R2" else "R3tig"
    robot_req = None

    deferred_r2_entry_ct = None
    if entry_mode == "short_r2":
        if stage == "2D-1":
            deferred_r2_entry_ct = _ctx_value(cycle_dict, "r2_2d1", "2D-1", ("2D-1",))
        elif stage == "2D-2":
            deferred_r2_entry_ct = _ctx_value(cycle_dict, "r2_2d2", "2D-2", ("2D-2",))
        else:
            alarm("ALARM 3 : 용접장 미지정")
        deferred_r2_entry_ct = _require_defined_ct(jNo, line_name, "R2", cycle_dict, deferred_r2_entry_ct)
    elif entry_mode == "gantry":
        gan_ct = _get_2d_main_entry_gan(jNo, line_name, cycle_dict, stage)
        yield env.process(run_named_gan_step(env, jNo, line_name, pipe_step, cycle_dict, "Gan", gan_ct))
        pipe_step += 1
    else:
        raise RuntimeError(f"알 수 없는 2D main entry_mode: {entry_mode}")

    if not fit_ready_evt.triggered:
        yield fit_ready_evt

    pos = None
    if stage in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage, jNo)

    if entry_mode == "short_r2":
        robot_req = resources[robot_name].request()
        yield robot_req
        yield env.process(run_line_step(
            env,
            line_name,
            jNo,
            pipe_step,
            process_named_on_held_resource(env, jNo, line_name, robot_name, deferred_r2_entry_ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    if robot_req is None:
        robot_req = resources[robot_name].request()
        yield robot_req

    fit_ct = _require_defined_ct(jNo, line_name, "Fit", cycle_dict,
                                 _ctx_value(cycle_dict, "fitup_spf", stage, (stage,)))
    tig1_ct = _require_defined_ct(jNo, line_name, robot_display, cycle_dict,
                                  _ctx_value(cycle_dict, "tig_spf", stage, (stage,)))
    tig2_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                  _ctx_value(cycle_dict, "tig2_spf", stage, (stage,)))
    mag_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict, _ctx_value(cycle_dict, "mag_spf", stage, (stage,)))
    final_gan_ct = _get_2d_main_final_gan(jNo, line_name, cycle_dict, stage)

    station_req = resources[stage].request()
    yield station_req

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Fit", fit_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, robot_display, tig1_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources[robot_name].release(robot_req)

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage, pos)
    resources[stage].release(station_req)

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", final_gan_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step


# 초보자 설명: 상황에 따라 일반 로봇 이동인지, 로봇+tig 결합인지 판단해 실행한다.
def run_robot_step(env, jNo, line_name, step_no, cycle_dict, robot_name, next_proc=None, from_gantry=False,
                   enforce_prev_step=True, combine_weld_hold=None):
    next_proc = norm_text(next_proc)
    is_weld_target = next_proc in WELD_PROCS
    is_main_line = line_name in {"Pipe", "Pipe2"}
    is_sub_line = line_name in {"Sub", "Sub2"}
    use_tig_display = bool(is_sub_line and robot_name in ROBOT_TIG_DISPLAY and is_weld_target)

    display_name, ct = get_robot_step_timing(jNo, line_name, robot_name, next_proc, from_gantry, cycle_dict,
                                             use_tig_display)

    if is_sub_line and display_name in ROBOT_TIG_DISPLAY.values() and is_weld_target:
        release_line_active_hold(jNo, line_name)
        yield from wait_for_main_weld_arrival(env, jNo, next_proc)
        use_combined_weld_hold = COMBINED_WELD_HOLD_ENABLED if combine_weld_hold is None else bool(combine_weld_hold)
        if use_combined_weld_hold:
            yield env.process(run_line_step(
                env,
                line_name,
                jNo,
                step_no,
                start_combined_weld_hold(env, jNo, line_name, display_name, next_proc, ct),
                enforce_prev_step=enforce_prev_step if not is_sub_line else False,
            ))
        else:
            yield env.process(run_line_step(
                env,
                line_name,
                jNo,
                step_no,
                process_robot_with_display(env, jNo, line_name, robot_name, display_name, ct),
                enforce_prev_step=enforce_prev_step if not is_sub_line else False,
            ))
        return

    yield env.process(run_line_step(env, line_name, jNo, step_no,
                                    process_robot_with_display(env, jNo, line_name, robot_name, display_name, ct),
                                    enforce_prev_step=enforce_prev_step))

    if is_main_line and is_weld_target:
        mark_main_weld_arrival(jNo, next_proc)


# 초보자 설명: Pipe2 라인의 일반 공정을 실행하는 얇은 래퍼다.
def run_pipe2_step(env, job, step_no, name, cycle_dict):
    yield env.process(run_line_step(env, "Pipe2", job, step_no, process(env, job, "Pipe2", name, cycle_dict)))


# 초보자 설명: Pipe2 라인에서 CutBuf를 거친 공정을 실행한다.
def run_pipe2_after_cutbuf_step(env, job, step_no, name, cycle_dict):
    yield env.process(
        run_line_step(env, "Pipe2", job, step_no, process_after_cutbuf(env, job, "Pipe2", name, cycle_dict)))


# 초보자 설명: Pipe2 라인의 Gan 공정을 실행한다.
def run_pipe2_gan_step(env, job, step_no, cycle_dict, gan_key):
    prev_gan_override = dyn_ct._override.get("Gan", None)
    dyn_ct.set("Gan", get_gan_step_cycle_time(job, "Pipe2", gan_key, cycle_dict))
    yield env.process(run_line_step(env, "Pipe2", job, step_no, process(env, job, "Pipe2", "Gan", cycle_dict)))
    if prev_gan_override is not None:
        dyn_ct.set("Gan", prev_gan_override)
    else:
        dyn_ct.set("Gan", None)


# =====================================================
# 제품 타입별 생산 라우팅(run_... 함수들)
# -----------------------------------------------------
# 아래 함수들은 제품 타입별로 실제 공정 순서를 적어 놓은 부분이다.
# 초보자는 이 구간을 '공정 흐름도 코드 버전'이라고 생각하면 이해하기 쉽다.
# =====================================================

# 초보자 설명: 짧은 2D 제품의 Pipe(main) 라인 흐름을 실행한다.
def run_short_2d_pipe(env, jNo, pipe_step, cycle_dict, weld, fit_ready_evt):
    bevel1_type, bevel2_type = get_2d_bevel_types(weld)
    yield env.process(
        run_line_step(env, "Pipe", jNo, pipe_step, process_after_cutbuf(env, jNo, "Pipe", "Con1", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, "Pipe", pipe_step, cycle_dict, 1, include_sen2=(bevel1_type == "J"))
    )
    yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Con2", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, "Pipe", pipe_step, cycle_dict, 2, include_sen2=(bevel2_type == "J"))
    )
    yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Conv3", cycle_dict)))
    pipe_step += 1
    entry_mode = "gantry" if norm_text(weld) == "3D" else "short_r2"
    pipe_step = yield env.process(
        run_2d_main_weld_sequence(env, jNo, pipe_step, cycle_dict, weld, fit_ready_evt, entry_mode))
    return pipe_step

# 초보자 설명: 짧은 2D 제품의 Sub 라인 흐름을 실행한다.
def run_short_2d_sub(env, jNo, sub_step, cycle_dict, weld, prod, fit_ready_evt, fit_bevel_proc):
    yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                    process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                    process_optional_future_step(env, jNo, "Sub", "FittingBuffer", cycle_dict)));
    sub_step += 1
    if prod != "2D":
        alarm("ALARM 4 : 스풀타입 미지정")
    if weld in ["2D-1", "2D-2"]:
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", weld, from_gantry=False));
        sub_step += 1
    elif weld == "3D":
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R3", "3D", from_gantry=False));
        sub_step += 1
    else:
        alarm("ALARM 3 : 용접장 미지정")
    if not fit_ready_evt.triggered:
        fit_ready_evt.succeed()
    return sub_step


# 초보자 설명: F2D 제품의 main 라인 공정 순서를 실행한다.
def run_f2d_main(env, jNo, pipe_step, cycle_dict, length, fit_ready_2d1_evt, fit_ready_3d_evt):
    """F2D main flow."""
    line_name = "Pipe"
    stage_2d1 = "2D-1"
    stage_3d = "3D"
    bevel1_type, bevel2_type = get_2d_bevel_types(stage_2d1)

    yield env.process(
        run_line_step(env, line_name, jNo, pipe_step, process_after_cutbuf(env, jNo, line_name, "Con1", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 1, include_sen2=(bevel1_type == "J"))
    )
    yield env.process(run_line_step(env, line_name, jNo, pipe_step, process(env, jNo, line_name, "Con2", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 2, include_sen2=(bevel2_type == "J"))
    )
    yield env.process(run_line_step(env, line_name, jNo, pipe_step, process(env, jNo, line_name, "Conv3", cycle_dict)))
    pipe_step += 1

    r2_req = None
    deferred_r2_entry_ct = None
    if length is not None and float(length) < 1700:
        deferred_r2_entry_ct = _require_defined_ct(jNo, line_name, "R2", cycle_dict,
                                                   _ctx_value(cycle_dict, "r2_2d1", stage_2d1, (stage_2d1,)))
    else:
        gan_entry_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                           _ctx_value(cycle_dict, "gan1_2d1", stage_2d1, (stage_2d1,)))
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_entry_ct),
        ))
        pipe_step += 1

    if not fit_ready_2d1_evt.triggered:
        yield fit_ready_2d1_evt

    pos = None
    if stage_2d1 in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_2d1, jNo)

    if deferred_r2_entry_ct is not None:
        r2_req = resources["R2"].request()
        yield r2_req
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, "R2", deferred_r2_entry_ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    if r2_req is None:
        r2_req = resources["R2"].request()
        yield r2_req

    fit_ct = _require_defined_ct(jNo, line_name, "Fit", cycle_dict,
                                 _ctx_value(cycle_dict, "fitup_spf", stage_2d1, (stage_2d1,)))
    r2tig_ct = _require_defined_ct(jNo, line_name, "R2tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_2d1, (stage_2d1,)))
    tig2_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                  _ctx_value(cycle_dict, "tig2_spf", stage_2d1, (stage_2d1,)))
    mag_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                 _ctx_value(cycle_dict, "mag_spf", stage_2d1, (stage_2d1,)))

    station_req = resources[stage_2d1].request()
    yield station_req

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Fit", fit_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R2tig", r2tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R2"].release(r2_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_2d1 in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_2d1, pos)
    resources[stage_2d1].release(station_req)

    gan_2d1_to_3d_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                           _ctx_value(cycle_dict, "gan4", stage_3d, (stage_3d,)))
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_2d1_to_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if not fit_ready_3d_evt.triggered:
        yield fit_ready_3d_evt

    angle_ct = _require_defined_ct(jNo, line_name, "Angle", cycle_dict,
                                   _ctx_value(cycle_dict, "fitup_3d", stage_3d, (stage_3d,)))
    r3tig_ct = _require_defined_ct(jNo, line_name, "R3tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_3d, (stage_3d,)))
    tig2_3d_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                     _ctx_value(cycle_dict, "tig2_spf", stage_3d, (stage_3d,)))
    mag_3d_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                    _ctx_value(cycle_dict, "mag_spf", stage_3d, (stage_3d,)))
    gan_out_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                     _ctx_value(cycle_dict, "out_3d", stage_3d, (stage_3d,)))

    pos = None
    if stage_3d in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_3d, jNo)
    r3_req = resources["R3"].request()
    yield r3_req
    station_req = resources[stage_3d].request()
    yield station_req

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Angle", angle_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R3tig", r3tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R3"].release(r3_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_3d in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_3d, pos)
    resources[stage_3d].release(station_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step

# 초보자 설명: F2D 제품의 sub 라인 공정 순서를 실행한다.
def run_f2d_sub(env, jNo, sub_step, cycle_dict, spool_name, fallback_ft, fit_ready_2d1_evt, fit_ready_3d_evt):
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)
    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-1", from_gantry=False));
    sub_step += 1
    if not fit_ready_2d1_evt.triggered:
        fit_ready_2d1_evt.succeed()
    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting(env, jNo, "Sub", fit2_proc, cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R3", "3D", from_gantry=False));
    sub_step += 1
    if not fit_ready_3d_evt.triggered:
        fit_ready_3d_evt.succeed()
    return sub_step


# 초보자 설명: F2D의 두 Sub 흐름을 병렬로 실행한다.
def run_f2d_sub_parallel(env, jNo, cycle_dict, spool_name, fallback_ft, fit_ready_2d1_evt, fit_ready_3d_evt):
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)
    lane1_done = env.event()

    # 초보자 설명: 'lane1' 단계의 처리를 맡는 보조 함수다.
    def lane1():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-1", from_gantry=False));
        sub_step += 1
        if not fit_ready_2d1_evt.triggered:
            fit_ready_2d1_evt.succeed()
        release_remaining_line_steps("Sub", jNo, sub_step - 1)
        if not lane1_done.triggered:
            lane1_done.succeed()

    # 초보자 설명: 'lane2' 단계의 처리를 맡는 보조 함수다.
    def lane2():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub2", fit2_proc, cycle_dict, 2)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub2", sub_step, cycle_dict, "R3", "3D", from_gantry=False));
        sub_step += 1
        if not fit_ready_3d_evt.triggered:
            fit_ready_3d_evt.succeed()
        release_remaining_line_steps("Sub2", jNo, sub_step - 1)

    env.process(lane1())
    env.process(lane2())
    yield lane1_done


# 초보자 설명: 긴 2D 제품의 Pipe(main) 라인 흐름을 실행한다.
def run_long_2d_pipe(env, jNo, pipe_step, cycle_dict, weld, fit_ready_evt):
    bevel1_type, bevel2_type = get_2d_bevel_types(weld)
    yield env.process(
        run_line_step(env, "Pipe", jNo, pipe_step, process_after_cutbuf(env, jNo, "Pipe", "Con1", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, "Pipe", pipe_step, cycle_dict, 1, include_sen2=(bevel1_type == "J"))
    )
    yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Con2", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, "Pipe", pipe_step, cycle_dict, 2, include_sen2=(bevel2_type == "J"))
    )
    yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Conv3", cycle_dict)))
    pipe_step += 1
    pipe_step = yield env.process(
        run_2d_main_weld_sequence(env, jNo, pipe_step, cycle_dict, weld, fit_ready_evt, "gantry"))
    return pipe_step

# 초보자 설명: 긴 2D 제품의 Sub 라인 흐름을 실행한다.
def run_long_2d_sub(env, jNo, sub_step, cycle_dict, weld, prod, fit_ready_evt, fit_bevel_proc):
    yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                    process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                    process_optional_future_step(env, jNo, "Sub", "FitBevel", cycle_dict)));
    sub_step += 1
    if prod == "2D":
        if weld in ["2D-1", "2D-2"]:
            yield env.process(
                run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", weld, from_gantry=False));
            sub_step += 1
        elif weld == "3D":
            yield env.process(
                run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R3", "3D", from_gantry=False));
            sub_step += 1
        else:
            alarm("ALARM 3 : 용접장 미지정")
    else:
        alarm("ALARM 4 : 스풀타입 미지정")
    if not fit_ready_evt.triggered:
        fit_ready_evt.succeed()
    return sub_step


# 초보자 설명: H3D 제품의 main/Pipe2 연동 공정을 실행한다.
def run_h3d_main(env, jNo, pipe_step, cycle_dict, h3d_sub_ready_2d2_evt, h3d_pipe2_ready_3d_evt):
    """H3D main flow."""
    line_name = "Pipe"
    stage_2d2 = "2D-2"
    stage_3d = "3D"
    bevel1_type, bevel2_type = get_2d_bevel_types(stage_2d2)
    bevel1_type_3d, bevel2_type_3d = get_2d_bevel_types(stage_3d)

    shared_occ_template = {
        "Cut": 1,
        "Con1": 1,
        "Sen1": 2,
        "Bevel1": 1,
        "Con2": 1,
        "Bevel2": 1,
        "Sen2": 1,
        "Conv3": 1,
    }
    shared_done = {
        (proc, occ): env.event()
        for proc, max_occ in shared_occ_template.items()
        for occ in range(1, max_occ + 1)
    }
    shared_done[("Cut", 1)].succeed()
    pipe_shared_occ = defaultdict(int)
    pipe2_shared_occ = defaultdict(int)

    # 초보자 설명: 여러 세부 공정을 순서대로 호출해 한 흐름을 실행하는 함수다.
    def run_pipe_shared(step_no, proc_name, proc_gen):
        pipe_shared_occ[proc_name] += 1
        occ = pipe_shared_occ[proc_name]
        yield env.process(run_line_step(env, line_name, jNo, step_no, proc_gen))
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            evt.succeed()

    # 초보자 설명: 여러 세부 공정을 순서대로 호출해 한 흐름을 실행하는 함수다.
    def run_pipe2_shared(step_no, proc_name, proc_gen):
        pipe2_shared_occ[proc_name] += 1
        occ = pipe2_shared_occ[proc_name]
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            yield evt
        yield env.process(run_line_step(env, "Pipe2", jNo, step_no, proc_gen))

    # 초보자 설명: 'pipe_group_runner' 단계의 처리를 맡는 보조 함수다.
    def pipe_group_runner(step_no, proc_name, ct, enforce_prev_step):
        pipe_shared_occ[proc_name] += 1
        occ = pipe_shared_occ[proc_name]
        yield env.process(run_line_step(
            env, line_name, jNo, step_no,
            process_named_on_held_resource(env, jNo, line_name, proc_name, ct),
            enforce_prev_step=enforce_prev_step,
        ))
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            evt.succeed()

    # 초보자 설명: 'pipe2_group_runner' 단계의 처리를 맡는 보조 함수다.
    def pipe2_group_runner(step_no, proc_name, ct, enforce_prev_step):
        pipe2_shared_occ[proc_name] += 1
        occ = pipe2_shared_occ[proc_name]
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            yield evt
        yield env.process(run_line_step(
            env, "Pipe2", jNo, step_no,
            process_named_on_held_resource(env, jNo, "Pipe2", proc_name, ct),
            enforce_prev_step=enforce_prev_step,
        ))

    # 초보자 설명: 'h3d_pipe2_flow' 단계의 처리를 맡는 보조 함수다.
    def h3d_pipe2_flow(start_step):
        pipe2_step = start_step
        yield env.process(run_pipe2_shared(
            pipe2_step, "Cut",
            process(env, jNo, "Pipe2", "Cut", cycle_dict),
        ))
        pipe2_step += 1
        yield CutBuf.put(1)

        yield env.process(run_pipe2_shared(
            pipe2_step, "Con1",
            process_after_cutbuf(env, jNo, "Pipe2", "Con1", cycle_dict),
        ))
        pipe2_step += 1
        pipe2_step = yield env.process(
            run_main_bevel_group(env, jNo, "Pipe2", pipe2_step, cycle_dict, 1,
                                 include_sen2=(bevel1_type_3d == "J"), step_runner=pipe2_group_runner)
        )
        yield env.process(run_pipe2_shared(
            pipe2_step, "Con2",
            process(env, jNo, "Pipe2", "Con2", cycle_dict),
        ))
        pipe2_step += 1
        pipe2_step = yield env.process(
            run_main_bevel_group(env, jNo, "Pipe2", pipe2_step, cycle_dict, 2,
                                 include_sen2=(bevel2_type_3d == "J"), step_runner=pipe2_group_runner)
        )
        yield env.process(run_pipe2_shared(
            pipe2_step, "Conv3",
            process(env, jNo, "Pipe2", "Conv3", cycle_dict),
        ))
        pipe2_step += 1

        gan_pipe2_ct = _require_defined_ct(jNo, "Pipe2", "Gan", cycle_dict,
                                           _ctx_value(cycle_dict, "gan1_3d", stage_3d, (stage_3d,)))
        yield env.process(run_line_step(
            env, "Pipe2", jNo, pipe2_step,
            process_named_with_resource(env, jNo, "Pipe2", "Gan", "Gan", gan_pipe2_ct),
        ))
        pipe2_step += 1
        if not h3d_pipe2_ready_3d_evt.triggered:
            h3d_pipe2_ready_3d_evt.succeed()
        release_remaining_line_steps("Pipe2", jNo, pipe2_step - 1)

    env.process(h3d_pipe2_flow(pipe_step + 1))

    yield env.process(run_pipe_shared(
        pipe_step, "Con1",
        process_after_cutbuf(env, jNo, line_name, "Con1", cycle_dict),
    ))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 1,
                             include_sen2=(bevel1_type == "J"), step_runner=pipe_group_runner)
    )
    yield env.process(run_pipe_shared(
        pipe_step, "Con2",
        process(env, jNo, line_name, "Con2", cycle_dict),
    ))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 2,
                             include_sen2=(bevel2_type == "J"), step_runner=pipe_group_runner)
    )
    yield env.process(run_pipe_shared(
        pipe_step, "Conv3",
        process(env, jNo, line_name, "Conv3", cycle_dict),
    ))
    pipe_step += 1

    r2_entry_ct = _require_defined_ct(jNo, line_name, "R2", cycle_dict,
                                      _ctx_value(cycle_dict, "r2_2d2", stage_2d2, (stage_2d2,)))

    if not h3d_sub_ready_2d2_evt.triggered:
        yield h3d_sub_ready_2d2_evt

    pos = None
    if stage_2d2 in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_2d2, jNo)

    r2_req = resources["R2"].request()
    yield r2_req
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R2", r2_entry_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    fit_ct = _require_defined_ct(jNo, line_name, "Fit", cycle_dict,
                                 _ctx_value(cycle_dict, "fitup_spf", stage_2d2, (stage_2d2,)))
    r2tig_ct = _require_defined_ct(jNo, line_name, "R2tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_2d2, (stage_2d2,)))
    tig2_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                  _ctx_value(cycle_dict, "tig2_spf", stage_2d2, (stage_2d2,)))
    mag_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                 _ctx_value(cycle_dict, "mag_spf", stage_2d2, (stage_2d2,)))

    station_req = resources[stage_2d2].request()
    yield station_req

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Fit", fit_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R2tig", r2tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R2"].release(r2_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_2d2 in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_2d2, pos)
    resources[stage_2d2].release(station_req)

    gan_to_3d_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                       _ctx_value(cycle_dict, "gan1_3d", stage_3d, (stage_3d,)))
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_to_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    r3_move_ct = _require_defined_ct(jNo, line_name, "R3", cycle_dict,
                                     _ctx_value(cycle_dict, "r3_2d2_to_3d", stage_3d, (stage_3d,)))
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "R3", "R3", r3_move_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if not h3d_pipe2_ready_3d_evt.triggered:
        yield h3d_pipe2_ready_3d_evt

    angle_ct = _require_defined_ct(jNo, line_name, "Angle", cycle_dict,
                                   _ctx_value(cycle_dict, "fitup_3d", stage_3d, (stage_3d,)))
    r3tig_ct = _require_defined_ct(jNo, line_name, "R3tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_3d, (stage_3d,)))
    tig2_3d_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                     _ctx_value(cycle_dict, "tig2_spf", stage_3d, (stage_3d,)))
    mag_3d_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                    _ctx_value(cycle_dict, "mag_spf", stage_3d, (stage_3d,)))
    gan_out_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                     _ctx_value(cycle_dict, "out_3d", stage_3d, (stage_3d,)))

    pos = None
    if stage_3d in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_3d, jNo)
    r3_hold_req = resources["R3"].request()
    yield r3_hold_req
    station_req = resources[stage_3d].request()
    yield station_req

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Angle", angle_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R3tig", r3tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R3"].release(r3_hold_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_3d in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_3d, pos)
    resources[stage_3d].release(station_req)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step

# 초보자 설명: H3D 제품의 Sub 라인 공정을 실행한다.
def run_h3d_sub(env, jNo, sub_step, cycle_dict, fit_bevel_proc, h3d_sub_ready_2d2_evt):
    yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                    process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-2", from_gantry=False));
    sub_step += 1
    if not h3d_sub_ready_2d2_evt.triggered:
        h3d_sub_ready_2d2_evt.succeed()
    return sub_step


# 초보자 설명: F3D 제품의 main/Pipe2 연동 공정을 실행한다.
def run_f3d_main(env, jNo, pipe_step, cycle_dict, length, f3d_fit1_ready_2d1_evt, f3d_fit2_ready_3d_evt):
    """F3D main flow."""
    line_name = "Pipe"
    stage_2d1 = "2D-1"
    stage_pipe2 = "2D-2"
    stage_3d = "3D"

    pipe2_start_evt = env.event()
    pipe2_ready_3d_evt = env.event()

    shared_occ_template = {
        "Cut": 1,
        "Con1": 1,
        "Sen1": 2,
        "Bevel1": 1,
        "Con2": 1,
        "Bevel2": 1,
        "Sen2": 1,
        "Conv3": 1,
    }
    shared_done = {
        (proc, occ): env.event()
        for proc, max_occ in shared_occ_template.items()
        for occ in range(1, max_occ + 1)
    }
    shared_done[("Cut", 1)].succeed()
    pipe_shared_occ = defaultdict(int)
    pipe2_shared_occ = defaultdict(int)

    # 초보자 설명: 여러 세부 공정을 순서대로 호출해 한 흐름을 실행하는 함수다.
    def run_pipe_shared(step_no, proc_name, proc_gen):
        pipe_shared_occ[proc_name] += 1
        occ = pipe_shared_occ[proc_name]
        yield env.process(run_line_step(env, line_name, jNo, step_no, proc_gen))
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            evt.succeed()

    # 초보자 설명: 여러 세부 공정을 순서대로 호출해 한 흐름을 실행하는 함수다.
    def run_pipe2_shared(step_no, proc_name, proc_gen):
        pipe2_shared_occ[proc_name] += 1
        occ = pipe2_shared_occ[proc_name]
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            yield evt
        yield env.process(run_line_step(env, "Pipe2", jNo, step_no, proc_gen))

    # 초보자 설명: 'pipe_group_runner' 단계의 처리를 맡는 보조 함수다.
    def pipe_group_runner(step_no, proc_name, ct, enforce_prev_step):
        pipe_shared_occ[proc_name] += 1
        occ = pipe_shared_occ[proc_name]
        yield env.process(run_line_step(
            env, line_name, jNo, step_no,
            process_named_on_held_resource(env, jNo, line_name, proc_name, ct),
            enforce_prev_step=enforce_prev_step,
        ))
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            evt.succeed()

    # 초보자 설명: 'pipe2_group_runner' 단계의 처리를 맡는 보조 함수다.
    def pipe2_group_runner(step_no, proc_name, ct, enforce_prev_step):
        pipe2_shared_occ[proc_name] += 1
        occ = pipe2_shared_occ[proc_name]
        evt = shared_done[(proc_name, occ)]
        if not evt.triggered:
            yield evt
        yield env.process(run_line_step(
            env, "Pipe2", jNo, step_no,
            process_named_on_held_resource(env, jNo, "Pipe2", proc_name, ct),
            enforce_prev_step=enforce_prev_step,
        ))

    # 초보자 설명: 'pipe2_flow' 단계의 처리를 맡는 보조 함수다.
    def pipe2_flow(start_step):
        pipe2_step = start_step
        if not pipe2_start_evt.triggered:
            yield pipe2_start_evt

        yield env.process(run_pipe2_shared(
            pipe2_step, "Cut",
            process(env, jNo, "Pipe2", "Cut", cycle_dict),
        ))
        pipe2_step += 1
        yield CutBuf.put(1)

        yield env.process(run_pipe2_shared(
            pipe2_step, "Con1",
            process_after_cutbuf(env, jNo, "Pipe2", "Con1", cycle_dict),
        ))
        pipe2_step += 1
        pipe2_step = yield env.process(
            run_main_bevel_group(env, jNo, "Pipe2", pipe2_step, cycle_dict, 1,
                                 include_sen2=False, step_runner=pipe2_group_runner)
        )
        yield env.process(run_pipe2_shared(
            pipe2_step, "Con2",
            process(env, jNo, "Pipe2", "Con2", cycle_dict),
        ))
        pipe2_step += 1
        pipe2_step = yield env.process(
            run_main_bevel_group(env, jNo, "Pipe2", pipe2_step, cycle_dict, 2,
                                 include_sen2=True, step_runner=pipe2_group_runner)
        )
        yield env.process(run_pipe2_shared(
            pipe2_step, "Conv3",
            process(env, jNo, "Pipe2", "Conv3", cycle_dict),
        ))
        pipe2_step += 1

        r2_scissors_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "R2",
            cycle_dict,
            _ctx_value(cycle_dict, "r2_2d2", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: Scissors -> S2 합류 지점",
        )
        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_with_resource(env, jNo, "Pipe2", "R2", "R2", r2_scissors_ct),
        ))
        pipe2_step += 1

        if not f3d_fit2_ready_3d_evt.triggered:
            yield f3d_fit2_ready_3d_evt

        pos = None
        if stage_pipe2 in STRICT_STATION_FIFO:
            _occ_no, pos = yield from wait_prev_station_fifo(env, stage_pipe2, jNo)

        r2_req = resources["R2"].request()
        yield r2_req

        r2_join_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "R2",
            cycle_dict,
            _ctx_value(cycle_dict, "r2_fit_2d2", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: S2 합류 지점 -> 2D-2 Fit",
        )
        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_on_held_resource(env, jNo, "Pipe2", "R2", r2_join_ct),
            enforce_prev_step=False,
        ))
        pipe2_step += 1

        fit2d2_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "Fit",
            cycle_dict,
            _ctx_value(cycle_dict, "fitup_spf", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: 2D-2 Fit",
        )
        r2tig2d2_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "R2tig",
            cycle_dict,
            _ctx_value(cycle_dict, "tig_spf", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: 2D-2 R2Tig1",
        )
        tig22d2_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "Tig2",
            cycle_dict,
            _ctx_value(cycle_dict, "tig2_spf", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: 2D-2 Tig2",
        )
        mag2d2_ct = _require_defined_ct(
            jNo,
            "Pipe2",
            "Mag",
            cycle_dict,
            _ctx_value(cycle_dict, "mag_spf", stage_pipe2, (stage_pipe2,)),
            detail="F3D M2 line: 2D-2 Mag",
        )

        station_req = resources[stage_pipe2].request()
        yield station_req

        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_on_held_resource(env, jNo, "Pipe2", "Fit", fit2d2_ct),
            enforce_prev_step=False,
        ))
        pipe2_step += 1

        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_on_held_resource(env, jNo, "Pipe2", "R2tig", r2tig2d2_ct),
            enforce_prev_step=False,
        ))
        pipe2_step += 1

        resources["R2"].release(r2_req)

        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_on_held_resource(env, jNo, "Pipe2", "Tig2", tig22d2_ct),
            enforce_prev_step=False,
        ))
        pipe2_step += 1

        yield env.process(run_line_step(
            env,
            "Pipe2",
            jNo,
            pipe2_step,
            process_named_on_held_resource(env, jNo, "Pipe2", "Mag", mag2d2_ct),
            enforce_prev_step=False,
        ))
        pipe2_step += 1

        if stage_pipe2 in STRICT_STATION_FIFO:
            mark_station_fifo_done(stage_pipe2, pos)
        resources[stage_pipe2].release(station_req)

        if not pipe2_ready_3d_evt.triggered:
            pipe2_ready_3d_evt.succeed()
        release_remaining_line_steps("Pipe2", jNo, pipe2_step - 1)

    env.process(pipe2_flow(pipe_step + 1))

    yield env.process(run_pipe_shared(
        pipe_step, "Con1",
        process_after_cutbuf(env, jNo, line_name, "Con1", cycle_dict),
    ))
    pipe_step += 1
    if not pipe2_start_evt.triggered:
        pipe2_start_evt.succeed()

    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 1,
                             include_sen2=True, step_runner=pipe_group_runner)
    )
    yield env.process(run_pipe_shared(
        pipe_step, "Con2",
        process(env, jNo, line_name, "Con2", cycle_dict),
    ))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 2,
                             include_sen2=False, step_runner=pipe_group_runner)
    )
    yield env.process(run_pipe_shared(
        pipe_step, "Conv3",
        process(env, jNo, line_name, "Conv3", cycle_dict),
    ))
    pipe_step += 1

    short_entry_2d1 = bool(length is not None and float(length) < 1700)
    deferred_r2_entry_ct = None
    r2_req = None

    if short_entry_2d1:
        deferred_r2_entry_ct = _require_defined_ct(
            jNo,
            line_name,
            "R2",
            cycle_dict,
            _ctx_value(cycle_dict, "r2_2d1", stage_2d1, (stage_2d1,)),
            detail="F3D M1 short entry: Scissors -> 2D-1",
        )
    else:
        gan_entry_ct = _require_defined_ct(
            jNo,
            line_name,
            "Gan",
            cycle_dict,
            _ctx_value(cycle_dict, "gan1_2d1", stage_2d1, (stage_2d1,)),
            detail="F3D M1 long entry: Scissors -> 2D-1",
        )
        yield env.process(run_line_step(
            env,
            line_name,
            jNo,
            pipe_step,
            process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_entry_ct),
        ))
        pipe_step += 1

    if not f3d_fit1_ready_2d1_evt.triggered:
        yield f3d_fit1_ready_2d1_evt

    pos = None
    if stage_2d1 in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_2d1, jNo)

    if deferred_r2_entry_ct is not None:
        r2_req = resources["R2"].request()
        yield r2_req
        yield env.process(run_line_step(
            env,
            line_name,
            jNo,
            pipe_step,
            process_named_on_held_resource(env, jNo, line_name, "R2", deferred_r2_entry_ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    if r2_req is None:
        r2_req = resources["R2"].request()
        yield r2_req

    fit_ct = _require_defined_ct(jNo, line_name, "Fit", cycle_dict,
                                 _ctx_value(cycle_dict, "fitup_spf", stage_2d1, (stage_2d1,)))
    r2tig_ct = _require_defined_ct(jNo, line_name, "R2tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_2d1, (stage_2d1,)))
    tig2_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                  _ctx_value(cycle_dict, "tig2_spf", stage_2d1, (stage_2d1,)))
    mag_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                 _ctx_value(cycle_dict, "mag_spf", stage_2d1, (stage_2d1,)))

    station_req = resources[stage_2d1].request()
    yield station_req

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Fit", fit_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R2tig", r2tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R2"].release(r2_req)

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_2d1 in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_2d1, pos)
    resources[stage_2d1].release(station_req)

    gan_to_3d_ct = _require_defined_ct(
        jNo,
        line_name,
        "Gan",
        cycle_dict,
        _ctx_value(cycle_dict, "gan4", stage_3d, (stage_3d,)),
    )
    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_to_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if not pipe2_ready_3d_evt.triggered:
        yield pipe2_ready_3d_evt

    angle_ct = _require_defined_ct(jNo, line_name, "Angle", cycle_dict,
                                   _ctx_value(cycle_dict, "fitup_3d", stage_3d, (stage_3d,)))
    r3tig_ct = _require_defined_ct(jNo, line_name, "R3tig", cycle_dict,
                                   _ctx_value(cycle_dict, "tig_spf", stage_3d, (stage_3d,)))
    tig2_3d_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                     _ctx_value(cycle_dict, "tig2_spf", stage_3d, (stage_3d,)))
    mag_3d_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                    _ctx_value(cycle_dict, "mag_spf", stage_3d, (stage_3d,)))
    gan_out_ct = _require_defined_ct(jNo, line_name, "Gan", cycle_dict,
                                     _ctx_value(cycle_dict, "out_3d", stage_3d, (stage_3d,)))

    pos = None
    if stage_3d in STRICT_STATION_FIFO:
        _occ_no, pos = yield from wait_prev_station_fifo(env, stage_3d, jNo)

    r3_req = resources["R3"].request()
    yield r3_req
    station_req = resources[stage_3d].request()
    yield station_req

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Angle", angle_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "R3tig", r3tig_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    resources["R3"].release(r3_req)

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Tig2", tig2_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_on_held_resource(env, jNo, line_name, "Mag", mag_3d_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if stage_3d in STRICT_STATION_FIFO:
        mark_station_fifo_done(stage_3d, pos)
    resources[stage_3d].release(station_req)

    yield env.process(run_line_step(
        env,
        line_name,
        jNo,
        pipe_step,
        process_named_with_resource(env, jNo, line_name, "Gan", "Gan", gan_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step

# 초보자 설명: F3D 제품의 Sub 라인 공정을 실행한다.
def run_f3d_sub(env, jNo, sub_step, cycle_dict, spool_name, fallback_ft, f3d_fit1_ready_2d1_evt, f3d_fit2_ready_3d_evt):
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)

    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-1", from_gantry=False));
    sub_step += 1
    if not f3d_fit1_ready_2d1_evt.triggered:
        f3d_fit1_ready_2d1_evt.succeed()

    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting(env, jNo, "Sub", fit2_proc, cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-2", from_gantry=False));
    sub_step += 1
    if not f3d_fit2_ready_3d_evt.triggered:
        f3d_fit2_ready_3d_evt.succeed()
    return sub_step


# 초보자 설명: F3D의 두 Sub 흐름을 병렬로 실행한다.
def run_f3d_sub_parallel(env, jNo, cycle_dict, spool_name, fallback_ft, f3d_fit1_ready_2d1_evt, f3d_fit2_ready_3d_evt):
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)
    lane1_done = env.event()

    # 초보자 설명: 'lane1' 단계의 처리를 맡는 보조 함수다.
    def lane1():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub", sub_step, cycle_dict, "R2", "2D-1", from_gantry=False));
        sub_step += 1
        if not f3d_fit1_ready_2d1_evt.triggered:
            f3d_fit1_ready_2d1_evt.succeed()
        release_remaining_line_steps("Sub", jNo, sub_step - 1)
        if not lane1_done.triggered:
            lane1_done.succeed()

    # 초보자 설명: 'lane2' 단계의 처리를 맡는 보조 함수다.
    def lane2():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub2", fit2_proc, cycle_dict, 2)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(
            run_robot_transport_only(env, jNo, "Sub2", sub_step, cycle_dict, "R2", "2D-2", from_gantry=False));
        sub_step += 1
        if not f3d_fit2_ready_3d_evt.triggered:
            f3d_fit2_ready_3d_evt.succeed()
        release_remaining_line_steps("Sub2", jNo, sub_step - 1)

    env.process(lane1())
    env.process(lane2())
    yield lane1_done


# 초보자 설명: 두 번 합류하는 Short 계열의 공통 최종 시퀀스를 실행한다.
def _run_dual_join_short_sequence(env, jNo, pipe_step, cycle_dict, fit1_ready_evt, fit2_ready_evt,
                                  main_r14_1_done_evt=None, main_turnover_done_evt=None):
    """F2DS/F2DL 공통 최종 단관 2회 용접 시퀀스.

    수정 포인트:
    - Pick1/Pick2 선행 준비 단계에서는 Short 용접장을 선점하지 않는다.
    - S1 준비 대기 중에는 R14도 오래 점유하지 않는다.
    - 실제 첫 Angle 공정이 시작되는 시점부터 두 번째 Mag 종료까지 Short 용접장을 연속 점유한다.
    - 두 번째 S2 합류 대기 중에도 R14를 불필요하게 오래 점유하지 않는다.
    """
    line_name = "Pipe"
    stage = "Short"

    pick1_ct = _require_defined_ct(jNo, line_name, "Pick1", cycle_dict,
                                   _ctx_value(cycle_dict, "r14_pick_1", stage, (stage,)))
    pick2_1_ct = _require_defined_ct(jNo, line_name, "Pick2", cycle_dict,
                                     _ctx_value(cycle_dict, "r14_pick_2", stage, (stage,)))
    angle1_ct = _require_defined_ct(jNo, line_name, "Angle", cycle_dict,
                                    _ctx_value(cycle_dict, "fitup3d_short_1", stage, (stage,)))
    r14tig1_ct = _require_defined_ct(jNo, line_name, "R14tig", cycle_dict,
                                     _ctx_value(cycle_dict, "tig_short_1", stage, (stage,)))
    tig2_1_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                    _ctx_value(cycle_dict, "tig2_short_1", stage, (stage,)))
    mag1_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                  _ctx_value(cycle_dict, "mag_short_1", stage, (stage,)))
    turnover_ct = _require_defined_ct(jNo, line_name, "TurnOv", cycle_dict,
                                      _ctx_value(cycle_dict, "turn_short_2", stage, (stage,)))
    pick2_2_ct = _require_defined_ct(jNo, line_name, "Pick2", cycle_dict,
                                     _ctx_value(cycle_dict, "r14_pick_2b", stage, (stage,)))
    angle2_ct = _require_defined_ct(jNo, line_name, "Angle", cycle_dict,
                                    _ctx_value(cycle_dict, "fitup3d_short_2", stage, (stage,)))
    r14tig2_ct = _require_defined_ct(jNo, line_name, "R14tig", cycle_dict,
                                     _ctx_value(cycle_dict, "tig_short_2", stage, (stage,)))
    tig2_2_ct = _require_defined_ct(jNo, line_name, "Tig2", cycle_dict,
                                    _ctx_value(cycle_dict, "tig2_short_2", stage, (stage,)))
    mag2_ct = _require_defined_ct(jNo, line_name, "Mag", cycle_dict,
                                  _ctx_value(cycle_dict, "mag_short_2", stage, (stage,)))
    r14_out_ct = _require_defined_ct(jNo, line_name, "R14", cycle_dict,
                                     _ctx_value(cycle_dict, "out_short", stage, (stage,)))

    # -------------------------
    # 선행 준비 1차: Pick1 -> Pick2
    # - Short 용접장은 아직 점유하지 않는다.
    # - R14는 Pick 작업 동안만 점유한다.
    # -------------------------
    r14_req = resources["R14"].request()
    yield r14_req
    for display_name, ct in [("Pick1", pick1_ct), ("Pick2", pick2_1_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1
    resources["R14"].release(r14_req)

    if fit1_ready_evt is not None and not fit1_ready_evt.triggered:
        yield fit1_ready_evt

    # -------------------------
    # 실제 Short 용접장 연속 점유 시작
    # Angle1 시작 시점부터 2차 Mag 끝까지 유지
    # -------------------------
    yield from acquire_short_station(env, jNo)

    # 1차 R14 용접 구간: Angle -> R14tig
    r14_req = resources["R14"].request()
    yield r14_req
    for display_name, ct in [("Angle", angle1_ct), ("R14tig", r14tig1_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1
    if main_r14_1_done_evt is not None and not main_r14_1_done_evt.triggered:
        main_r14_1_done_evt.succeed()
    resources["R14"].release(r14_req)

    for display_name, ct in [("Tig2", tig2_1_ct), ("Mag", mag1_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    # 2차 선행 준비: TurnOv -> Pick2
    # - Short 용접장은 계속 점유
    # - R14는 준비 작업 동안만 점유 후 해제
    r14_req = resources["R14"].request()
    yield r14_req
    for display_name, ct in [("TurnOv", turnover_ct), ("Pick2", pick2_2_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1
    if main_turnover_done_evt is not None and not main_turnover_done_evt.triggered:
        main_turnover_done_evt.succeed()
    resources["R14"].release(r14_req)

    if fit2_ready_evt is not None and not fit2_ready_evt.triggered:
        yield fit2_ready_evt

    # 2차 R14 용접 구간: Angle -> R14tig
    r14_req = resources["R14"].request()
    yield r14_req
    for display_name, ct in [("Angle", angle2_ct), ("R14tig", r14tig2_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1
    resources["R14"].release(r14_req)

    for display_name, ct in [("Tig2", tig2_2_ct), ("Mag", mag2_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    release_short_station(env)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "R14", "R14", r14_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step

# 초보자 설명: F2DS 계열 Short 제품의 main 라인 공정을 실행한다.
def run_fsshrt_main(env, jNo, pipe_step, cycle_dict, fs_fit1_ready_evt, fs_fit2_ready_evt, fs_main_r14_1_done_evt,
                    fs_main_turnover_done_evt):
    """F2DS(사용자 표현 FlShrt) M1 공정.

    Cut 이후:
    FitMov(BM) -> ShrtBev1(BI) -> FitMov(BM) -> ShrtBev2(BK) -> R2(Z)
    -> Pick1(AM) -> Pick2(AN) -> [S1 합류] -> Angle(AP) -> R14Tig1(AQ)
    -> Tig2(AR) -> Mag(AS) -> TurnOv(AT) -> Pick2(AU) -> [S2 합류]
    -> Angle(AV) -> R14Tig1(AW) -> Tig2(AX) -> Mag(AY) -> R14(BC)
    """
    line_name = "Pipe"

    yield from wait_prev_short_s_prep(env, jNo)

    fitmov1_ct = _require_defined_ct(jNo, line_name, "FitMov", cycle_dict,
                                     _ctx_short_prep_value(cycle_dict, "fitmov_short"))
    shrt_bev1_ct = _require_defined_ct(jNo, line_name, "ShrtBev1", cycle_dict,
                                       _ctx_short_prep_value(cycle_dict, ("shrt_bev1_j", "shrt_bev1_v")))
    fitmov2_ct = _require_defined_ct(jNo, line_name, "FitMov", cycle_dict,
                                     _ctx_short_prep_value(cycle_dict, "fitmov_short"))
    shrt_bev2_ct = _require_defined_ct(jNo, line_name, "ShrtBev2", cycle_dict,
                                       _ctx_short_prep_value(cycle_dict, ("shrt_bev2_j", "shrt_bev2_v")))

    for display_name, resource_name, ct in [
        ("FitMov", "FitMov", fitmov1_ct),
        ("ShrtBev1", "sShBevel", shrt_bev1_ct),
        ("FitMov", "FitMov", fitmov2_ct),
        ("ShrtBev2", "sShBevel", shrt_bev2_ct),
    ]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_with_resource(env, jNo, line_name, resource_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "R2", cycle_dict),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    mark_short_s_prep_done(jNo)

    pipe_step = yield env.process(_run_dual_join_short_sequence(
        env, jNo, pipe_step, cycle_dict,
        fs_fit1_ready_evt, fs_fit2_ready_evt,
        fs_main_r14_1_done_evt, fs_main_turnover_done_evt,
    ))

    mark_short_s_first_short_done(jNo)
    mark_short_s_weld_done(jNo)
    return pipe_step


# 초보자 설명: F2DS 계열 Short 제품의 sub 라인 공정을 실행한다.
def run_fsshrt_sub(env, jNo, sub_step, cycle_dict, spool_name, fs_fit1_ready_evt, fs_fit2_ready_evt,
                   fs_main_r14_1_done_evt, fs_main_turnover_done_evt):
    fit1_ft, fit2_ft = spool_fit_detail.get(spool_name, ("", ""))
    fit1_proc = get_fit_proc_from_ft(fit1_ft)
    fit2_proc = get_fit_proc_from_ft(fit2_ft)

    # S1/S2는 모두 R2 준비까지만 수행하고,
    # 실제 R14tig는 M1(Pipe) 라인에서 수행한다.
    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict)));
    sub_step += 1
    if not fs_fit1_ready_evt.triggered:
        fs_fit1_ready_evt.succeed()

    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting(env, jNo, "Sub", fit2_proc, cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict)));
    sub_step += 1
    if not fs_fit2_ready_evt.triggered:
        fs_fit2_ready_evt.succeed()
    return sub_step


# 초보자 설명: F2DS 계열의 두 Sub 라인을 병렬로 실행한다.
def run_fsshrt_sub_parallel(env, jNo, cycle_dict, spool_name, fs_fit1_ready_evt, fs_fit2_ready_evt,
                            fs_main_r14_1_done_evt, fs_main_turnover_done_evt):
    fit1_ft, fit2_ft = spool_fit_detail.get(spool_name, ("", ""))
    fit1_proc = get_fit_proc_from_ft(fit1_ft)
    fit2_proc = get_fit_proc_from_ft(fit2_ft)
    lane1_done = env.event()
    lane2_done = env.event()

    # 초보자 설명: 'lane1' 단계의 처리를 맡는 보조 함수다.
    def lane1():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict),
                                        enforce_prev_step=False));
        sub_step += 1
        if not fs_fit1_ready_evt.triggered:
            fs_fit1_ready_evt.succeed()
        release_remaining_line_steps("Sub", jNo, sub_step - 1)
        if not lane1_done.triggered:
            lane1_done.succeed()

    # 초보자 설명: 'lane2' 단계의 처리를 맡는 보조 함수다.
    def lane2():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub2", fit2_proc, cycle_dict, 2)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R2", cycle_dict),
                                        enforce_prev_step=False));
        sub_step += 1
        if not fs_fit2_ready_evt.triggered:
            fs_fit2_ready_evt.succeed()
        release_remaining_line_steps("Sub2", jNo, sub_step - 1)
        if not lane2_done.triggered:
            lane2_done.succeed()

    env.process(lane1())
    env.process(lane2())
    yield lane1_done & lane2_done


# 초보자 설명: F2DL 계열 Short 제품의 main 라인 공정을 실행한다.
def run_flshrt_main(env, jNo, pipe_step, cycle_dict, fl_fit1_ready_evt, fl_fit2_ready_evt, fl_main_r14_1_done_evt,
                    fl_main_turnover_done_evt):
    """F2DL(사용자 표현 FsShrt) M1 공정."""
    line_name = "Pipe"

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_after_cutbuf(env, jNo, line_name, "Con1", cycle_dict),
    ))
    pipe_step += 1

    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 1, include_sen2=True)
    )
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "Con2", cycle_dict),
    ))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 2, include_sen2=True)
    )
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "Conv3", cycle_dict),
    ))
    pipe_step += 1
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "R2", cycle_dict),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    pipe_step = yield env.process(_run_dual_join_short_sequence(
        env, jNo, pipe_step, cycle_dict,
        fl_fit1_ready_evt, fl_fit2_ready_evt,
        fl_main_r14_1_done_evt, fl_main_turnover_done_evt,
    ))
    return pipe_step

# 초보자 설명: F2DL 계열 Short 제품의 sub 라인 공정을 실행한다.
def run_flshrt_sub(env, jNo, sub_step, cycle_dict, spool_name, fl_fit1_ready_evt, fl_fit2_ready_evt,
                   fl_main_r14_1_done_evt, fl_main_turnover_done_evt):
    fit1_ft, fit2_ft = spool_fit_detail.get(spool_name, ("", ""))
    fit1_proc = get_fit_proc_from_ft(fit1_ft)
    fit2_proc = get_fit_proc_from_ft(fit2_ft)
    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict)));
    sub_step += 1
    if not fl_fit1_ready_evt.triggered:
        fl_fit1_ready_evt.succeed()
    yield env.process(
        run_line_step(env, "Sub", jNo, sub_step, process_sub_fitting(env, jNo, "Sub", fit2_proc, cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
    sub_step += 1
    yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict)));
    sub_step += 1
    if not fl_fit2_ready_evt.triggered:
        fl_fit2_ready_evt.succeed()
    return sub_step


# 초보자 설명: F2DL 계열의 두 Sub 라인을 병렬로 실행한다.
def run_flshrt_sub_parallel(env, jNo, cycle_dict, spool_name, fl_fit1_ready_evt, fl_fit2_ready_evt,
                            fl_main_r14_1_done_evt, fl_main_turnover_done_evt):
    fit1_ft, fit2_ft = spool_fit_detail.get(spool_name, ("", ""))
    fit1_proc = get_fit_proc_from_ft(fit1_ft)
    fit2_proc = get_fit_proc_from_ft(fit2_ft)
    lane1_done = env.event()
    lane2_done = env.event()

    # 초보자 설명: 'lane1' 단계의 처리를 맡는 보조 함수다.
    def lane1():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict)));
        sub_step += 1
        if not fl_fit1_ready_evt.triggered:
            fl_fit1_ready_evt.succeed()
        release_remaining_line_steps("Sub", jNo, sub_step - 1)
        if not lane1_done.triggered:
            lane1_done.succeed()

    # 초보자 설명: 'lane2' 단계의 처리를 맡는 보조 함수다.
    def lane2():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub2", fit2_proc, cycle_dict, 2)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R2", cycle_dict),
                                        enforce_prev_step=False));
        sub_step += 1
        if not fl_fit2_ready_evt.triggered:
            fl_fit2_ready_evt.succeed()
        release_remaining_line_steps("Sub2", jNo, sub_step - 1)
        if not lane2_done.triggered:
            lane2_done.succeed()

    env.process(lane1())
    env.process(lane2())
    yield lane1_done & lane2_done


# 초보자 설명: 2DS 제품의 main 라인 공정을 실행한다.
def run_2ds_main(env, jNo, pipe_step, cycle_dict, fit_ready_evt):
    """2DS 전용 main flow.

    Cut은 caller에서 이미 수행된 상태이며, 여기서는 Cut 이후 공정부터 처리한다.
    FitMov -> ShrtBev1 -> FitMov -> ShrtBev2 -> R1 -> R2
    -> Pick1 -> Pick2 -> Fit -> R14Tig1 -> Tig2 -> Mag -> R14(out)

    자원 규칙
    - R14 자원: Pick1 ~ R14Tig1 동안 점유 후 해제
    - Short 자원: Pick1 ~ Mag 동안 점유 후 해제
    """
    line_name = "Pipe"
    stage = "Short"

    yield from wait_prev_short_s_prep(env, jNo)

    fitmov1_ct = _require_defined_ct(
        jNo, line_name, "FitMov", cycle_dict,
        _ctx_short_prep_value(cycle_dict, "fitmov_short"),
    )
    shrt_bev1_ct = _require_defined_ct(
        jNo, line_name, "ShrtBev1", cycle_dict,
        _ctx_short_prep_value(cycle_dict, ("shrt_bev1_j", "shrt_bev1_v")),
    )
    fitmov2_ct = _require_defined_ct(
        jNo, line_name, "FitMov", cycle_dict,
        _ctx_short_prep_value(cycle_dict, "fitmov_short"),
    )
    shrt_bev2_ct = _require_defined_ct(
        jNo, line_name, "ShrtBev2", cycle_dict,
        _ctx_short_prep_value(cycle_dict, ("shrt_bev2_j", "shrt_bev2_v")),
    )
    r1_ct = _require_defined_ct(
        jNo, line_name, "R1", cycle_dict,
        _ctx_short_prep_value(cycle_dict, "r1_short"),
    )

    for display_name, resource_name, ct in [
        ("FitMov", "FitMov", fitmov1_ct),
        ("ShrtBev1", "sShBevel", shrt_bev1_ct),
        ("FitMov", "FitMov", fitmov2_ct),
        ("ShrtBev2", "sShBevel", shrt_bev2_ct),
        ("R1", "R1", r1_ct),
    ]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_with_resource(env, jNo, line_name, resource_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    mark_short_s_prep_done(jNo)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "R2", cycle_dict),
        enforce_prev_step=False,
    ))
    pipe_step += 1

    if not fit_ready_evt.triggered:
        yield fit_ready_evt

    pick1_ct = _require_defined_ct(
        jNo, line_name, "Pick1", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_1", stage, (stage,)),
    )
    pick2_ct = _require_defined_ct(
        jNo, line_name, "Pick2", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_2", stage, (stage,)),
    )
    fit_ct = _require_defined_ct(
        jNo, line_name, "Fit", cycle_dict,
        _ctx_value(cycle_dict, "fitup_short_1", stage, (stage,)),
    )
    r14tig_ct = _require_defined_ct(
        jNo, line_name, "R14tig", cycle_dict,
        _ctx_value(cycle_dict, "tig_short_1", stage, (stage,)),
    )
    tig2_ct = _require_defined_ct(
        jNo, line_name, "Tig2", cycle_dict,
        _ctx_value(cycle_dict, "tig2_short_1", stage, (stage,)),
    )
    mag_ct = _require_defined_ct(
        jNo, line_name, "Mag", cycle_dict,
        _ctx_value(cycle_dict, "mag_short_1", stage, (stage,)),
    )
    r14_out_ct = _require_defined_ct(
        jNo, line_name, "R14", cycle_dict,
        _ctx_value(cycle_dict, "out_short", stage, (stage,)),
    )

    yield from acquire_short_station(env, jNo)
    r14_req = resources["R14"].request()
    yield r14_req

    for display_name, ct in [
        ("Pick1", pick1_ct),
        ("Pick2", pick2_ct),
        ("Fit", fit_ct),
        ("R14tig", r14tig_ct),
    ]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    resources["R14"].release(r14_req)

    for display_name, ct in [("Tig2", tig2_ct), ("Mag", mag_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    mark_short_s_first_short_done(jNo)
    mark_short_s_weld_done(jNo)
    release_short_station(env)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "R14", "R14", r14_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step


# 초보자 설명: 2DL 제품의 main 라인 공정을 실행한다.
def run_2dl_main(env, jNo, pipe_step, cycle_dict, fit_ready_evt):
    """2DL 전용 main flow."""
    line_name = "Pipe"
    stage = "Short"

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_after_cutbuf(env, jNo, line_name, "Con1", cycle_dict),
    ))
    pipe_step += 1

    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 1, include_sen2=True)
    )
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "Con2", cycle_dict),
    ))
    pipe_step += 1
    pipe_step = yield env.process(
        run_main_bevel_group(env, jNo, line_name, pipe_step, cycle_dict, 2, include_sen2=False)
    )
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "Conv3", cycle_dict),
    ))
    pipe_step += 1
    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process(env, jNo, line_name, "R2", cycle_dict),
    ))
    pipe_step += 1

    if not fit_ready_evt.triggered:
        yield fit_ready_evt

    pick1_ct = _require_defined_ct(
        jNo, line_name, "Pick1", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_1", stage, (stage,)),
    )
    pick2_ct = _require_defined_ct(
        jNo, line_name, "Pick2", cycle_dict,
        _ctx_value(cycle_dict, "r14_pick_2", stage, (stage,)),
    )
    fit_ct = _require_defined_ct(
        jNo, line_name, "Fit", cycle_dict,
        _ctx_value(cycle_dict, "fitup_short_1", stage, (stage,)),
    )
    r14tig_ct = _require_defined_ct(
        jNo, line_name, "R14tig", cycle_dict,
        _ctx_value(cycle_dict, "tig_short_1", stage, (stage,)),
    )
    tig2_ct = _require_defined_ct(
        jNo, line_name, "Tig2", cycle_dict,
        _ctx_value(cycle_dict, "tig2_short_1", stage, (stage,)),
    )
    mag_ct = _require_defined_ct(
        jNo, line_name, "Mag", cycle_dict,
        _ctx_value(cycle_dict, "mag_short_1", stage, (stage,)),
    )
    r14_out_ct = _require_defined_ct(
        jNo, line_name, "R14", cycle_dict,
        _ctx_value(cycle_dict, "out_short", stage, (stage,)),
    )

    yield from acquire_short_station(env, jNo)
    r14_req = resources["R14"].request()
    yield r14_req

    for display_name, ct in [
        ("Pick1", pick1_ct),
        ("Pick2", pick2_ct),
        ("Fit", fit_ct),
        ("R14tig", r14tig_ct),
    ]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    resources["R14"].release(r14_req)

    for display_name, ct in [("Tig2", tig2_ct), ("Mag", mag_ct)]:
        yield env.process(run_line_step(
            env, line_name, jNo, pipe_step,
            process_named_on_held_resource(env, jNo, line_name, display_name, ct),
            enforce_prev_step=False,
        ))
        pipe_step += 1

    release_short_station(env)

    yield env.process(run_line_step(
        env, line_name, jNo, pipe_step,
        process_named_with_resource(env, jNo, line_name, "R14", "R14", r14_out_ct),
        enforce_prev_step=False,
    ))
    pipe_step += 1
    return pipe_step

# 초보자 설명: FF 제품의 두 Sub 흐름을 병렬로 실행한다.
def run_ff_sub_parallel(env, jNo, cycle_dict, spool_name, fallback_ft=None, ff_fit1_ready_evt=None,
                        ff_fit2_ready_evt=None):
    fit1_proc, fit2_proc = get_two_fit_procs_for_spool(spool_name, fallback_ft)
    lane1_done = env.event()
    lane2_done = env.event()

    # 초보자 설명: 'lane1' 단계의 처리를 맡는 보조 함수다.
    def lane1():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub", fit1_proc, cycle_dict, 1)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict),
                                        enforce_prev_step=False));
        sub_step += 1
        if ff_fit1_ready_evt is not None and not ff_fit1_ready_evt.triggered:
            ff_fit1_ready_evt.succeed()
        release_remaining_line_steps("Sub", jNo, sub_step - 1)
        if not lane1_done.triggered:
            lane1_done.succeed()

    # 초보자 설명: 'lane2' 단계의 처리를 맡는 보조 함수다.
    def lane2():
        sub_step = 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step,
                                        process_sub_fitting_first(env, jNo, "Sub2", fit2_proc, cycle_dict, 2)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R1", cycle_dict)));
        sub_step += 1
        yield env.process(run_line_step(env, "Sub2", jNo, sub_step, process(env, jNo, "Sub2", "R2", cycle_dict),
                                        enforce_prev_step=False));
        sub_step += 1
        if ff_fit2_ready_evt is not None and not ff_fit2_ready_evt.triggered:
            ff_fit2_ready_evt.succeed()
        release_remaining_line_steps("Sub2", jNo, sub_step - 1)
        if not lane2_done.triggered:
            lane2_done.succeed()

    env.process(lane1())
    env.process(lane2())
    yield lane1_done & lane2_done


# 초보자 설명: 1차 시뮬레이션 결과에서 R2 유휴 구간을 뽑아낸다.
def extract_idle_time_r2(gantt_rows, min_idle=2.0):
    r2_steps = sorted([g for g in gantt_rows if g[2] in ["R2", "R2tig"]], key=lambda x: (x[3], x[4]))
    idle_list = []
    if not r2_steps:
        return idle_list
    for i in range(len(r2_steps) - 1):
        cur = r2_steps[i]
        nxt = r2_steps[i + 1]
        gap = safe_ct(nxt[3], 0) - safe_ct(cur[4], 0)
        if gap >= min_idle:
            idle_list.append({"start": safe_ct(cur[4], 0), "end": safe_ct(nxt[3], 0), "gap": gap, "from_job": cur[0],
                              "to_job": nxt[0]})
    return idle_list


# 초보자 설명: plan2 Job을 R2 유휴 시간대에 넣기 위한 시작 오프셋을 계산한다.
def calc_offset_to_r2_for_plan2(plan_row):
    ctNum = find_match_row_by_plan([plan_row], 1)
    cycle_dict = build_cycle_dict(CT[ctNum - 1], plan_row=plan_row) if ctNum else DEFAULT_CYCLE.copy()
    prod = norm_text(plan_row[4])
    ft = norm_text(plan_row[7])
    fit_proc = get_fit_proc_from_ft(ft)
    if prod == "2DS":
        return (
            (_ctx_value(cycle_dict, "cut", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, "fitmov_short", 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, ("shrt_bev1_j", "shrt_bev1_v"), 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, "fitmov_short", 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, ("shrt_bev2_j", "shrt_bev2_v"), 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, "r1_short", 0.0) or 0.0)
        )
    if prod == "2DL":
        return (
            _ctx_value(cycle_dict, "cut", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "conv1", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "sen1_1", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "bev1_j", "Short", ("Short",), 0.0)
            or _ctx_value(cycle_dict, "bev1_v", "Short", ("Short",), 0.0)
            or 0.0
        ) + (
            _ctx_value(cycle_dict, "sen2_1", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "conv2", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "sen1_2", "Short", ("Short",), 0.0) or 0.0
        ) + (
            _ctx_value(cycle_dict, "bev2_j", "Short", ("Short",), 0.0)
            or _ctx_value(cycle_dict, "bev2_v", "Short", ("Short",), 0.0)
            or 0.0
        ) + (
            _ctx_value(cycle_dict, "conv3", "Short", ("Short",), 0.0) or 0.0
        )
    if prod == "F2DS":
        return (
            (_ctx_value(cycle_dict, "cut", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, "fitmov_short", 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, ("shrt_bev1_j", "shrt_bev1_v"), 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, "fitmov_short", 0.0) or 0.0)
            + (_ctx_short_prep_value(cycle_dict, ("shrt_bev2_j", "shrt_bev2_v"), 0.0) or 0.0)
        )
    if prod == "F2DL":
        return (
            (_ctx_value(cycle_dict, "cut", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_value(cycle_dict, "conv1", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_value(cycle_dict, "sen1_1", "Short", ("Short",), 0.0) or 0.0)
            + ((_ctx_value(cycle_dict, "bev1_j", "Short", ("Short",), 0.0) or _ctx_value(cycle_dict, "bev1_v", "Short", ("Short",), 0.0)) or 0.0)
            + (_ctx_value(cycle_dict, "sen2_1", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_value(cycle_dict, "conv2", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_value(cycle_dict, "sen1_2", "Short", ("Short",), 0.0) or 0.0)
            + ((_ctx_value(cycle_dict, "bev2_j", "Short", ("Short",), 0.0) or _ctx_value(cycle_dict, "bev2_v", "Short", ("Short",), 0.0)) or 0.0)
            + (_ctx_value(cycle_dict, "sen2_2", "Short", ("Short",), 0.0) or 0.0)
            + (_ctx_value(cycle_dict, "conv3", "Short", ("Short",), 0.0) or 0.0)
        )
    if prod == "FF":
        fit_ct = cycle_dict.get(fit_proc, cycle_dict.get("FitMov", 0))
        return fit_ct + cycle_dict.get("R1", 0)
    return 0.0


# 초보자 설명: plan2를 idle 삽입 규칙에 맞는 순서로 정렬한다.
def order_plan2_for_idle_insert(plan2_rows):
    type_priority = {"2DS": 0, "2DL": 1, "F2DS": 2, "F2DL": 3, "FF": 4, "ff": 4}
    return sorted(plan2_rows, key=lambda r: (-_inch_to_float(r[2]), type_priority.get(norm_text(r[4]), 99)))


# 초보자 설명: plan2 각 Job에 강제 Cut 시작 시각을 반영한다.
def apply_forced_cut_starts_to_plan2(plan2_rows, idleTimeR2):
    ordered = order_plan2_for_idle_insert(plan2_rows)
    result = []
    for idx, row in enumerate(ordered):
        rr = list(row)
        if idx < len(idleTimeR2):
            idle_start = idleTimeR2[idx]["start"]
            offset = calc_offset_to_r2_for_plan2(tuple(rr))
            rr[8] = max(0.0, idle_start - offset)
        else:
            rr[8] = None
        result.append(tuple(rr))
    return renumber_plan_rows(result)


# 초보자 설명: 각 Job의 최초 시작 시각을 맵으로 만든다.
def _get_job_first_start_map(gantt_rows):
    job_first = {}
    for job_id, line, proc, start, end in gantt_rows:
        s = safe_ct(start, 0.0)
        if job_id not in job_first or s < job_first[job_id]:
            job_first[job_id] = s
    return job_first


# 초보자 설명: 여러 계획 조각을 실제 시작 시각 기준 최종 plan으로 합친다.
def combine_final_plan_by_start_time(srtPlan_rows, plan2_rows, first_gantt_rows):
    """규칙 기반 최종 plan 병합.

    핵심 원칙:
    1) srtPlan_rows 안의 기존 순서(사이즈 정렬 + 2D/F2D/H3D/F3D 배치 + 용접장 할당)는 절대 깨지지 않아야 한다.
    2) Short 계열(plan2_rows)만 R2 idle time 기반 forced cut start(rr[8])에 맞춰
       srtPlan 사이에 끼워 넣는다.

    기존 구현은 srtPlan_rows 자체를 first_gantt_rows의 실제 시작시각으로 다시 정렬하면서
    규칙 기반의 원래 순서를 깨뜨리고 있었다. 여기서는 srtPlan 순서를 고정한 채,
    각 long plan row의 "첫 시작 시각"을 삽입 기준점으로만 사용한다.
    """
    srt_job_first = _get_job_first_start_map(first_gantt_rows)

    long_rows = []
    for row in srtPlan_rows:
        rr = list(row)
        if len(rr) < 9:
            rr.append(None)
        job_no = rr[0]
        first_start = srt_job_first.get(job_no, 10 ** 15)
        long_rows.append((first_start, tuple(rr)))

    short_rows = []
    for row in plan2_rows:
        rr = list(row)
        if len(rr) < 9:
            rr.append(None)
        sort_time = safe_ct(rr[8], 10 ** 15)
        short_rows.append((sort_time, tuple(rr)))
    short_rows.sort(key=lambda x: x[0])

    merged = []
    short_idx = 0
    for first_start, long_row in long_rows:
        while short_idx < len(short_rows) and short_rows[short_idx][0] < first_start:
            merged.append(short_rows[short_idx][1])
            short_idx += 1
        merged.append(long_row)

    while short_idx < len(short_rows):
        merged.append(short_rows[short_idx][1])
        short_idx += 1

    return renumber_plan_rows(merged)


# 초보자 설명: 시뮬레이션용 plan 행을 저장용 전체 형식으로 확장한다.
def convert_sim_plan_row_to_full_row(sim_row):
    no, spool, inch, sch, prod, weld, length, fit, _forced = sim_row[:9]
    fit1_ft, fit2_ft = spool_fit_detail.get(spool, ("", ""))
    out_length1 = length
    out_length2 = 0
    out_fit1 = fit1_ft if fit1_ft != "" else 0
    out_fit2 = fit2_ft if fit2_ft != "" else 0
    if (out_fit1 == 0 or out_fit1 == "") and fit not in [None, "", 0]:
        out_fit1 = fit
    return (no, spool, inch, sch, prod, weld, out_length1, out_length2, out_fit1, out_fit2)


# 초보자 설명: 엑셀 저장용 plan 전체 행 목록을 만든다.
def build_full_format_plan_rows(sim_plan_rows):
    return [tuple([idx] + list(convert_sim_plan_row_to_full_row(row))[1:]) for idx, row in
            enumerate(sim_plan_rows, start=1)]


# =====================================================
# Job 디스패처
# -----------------------------------------------------
# job() 함수는 Job 한 개를 받아 어떤 라우팅 함수를 실행할지 결정한다.
# 즉, 제품 타입별 run_... 함수들의 관문 역할을 한다.
# =====================================================

# 초보자 설명: Job 한 개가 어떤 제품 타입인지 보고 맞는 라우팅 함수를 호출하는 핵심 디스패처다.
def job(env, jNo, name, inch, Sch, prod, weld=None, length=None, ft=None, forced_cut_start=None):
    ctNum = find_match_row(jNo)
    if ctNum:
        cycle_dict = build_cycle_dict(CT[ctNum - 1], plan_row=plan[jNo - 1])
    else:
        raise ValueError(f"CycleTime.xlsx에서 매칭되는 공정 조건을 찾지 못했습니다.\nJob No: {jNo}\nPlan Row: {plan[jNo - 1]}")
    pipe_ready = env.event();
    sub_ready = env.event();
    fit_ready = env.event();
    ff_sub_fl_ready = env.event()
    f2d_fit_ready_2d1 = env.event();
    f2d_fit_ready_3d = env.event();
    h3d_sub_ready_2d2 = env.event();
    h3d_pipe2_ready_3d = env.event()
    f3d_fit1_ready_2d1 = env.event();
    f3d_fit2_ready_3d = env.event()
    fs_fit1_ready_short = env.event();
    fs_fit2_ready_short = env.event();
    fs_main_r14_1_done = env.event();
    fs_main_turnover_done = env.event()
    fl_fit1_ready_short = env.event();
    fl_fit2_ready_short = env.event();
    fl_main_r14_1_done = env.event();
    fl_main_turnover_done = env.event()
    ff_fit1_ready_short = env.event();
    ff_fit2_ready_short = env.event()

    # 초보자 설명: 'pipe_flow' 단계의 처리를 맡는 보조 함수다.
    def pipe_flow():
        if forced_cut_start is not None and prod != "FF":
            wait_t = safe_ct(forced_cut_start, 0.0) - env.now
            if wait_t > 0:
                yield env.timeout(wait_t)
        if prod not in ["H3D", "F3D"]:
            release_remaining_line_steps("Pipe2", jNo, 0)
        ft_str = norm_text(ft)
        fit_bevel_proc = get_fit_proc_from_ft(ft_str)
        pipe_step = 1
        if prod == "FF":
            if not PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                if fit_bevel_proc == "fl":
                    yield env.process(
                        run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "fl", cycle_dict)));
                    pipe_step += 1
                    yield env.process(run_line_step(env, "Sub", jNo, 1, process(env, jNo, "Sub", "fl", cycle_dict)))
                    if not ff_sub_fl_ready.triggered: ff_sub_fl_ready.succeed()
                else:
                    yield env.process(run_line_step(env, "Pipe", jNo, pipe_step,
                                                    process(env, jNo, "Pipe", fit_bevel_proc, cycle_dict)));
                    pipe_step += 1
                    if not ff_sub_fl_ready.triggered: ff_sub_fl_ready.succeed()
                for proc in ["R1", "R2", "R14"]:
                    yield env.process(
                        run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", proc, cycle_dict)));
                    pipe_step += 1
                release_remaining_line_steps("Pipe", jNo, pipe_step - 1);
                pipe_ready.succeed();
                return
            # FF는 두 개의 제품을 모두 Sub flow에서 병렬 준비한 뒤,
            # 두 제품이 모두 Short 용접장 도착 준비를 마치면 최종 Short 공정을 시작한다.
            release_remaining_line_steps("Pipe", jNo, 0)
            pipe_ready.succeed()
            return
        if prod in ["H3D", "F3D", "F2DS", "F2DL"]:
            yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Cut", cycle_dict)));
            pipe_step += 1
            # Short 계열 중 F2DS는 Con1 공정을 거치지 않으므로 CutBuf를 점유하면 안 됨
            if prod != "F2DS":
                yield CutBuf.put(1)
            if prod == "H3D":
                pipe_step = yield env.process(
                    run_h3d_main(env, jNo, pipe_step, cycle_dict, h3d_sub_ready_2d2, h3d_pipe2_ready_3d))
            elif prod == "F3D":
                pipe_step = yield env.process(
                    run_f3d_main(env, jNo, pipe_step, cycle_dict, length, f3d_fit1_ready_2d1, f3d_fit2_ready_3d))
            elif prod == "F2DS":
                pipe_step = yield env.process(
                    run_fsshrt_main(env, jNo, pipe_step, cycle_dict, fs_fit1_ready_short, fs_fit2_ready_short,
                                    fs_main_r14_1_done, fs_main_turnover_done))
            else:
                pipe_step = yield env.process(
                    run_flshrt_main(env, jNo, pipe_step, cycle_dict, fl_fit1_ready_short, fl_fit2_ready_short,
                                    fl_main_r14_1_done, fl_main_turnover_done))
            release_remaining_line_steps("Pipe", jNo, pipe_step - 1);
            pipe_ready.succeed();
            return
        yield env.process(run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "Cut", cycle_dict)));
        pipe_step += 1
        # 2DS는 Con1 없이 FitMov/sShBevel로 바로 진행하므로 CutBuf를 점유하면 안 됨
        if prod != "2DS":
            yield CutBuf.put(1)
        if prod == "2D" and length is not None and 600 <= float(length) < 1700:
            pipe_step = yield env.process(run_short_2d_pipe(env, jNo, pipe_step, cycle_dict, weld, fit_ready))
        elif prod == "2D" and length is not None and 1700 <= float(length) < 10800:
            pipe_step = yield env.process(run_long_2d_pipe(env, jNo, pipe_step, cycle_dict, weld, fit_ready))
        elif prod == "F2D":
            pipe_step = yield env.process(
                run_f2d_main(env, jNo, pipe_step, cycle_dict, length, f2d_fit_ready_2d1, f2d_fit_ready_3d))
        elif prod == "2DS":
            pipe_step = yield env.process(run_2ds_main(env, jNo, pipe_step, cycle_dict, fit_ready))
        elif prod == "2DL":
            pipe_step = yield env.process(run_2dl_main(env, jNo, pipe_step, cycle_dict, fit_ready))
        else:
            yield env.process(run_line_step(env, "Pipe", jNo, pipe_step,
                                            process_after_cutbuf(env, jNo, "Pipe", "Con1", cycle_dict)));
            pipe_step += 1
            for proc in ["Sen1", "Bevel1", "Con2", "Bevel2"]:
                yield env.process(
                    run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", proc, cycle_dict)));
                pipe_step += 1
            if prod == "2D":
                if length is not None and length < 1700:
                    if weld in ["2D-1", "2D-2"]:
                        yield env.process(
                            run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R2", weld, from_gantry=False));
                        pipe_step += 1
                        if not fit_ready.triggered: yield fit_ready
                        yield env.process(
                            run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", weld, cycle_dict)));
                        pipe_step += 1
                    else:
                        yield env.process(run_gan_step(env, jNo, "Pipe", pipe_step, cycle_dict, "Gan1"));
                        pipe_step += 1
                        yield env.process(
                            run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R3", "3D", from_gantry=True));
                        pipe_step += 1
                        if not fit_ready.triggered: yield fit_ready
                        yield env.process(
                            run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "3D", cycle_dict),
                                          enforce_prev_step=False));
                        pipe_step += 1
                else:
                    yield env.process(run_gan_step(env, jNo, "Pipe", pipe_step, cycle_dict, "Gan1"));
                    pipe_step += 1
                    if weld in ["2D-1", "2D-2"]:
                        yield env.process(
                            run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R2", weld, from_gantry=True));
                        pipe_step += 1
                    else:
                        yield env.process(
                            run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R3", "3D", from_gantry=True));
                        pipe_step += 1
                    if not fit_ready.triggered: yield fit_ready
                    target = weld if weld in ["2D-1", "2D-2"] else "3D"
                    yield env.process(
                        run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", target, cycle_dict),
                                      enforce_prev_step=False));
                    pipe_step += 1
            elif prod == "Short":
                yield env.process(
                    run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "R1", cycle_dict),
                                  enforce_prev_step=False));
                pipe_step += 1
                yield env.process(
                    run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "R2", cycle_dict),
                                  enforce_prev_step=False));
                pipe_step += 1
                yield env.process(
                    run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R14", "Short", from_gantry=False,
                                   enforce_prev_step=False));
                pipe_step += 1
            elif prod == "2DL":
                yield env.process(
                    run_line_step(env, "Pipe", jNo, pipe_step, process(env, jNo, "Pipe", "R2", cycle_dict)));
                pipe_step += 1
                # Short 용접 공간은 모든 Short 계열 Job이 공유하므로,
                # 이전 Short Job의 마지막 Short 공정이 끝난 뒤에만 첫 투입(R14)이 가능하다.
                yield from wait_prev_short_s_weld(env, jNo)
                yield env.process(
                    run_robot_step(env, jNo, "Pipe", pipe_step, cycle_dict, "R14", "Short", from_gantry=False));
                pipe_step += 1
        release_remaining_line_steps("Pipe", jNo, pipe_step - 1);
        pipe_ready.succeed()

    # 초보자 설명: 'fitting_flow' 단계의 처리를 맡는 보조 함수다.
    def fitting_flow():
        ft_str = norm_text(ft)
        fit_bevel_proc = get_fit_proc_from_ft(ft_str)
        sub_step = 1
        if prod == "2D" and length is not None and 600 <= float(length) < 1700:
            sub_step = yield env.process(
                run_short_2d_sub(env, jNo, sub_step, cycle_dict, weld, prod, fit_ready, fit_bevel_proc))
        elif prod == "2D" and length is not None and 1700 <= float(length) < 10800:
            sub_step = yield env.process(
                run_long_2d_sub(env, jNo, sub_step, cycle_dict, weld, prod, fit_ready, fit_bevel_proc))
        elif prod == "F2D":
            if PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                yield env.process(
                    run_f2d_sub_parallel(env, jNo, cycle_dict, name, ft, f2d_fit_ready_2d1, f2d_fit_ready_3d))
                sub_ready.succeed()
                return
            sub_step = yield env.process(
                run_f2d_sub(env, jNo, sub_step, cycle_dict, name, ft, f2d_fit_ready_2d1, f2d_fit_ready_3d))
        elif prod == "H3D":
            sub_step = yield env.process(run_h3d_sub(env, jNo, sub_step, cycle_dict, fit_bevel_proc, h3d_sub_ready_2d2))
        elif prod == "F3D":
            if PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                yield env.process(
                    run_f3d_sub_parallel(env, jNo, cycle_dict, name, ft, f3d_fit1_ready_2d1, f3d_fit2_ready_3d))
                sub_ready.succeed()
                return
            sub_step = yield env.process(
                run_f3d_sub(env, jNo, sub_step, cycle_dict, name, ft, f3d_fit1_ready_2d1, f3d_fit2_ready_3d))
        elif prod == "F2DS":
            if PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                yield env.process(
                    run_fsshrt_sub_parallel(env, jNo, cycle_dict, name, fs_fit1_ready_short, fs_fit2_ready_short,
                                            fs_main_r14_1_done, fs_main_turnover_done))
                sub_ready.succeed()
                return
            sub_step = yield env.process(
                run_fsshrt_sub(env, jNo, sub_step, cycle_dict, name, fs_fit1_ready_short, fs_fit2_ready_short,
                               fs_main_r14_1_done, fs_main_turnover_done))
        elif prod == "F2DL":
            if PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                yield env.process(
                    run_flshrt_sub_parallel(env, jNo, cycle_dict, name, fl_fit1_ready_short, fl_fit2_ready_short,
                                            fl_main_r14_1_done, fl_main_turnover_done))
                sub_ready.succeed()
                return
            sub_step = yield env.process(
                run_flshrt_sub(env, jNo, sub_step, cycle_dict, name, fl_fit1_ready_short, fl_fit2_ready_short,
                               fl_main_r14_1_done, fl_main_turnover_done))
        elif prod == "2D":
            yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                            process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
            sub_step += 1
            if weld in ["2D-1", "2D-2"]:
                yield env.process(run_robot_step(env, jNo, "Sub", sub_step, cycle_dict, "R2", weld, from_gantry=False))
                sub_step += 1
            elif weld == "3D":
                yield env.process(run_robot_step(env, jNo, "Sub", sub_step, cycle_dict, "R3", "3D", from_gantry=False))
                sub_step += 1
            else:
                alarm("ALARM 3 : 용접장 미지정")
            if ft_str in ["45el", "90el", "fl", "Tee", "ReTee", "2DS", "sShBevel"] and not fit_ready.triggered:
                fit_ready.succeed()
        elif prod == "2DL":
            yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                            process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict),
                                            enforce_prev_step=False));
            sub_step += 1
            if ft_str in ["45el", "90el", "fl", "Tee", "ReTee", "2DS", "sShBevel"] and not fit_ready.triggered:
                fit_ready.succeed()
        elif prod == "2DS":
            yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                            process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict),
                                            enforce_prev_step=False));
            sub_step += 1
            if ft_str in ["45el", "90el", "fl", "Tee", "ReTee", "2DS", "sShBevel"] and not fit_ready.triggered:
                fit_ready.succeed()
        elif prod == "Short":
            yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                            process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict, 1)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R1", cycle_dict)));
            sub_step += 1
            yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", "R2", cycle_dict),
                                            enforce_prev_step=False));
            sub_step += 1
            yield env.process(run_robot_step(env, jNo, "Sub", sub_step, cycle_dict, "R14", "Short", from_gantry=False));
            sub_step += 1
            if ft_str in ["45el", "90el", "fl", "Tee", "ReTee", "2DS", "sShBevel"] and not fit_ready.triggered:
                fit_ready.succeed()
        elif prod == "FF":
            if PARALLEL_SUB_MULTI_PRODUCTS_ENABLED:
                yield env.process(
                    run_ff_sub_parallel(env, jNo, cycle_dict, name, ft, ff_fit1_ready_short, ff_fit2_ready_short))
                sub_ready.succeed()
                return
            if fit_bevel_proc == "fl":
                if not ff_sub_fl_ready.triggered:
                    yield ff_sub_fl_ready
                sub_step = 2
            else:
                yield env.process(run_line_step(env, "Sub", jNo, sub_step,
                                                process_sub_fitting_first(env, jNo, "Sub", fit_bevel_proc, cycle_dict,
                                                                          1)));
                sub_step += 1
            for proc in ["R1", "R2", "R14"]:
                yield env.process(run_line_step(env, "Sub", jNo, sub_step, process(env, jNo, "Sub", proc, cycle_dict)));
                sub_step += 1
        release_remaining_line_steps("Sub", jNo, sub_step - 1);
        sub_ready.succeed()

    env.process(pipe_flow())
    env.process(fitting_flow())
    yield pipe_ready & sub_ready
    final_step = 49
    if prod == "FF":
        final_step = yield env.process(
            run_ff_main_final_sequence(env, jNo, final_step, cycle_dict, ff_fit1_ready_short, ff_fit2_ready_short)
        )
    elif prod not in ["2D", "F2D", "H3D", "F3D", "F2DS", "F2DL", "2DS", "2DL"]:
        yield env.process(
            run_line_step(env, "Pipe", jNo, final_step, process(env, jNo, "Pipe", weld if weld else "3D", cycle_dict),
                          enforce_prev_step=False))
        if prod in ["Short", "2DL"]:
            mark_short_s_weld_done(jNo)

    release_all_job_line_active_holds(jNo)


# =====================================================
# 결과 정리 / 간트 차트 / ProcessTime 저장
# -----------------------------------------------------
# 시뮬레이션이 끝난 뒤 사람이 보기 좋은 형태로 결과를 정리하는 구간이다.
# =====================================================

# 초보자 설명: ProcessTime 엑셀에 넣기 좋도록 시간 문자열을 꾸민다.
def _format_process_time_value(v):
    if v is None:
        return ""
    try:
        fv = float(v)
    except Exception:
        return str(v)
    if abs(fv - round(fv)) < 1e-9:
        return str(int(round(fv)))
    return f"{fv:.1f}".rstrip("0").rstrip(".")


DISPLAY_SUB_SPLIT_PROCS = {"45el", "90el", "fl", "Tee", "ReTee", "sShBevel"}
DISPLAY_LINE_ORDER = {"Pipe": 0, "Pipe2": 1, "Sub": 2, "Sub2": 3}


# 초보자 설명: 화면 표시용 간트 행을 구성한다.
def build_display_gantt_rows(gantt_rows):
    """차트/로그 표시용 gantt 라인을 정리한다.

    현재 시뮬레이션 로직은 Sub / Sub2를 실제로 별도 라인으로 기록한다.
    따라서 raw gantt 안에 이미 Sub2가 존재하면 그 기록을 그대로 사용해야
    차트와 ProcessTime이 완전히 일치한다.

    다만, 과거 버전 호환을 위해 Sub2가 전혀 없는 legacy 로그에 대해서만
    예전 방식의 표시용 재분배를 적용한다.
    """
    has_explicit_sub2 = any(len(row) == 5 and str(row[1]) == "Sub2" for row in gantt_rows)
    if has_explicit_sub2:
        return list(gantt_rows)

    indexed_sub_tasks = defaultdict(list)
    for idx, row in enumerate(gantt_rows):
        if len(row) != 5:
            continue
        job_id, line_name, proc, start, end = row
        if str(line_name) != "Sub":
            continue
        s = safe_ct(start, 0.0)
        e = safe_ct(end, s)
        indexed_sub_tasks[int(job_id)].append((idx, str(proc), s, e))

    sub2_indices = set()
    for job_id, tasks in indexed_sub_tasks.items():
        tasks.sort(key=lambda x: (x[2], x[3], x[0]))
        fit_positions = [pos for pos, (_idx, proc, _s, _e) in enumerate(tasks) if proc in DISPLAY_SUB_SPLIT_PROCS]
        if len(fit_positions) < 2:
            continue
        split_pos = fit_positions[1]
        for idx, _proc, _s, _e in tasks[split_pos:]:
            sub2_indices.add(idx)

    display_rows = []
    for idx, row in enumerate(gantt_rows):
        if len(row) != 5:
            display_rows.append(row)
            continue
        job_id, line_name, proc, start, end = row
        if idx in sub2_indices and str(line_name) == "Sub":
            display_rows.append((job_id, "Sub2", proc, start, end))
        else:
            display_rows.append(row)
    return display_rows


# 초보자 설명: 차트 그리기용 간트 행을 구성한다.
def build_chart_gantt_rows(gantt_rows):
    """간트 차트 표시용 gantt 라인을 정리한다.

    ProcessTime은 실제 기록 라인(Sub/Sub2)을 그대로 유지한다.
    반면 차트에서는 R2tig/R3tig/R14tig bar를 실제 용접 수행 위치인
    M1(Pipe) 라인에 표시한다.

    단, S1/S2에서 이어지는 합류 점선은 기존처럼 해당 tig 공정 시작점으로
    연결되어야 하므로, 차트용 라인 이동은 표시 단계에서만 적용한다.
    """
    display_rows = build_display_gantt_rows(gantt_rows)
    tig_procs = set(ROBOT_TIG_DISPLAY.values())
    chart_rows = []
    for row in display_rows:
        if len(row) != 5:
            chart_rows.append(row)
            continue
        job_id, line_name, proc, start, end = row
        if str(proc) in tig_procs and str(line_name) in ["Sub", "Sub2"]:
            chart_rows.append((job_id, "Pipe", proc, start, end))
        else:
            chart_rows.append(row)
    return chart_rows


# 초보자 설명: 간트 행을 원하는 순서로 정렬하기 위한 기준이다.
def _line_order_key(line_name):
    return (DISPLAY_LINE_ORDER.get(line_name, 99), str(line_name))


# 초보자 설명: 엑셀 템플릿 행의 서식을 다른 행으로 복사한다.
def _copy_template_row_style(ws, template_row_idx, target_row_idx, max_col):
    src_dim = ws.row_dimensions[template_row_idx]
    if src_dim.height is not None:
        ws.row_dimensions[target_row_idx].height = src_dim.height

    for col_idx in range(1, max_col + 1):
        src_cell = ws.cell(template_row_idx, col_idx)
        dst_cell = ws.cell(target_row_idx, col_idx)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            dst_cell.number_format = src_cell.number_format
        if src_cell.font:
            dst_cell.font = copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)


# 초보자 설명: 공정 시간을 여러 줄 형식의 엑셀로 저장한다.
def save_process_time_multirow(output_path, full_plan_rows, gantt_rows):
    recreate_output_file(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers_top = [
        "번호", "Spool No.", "Inch", "Sch.", "Length1", "Length2", "피팅1", "피팅2", "스풀타입", "용접장", "공정 시간 로그"
    ]
    for col_idx, value in enumerate(headers_top, start=1):
        ws.cell(1, col_idx).value = value

    for col_idx in range(1, 11):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
    ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=33)
    ws.cell(2, 11).value = "Log1"
    for i in range(12, 34):
        ws.cell(2, i).value = f"Log{i - 10}"

    default_widths = {
        1: 8, 2: 24, 3: 10, 4: 10, 5: 12, 6: 12, 7: 14, 8: 14, 9: 12, 10: 12
    }
    for col_idx, width in default_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(11, 34):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx in (1, 2):
        ws.row_dimensions[row_idx].height = 22

    display_gantt_rows = build_display_gantt_rows(gantt_rows)

    gantt_map = defaultdict(lambda: defaultdict(list))
    for job_id, line_name, proc, start, end in display_gantt_rows:
        s = safe_ct(start, 0.0)
        e = safe_ct(end, s)
        gantt_map[int(job_id)][str(line_name)].append((str(proc), s, e))

    for job_id in gantt_map:
        for line_name in gantt_map[job_id]:
            gantt_map[job_id][line_name].sort(key=lambda x: (x[1], x[2], x[0]))

    base_result_col = 11  # K
    current_row = 3

    for full_row in full_plan_rows:
        job_no = int(full_row[0])
        spool_no = full_row[1] if len(full_row) > 1 else None
        inch = full_row[2] if len(full_row) > 2 else None
        sch = full_row[3] if len(full_row) > 3 else None
        prod = full_row[4] if len(full_row) > 4 else None
        weld = full_row[5] if len(full_row) > 5 else None
        length1 = full_row[6] if len(full_row) > 6 else None
        length2 = full_row[7] if len(full_row) > 7 else None
        fit1 = full_row[8] if len(full_row) > 8 else None
        fit2 = full_row[9] if len(full_row) > 9 else None

        line_dict = gantt_map.get(job_no, {})
        ordered_lines = sorted(line_dict.keys(), key=_line_order_key)
        if not ordered_lines:
            ordered_lines = ["Pipe"]
            line_dict = {"Pipe": []}

        line_rows = []
        for line_name in ordered_lines:
            tasks = line_dict.get(line_name, [])
            formatted = [
                f"{proc}: {_format_process_time_value(start)} - {_format_process_time_value(end)}"
                for proc, start, end in tasks
            ]
            line_rows.append(formatted)

        block_rows = max(1, len(line_rows))
        for offset in range(block_rows):
            target_row = current_row + offset
            ws.row_dimensions[target_row].height = 20

        ws.cell(current_row, 1).value = job_no
        ws.cell(current_row, 2).value = spool_no
        ws.cell(current_row, 3).value = inch
        ws.cell(current_row, 4).value = sch
        ws.cell(current_row, 5).value = length1
        ws.cell(current_row, 6).value = length2
        ws.cell(current_row, 7).value = fit1
        ws.cell(current_row, 8).value = fit2
        ws.cell(current_row, 9).value = prod
        ws.cell(current_row, 10).value = weld

        for row_offset, formatted_tasks in enumerate(line_rows):
            target_row = current_row + row_offset
            for idx, text in enumerate(formatted_tasks):
                if base_result_col + idx <= 33:
                    ws.cell(target_row, base_result_col + idx).value = text
                else:
                    ws.cell(target_row, base_result_col + idx).value = text
                    ws.column_dimensions[get_column_letter(base_result_col + idx)].width = 18

        if block_rows > 1:
            for col_idx in range(1, 11):
                ws.merge_cells(start_row=current_row, start_column=col_idx, end_row=current_row + block_rows - 1,
                               end_column=col_idx)

        current_row += block_rows

    wb.save(output_path)


# =====================================================
# 시뮬레이션 실행 파이프라인
# -----------------------------------------------------
# 여기서는 plan 전체를 읽어 SimPy 환경을 돌리고
# 간트, makespan, ProcessTime 결과를 최종적으로 계산한다.
# =====================================================

# 초보자 설명: 현재 plan 전체를 SimPy로 실행하고 간트/공정시간 결과를 만든다.
def run_simulation(current_plan, output_png=None, render_chart=True, combined_weld_hold=True,
                   parallel_sub_products=True):
    global plan, idleTimeR2, gantt, COMBINED_WELD_HOLD_ENABLED, PARALLEL_SUB_MULTI_PRODUCTS_ENABLED

    COMBINED_WELD_HOLD_ENABLED = bool(combined_weld_hold)
    PARALLEL_SUB_MULTI_PRODUCTS_ENABLED = bool(parallel_sub_products)
    plan = renumber_plan_rows(current_plan)
    reset_runtime_state(plan)
    rebuild_station_fifo_tokens(plan)
    rebuild_sub_start_fifo_tokens(plan)

    for row in plan:
        if len(row) < 8:
            continue
        no, name, inch, Sch, p, w, l, ft = row[:8]
        forced_cut_start = row[8] if len(row) > 8 else None
        env.process(job(env, no, name, inch, Sch, p, w, l, ft, forced_cut_start))

    env.run()

    weld_proc_set = {"2D-1", "2D-2", "3D", "Short"}
    inch_by_job = {}
    joint_by_job = defaultdict(int)
    for row in plan:
        if len(row) < 3:
            continue
        job_no = row[0]
        inch_val = safe_ct(row[2], 0.0)
        inch_by_job[job_no] = inch_val
        counts = count_station_uses(row)
        joint_by_job[job_no] = sum(int(counts.get(proc, 0)) for proc in weld_proc_set)

    JointNum = sum(joint_by_job.values())
    Dia = sum(joint_by_job[job_id] * inch_by_job.get(job_id, 0.0) for job_id in joint_by_job)
    scrap_pct = 0.0
    idleTimeR2 = extract_idle_time_r2(gantt, min_idle=2.0)

    if not render_chart:
        return {"gantt": list(gantt), "idleTimeR2": list(idleTimeR2), "plan": list(plan)}

    chart_gantt = build_chart_gantt_rows(gantt)

    START_TIME = 0
    fig_h = max(10, len(plan) * 0.56)
    fig, ax = plt.subplots(figsize=(18, fig_h))

    # 초보자 설명: 현재 문맥에 필요한 값이나 이름을 계산해서 돌려주는 함수다.
    def get_line_y(job_id, line):
        base_y = (int(job_id) - 1) * 4
        if line == "Pipe":
            return base_y
        if line == "Pipe2":
            return base_y + 1
        if line == "Sub":
            return base_y + 2
        return base_y + 3

    # -------------------------------------------------
    # bar 그리기
    # - 개별 barh 반복 호출 대신 broken_barh 그룹화로 렌더링 비용을 줄인다.
    # -------------------------------------------------
    broken_bar_groups = defaultdict(lambda: {"xranges": [], "colors": []})
    for job_id, line, proc, start, end in chart_gantt:
        s = safe_ct(start, 0.0)
        e = safe_ct(end, s)
        if e < s:
            e = s
        y = get_line_y(job_id, line)
        group = broken_bar_groups[y]
        group["xranges"].append((s + START_TIME, e - s))
        group["colors"].append(COLOR.get(proc, "gray"))

    for y, group in broken_bar_groups.items():
        ax.broken_barh(
            group["xranges"],
            (y - 0.35, 0.7),
            facecolors=group["colors"],
            edgecolors="white",
            linewidth=0.6,
        )

    # -------------------------------------------------
    # 작업맵 생성
    # -------------------------------------------------
    task_map = defaultdict(lambda: defaultdict(list))
    for job_id, line, proc, start, end in chart_gantt:
        s = safe_ct(start, 0.0)
        e = safe_ct(end, s)
        task_map[job_id][line].append((proc, s, e))

    for job_id in task_map:
        for line_name in task_map[job_id]:
            task_map[job_id][line_name].sort(key=lambda x: x[1])

    # -------------------------------------------------
    # main / fitting flow 점선 연결
    # main flow = 붉은 점선
    # sub flow = 파란 점선
    # -------------------------------------------------
    # 초보자 설명: 'draw_dashed_horizontal' 단계의 처리를 맡는 보조 함수다.
    def draw_dashed_horizontal(x_from, x_to, y, color, linewidth=0.4, zorder=5, dash_len=2.0, gap_len=1.5):
        x1 = safe_ct(x_from, 0.0)
        x2 = safe_ct(x_to, x1)
        if x2 < x1:
            x1, x2 = x2, x1
        if abs(x2 - x1) < 1e-9:
            return
        ax.plot(
            [x1 + START_TIME, x2 + START_TIME],
            [y, y],
            linestyle=(0, (dash_len, gap_len)),
            linewidth=linewidth,
            color=color,
            zorder=zorder,
            dash_capstyle="butt",
            solid_capstyle="butt",
        )

    # 초보자 설명: 'draw_dashed_vertical' 단계의 처리를 맡는 보조 함수다.
    def draw_dashed_vertical(x, y_from, y_to, color, linewidth=0.4, zorder=6, dash_len=0.09, gap_len=0.06):
        yy1 = float(y_from)
        yy2 = float(y_to)
        if yy2 < yy1:
            yy1, yy2 = yy2, yy1
        if abs(yy2 - yy1) < 1e-9:
            return
        ax.plot(
            [x + START_TIME, x + START_TIME],
            [yy1, yy2],
            linestyle=(0, (dash_len, gap_len)),
            linewidth=linewidth,
            color=color,
            zorder=zorder,
            dash_capstyle="butt",
            solid_capstyle="butt",
        )

    for job_id, line_dict in task_map.items():
        for line_name, tasks in line_dict.items():
            if len(tasks) < 2:
                continue

            color = "red" if line_name in ["Pipe", "Pipe2"] else "blue"
            y = get_line_y(job_id, line_name)

            for i in range(len(tasks) - 1):
                _, s1, e1 = tasks[i]
                _, s2, e2 = tasks[i + 1]
                draw_dashed_horizontal(e1, s2, y, color=color, linewidth=0.4, zorder=5)

    # -------------------------------------------------
    # fitting -> main 합류선 ('ㄴ'자)
    # - fitting flow의 제품이 2개인 경우:
    #   첫번째 제품은 첫번째 용접장으로,
    #   두번째 제품은 main flow의 두번째 용접장으로 연결
    # -------------------------------------------------
    # 초보자 설명: 현재 문맥에 필요한 값이나 이름을 계산해서 돌려주는 함수다.
    def get_nth_task(line_tasks, proc_name, n=1):
        cnt = 0
        for t in line_tasks:
            if t[0] == proc_name:
                cnt += 1
                if cnt == n:
                    return t
        return None

    # 초보자 설명: 조건에 맞는 행이나 값을 찾아 반환하는 함수다.
    def find_target_task_for_job(job_no, proc_name, preferred_lines=("Pipe", "Pipe2"), nth=1):
        for ln in preferred_lines:
            if ln in task_map[job_no]:
                t = get_nth_task(task_map[job_no][ln], proc_name, nth)
                if t is not None:
                    return ln, t
        return None, None

    # 초보자 설명: CycleTime 값을 찾거나 해석하는 데 쓰는 보조 함수다.
    def draw_l_connect(x_from, y_from, x_to, y_to, color, linewidth=0.4, zorder=6):
        x_from = safe_ct(x_from, 0.0)
        x_to = safe_ct(x_to, x_from)
        if abs(x_to - x_from) < 0.25:
            elbow_x = x_to + 0.8
            draw_dashed_horizontal(x_from, elbow_x, y_from, color=color, linewidth=linewidth, zorder=zorder)
            draw_dashed_vertical(elbow_x, y_from, y_to, color=color, linewidth=linewidth, zorder=zorder)
            draw_dashed_horizontal(elbow_x, x_to, y_to, color=color, linewidth=linewidth, zorder=zorder)
            return
        draw_dashed_horizontal(x_from, x_to, y_from, color=color, linewidth=linewidth, zorder=zorder)
        draw_dashed_vertical(x_to, y_from, y_to, color=color, linewidth=linewidth, zorder=zorder)

    for row in plan:
        if len(row) < 6:
            continue

        jNo, name, inch, Sch, prod, weld = row[:6]

        if jNo not in task_map:
            continue
        merge_targets = []

        if prod == "2D":
            if weld in ["2D-1", "2D-2"]:
                merge_targets.append((1, "R2tig", "Pipe", 1))
            elif weld == "3D":
                merge_targets.append((1, "R3tig", "Pipe", 1))
        elif prod == "F2D":
            merge_targets.append((1, "R2tig", "Pipe", 1))
            merge_targets.append((2, "R3tig", "Pipe", 1))
        elif prod == "H3D":
            merge_targets.append((1, "R2tig", "Pipe", 1))
        elif prod == "F3D":
            merge_targets.append((1, "R2tig", "Pipe", 1))
            merge_targets.append((2, "R2", "Pipe2", 2))
        elif prod in ["F2DS", "F2DL"]:
            merge_targets.append((1, "R14tig", "Pipe", 1))
            merge_targets.append((2, "R14tig", "Pipe", 2))
        elif prod == "FF":
            merge_targets.append((1, "Pick1", "Pipe", 1))
            merge_targets.append((2, "Pick1", "Pipe", 1))
        elif prod == "2DL":
            merge_targets.append((1, "Pick1", "Pipe", 1))
        elif prod in ["Short", "2DS"]:
            merge_targets.append((1, "R14tig", "Pipe", 1))

        for sub_idx, target_proc, target_line_hint, target_nth in merge_targets:
            source_sub_line = "Sub" if sub_idx == 1 else "Sub2"
            sub_tasks = task_map[jNo].get(source_sub_line, [])
            if not sub_tasks:
                continue

            # 합류선은 해당 sub 라인의 마지막 공정 끝점에서 시작해야
            # R2/R3/R14 -> R*tig 또는 Short 와 같은 실제 합류 위치가 정확히 보인다.
            sub_task = sub_tasks[-1]
            _, sub_s, sub_e = sub_task
            y_sub = get_line_y(jNo, source_sub_line)

            main_line, main_task = find_target_task_for_job(
                jNo,
                target_proc,
                preferred_lines=(target_line_hint, "Pipe2"),
                nth=target_nth,
            )

            if main_task is None:
                continue

            _, main_s, main_e = main_task
            y_main = get_line_y(jNo, main_line)

            draw_l_connect(
                x_from=sub_e,
                y_from=y_sub,
                x_to=main_s,
                y_to=y_main,
                color="blue",
                linewidth=0.4,
                zorder=6,
            )

    # -------------------------------------------------
    # Pipe2 -> main 두번째 용접장 합류선 ('ㄴ'자)
    # - H3D / F3D의 경우 두번째 제품(Pipe2)도
    #   최종적으로 main flow의 두번째 용접장(3D)에서 합류됨
    # -------------------------------------------------
    for row in plan:
        if len(row) < 6:
            continue

        jNo, name, inch, Sch, prod, weld = row[:6]

        if prod not in ["H3D", "F3D"]:
            continue
        if jNo not in task_map:
            continue
        if "Pipe2" not in task_map[jNo]:
            continue
        if len(task_map[jNo]["Pipe2"]) == 0:
            continue
        if "Pipe" not in task_map[jNo]:
            continue

        pipe2_last = task_map[jNo]["Pipe2"][-1]
        _, p2_s, p2_e = pipe2_last

        target_proc = "Angle" if prod in ["H3D", "F3D"] else "3D"
        main_line, main_task = find_target_task_for_job(jNo, target_proc, preferred_lines=("Pipe",), nth=1)
        if main_task is None:
            continue

        _, main_s, main_e = main_task
        y_pipe2 = get_line_y(jNo, "Pipe2")
        y_main = get_line_y(jNo, main_line)

        draw_l_connect(
            x_from=p2_e,
            y_from=y_pipe2,
            x_to=main_s,
            y_to=y_main,
            color="red",
            linewidth=0.6,
            zorder=6,
        )

    # -------------------------------------------------
    # job 구분선 / 축
    # -------------------------------------------------
    minor_hlines = []
    for job_id in range(1, len(plan) + 1):
        for line_name in ["Pipe", "Pipe2", "Sub", "Sub2"]:
            minor_hlines.append(get_line_y(job_id, line_name) + 0.5)
    if minor_hlines:
        ax.hlines(minor_hlines, xmin=START_TIME, xmax=max((end for _, _, _, _, end in gantt), default=0) + START_TIME,
                  colors="#d0d0d0", linewidth=0.35, zorder=0)

    major_hlines = [i * 4 - 0.5 for i in range(0, len(plan) + 1)]
    if major_hlines:
        ax.hlines(major_hlines, xmin=START_TIME, xmax=max((end for _, _, _, _, end in gantt), default=0) + START_TIME,
                  colors="black", linewidth=0.8)

    ax.set_yticks([((i - 1) * 4) + 1.5 for i in range(1, len(plan) + 1)])

    job_labels = []
    for i, row in enumerate(plan, start=1):
        job_no = row[0] if len(row) > 0 and row[0] is not None else i
        prod_val = str(row[4]).strip() if len(row) > 4 and row[4] is not None else "-"
        weld_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
        inch_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "-"
        ft_val = str(row[7]).strip() if len(row) > 7 and row[7] is not None else "-"
        weld_suffix = get_weld_station_label(prod_val, weld_val)
        job_labels.append(f"{job_no}. {prod_val}{weld_suffix}, {inch_val}in, {ft_val}")

    ax.set_yticklabels(job_labels)
    ax.tick_params(axis="y", labelsize=7)
    ax.invert_yaxis()
    ax.set_ylim(len(plan) * 4 - 0.5, -0.8)

    ax.set_xlabel("Time")
    ax.set_ylabel("Production Jobs")

    # 초보자 설명: 'format_time' 단계의 처리를 맡는 보조 함수다.
    def format_time(x, pos):
        hour = int(x // 60)
        minute = int(x % 60)
        return f"{hour:02d}:{minute:02d}"

    ax.xaxis.set_major_formatter(FuncFormatter(format_time))

    used_set = set([g[2] for g in chart_gantt])
    used = [k for k in COLOR.keys() if k in used_set]
    handles = [plt.Rectangle((0, 0), 1, 1, color=COLOR.get(u, "gray")) for u in used]
    legend = ax.legend(
        handles,
        used,
        loc="upper left",
        bbox_to_anchor=(1.01, 1.00),
        borderaxespad=0.0,
        fontsize=9,
    )

    makespan_minutes = max((end for _, _, _, _, end in gantt), default=0)
    makespan_total_minutes = int(round(makespan_minutes))
    makespan_hh = makespan_total_minutes // 60
    makespan_mm = makespan_total_minutes % 60
    info_text = (
        f"joint : {JointNum}\n"
        f"Dia inch : {Dia}\n"
        f"Scrap : {scrap_pct:.2f}%\n"
        f"make span : {makespan_hh:02d}:{makespan_mm:02d}"
    )
    info_box = AnchoredText(
        info_text,
        loc="upper left",
        prop=dict(size=9),
        frameon=True,
        bbox_to_anchor=(1.01, 0.40),
        bbox_transform=ax.transAxes,
        borderpad=0.0,
        pad=0.4,
    )
    info_box.patch.set_facecolor(legend.get_frame().get_facecolor())
    info_box.patch.set_edgecolor(legend.get_frame().get_edgecolor())
    info_box.patch.set_linewidth(legend.get_frame().get_linewidth())
    if legend.get_frame().get_alpha() is not None:
        info_box.patch.set_alpha(legend.get_frame().get_alpha())
    info_box.txt._text.set_fontsize(9)
    info_box.txt._text.set_multialignment("left")
    ax.add_artist(info_box)

    plt.tight_layout()
    fig.canvas.draw()

    base_dpi = float(fig.dpi)
    target_dpi = 150.0 if output_png is not None else base_dpi
    dpi_scale = target_dpi / max(1.0, base_dpi)
    fig_w_px = float(fig.get_figwidth()) * target_dpi
    fig_h_px = float(fig.get_figheight()) * target_dpi
    ax_bbox = ax.get_window_extent(renderer=fig.canvas.get_renderer())
    plot_box = {
        "left": float(ax_bbox.x0) * dpi_scale,
        "right": float(ax_bbox.x1) * dpi_scale,
        "top": float(fig.canvas.get_width_height()[1] - ax_bbox.y1) * dpi_scale,
        "bottom": float(fig.canvas.get_width_height()[1] - ax_bbox.y0) * dpi_scale,
    }
    job_bands = []
    canvas_h_px = float(fig.canvas.get_width_height()[1])
    for idx in range(len(plan)):
        y_top_data = idx * 4 - 0.5
        y_bottom_data = idx * 4 + 3.5
        _x1, y_top_disp = ax.transData.transform((0.0, y_top_data))
        _x2, y_bottom_disp = ax.transData.transform((0.0, y_bottom_data))
        top_px = (canvas_h_px - max(y_top_disp, y_bottom_disp)) * dpi_scale
        bottom_px = (canvas_h_px - min(y_top_disp, y_bottom_disp)) * dpi_scale
        job_bands.append({"index": idx, "top": float(top_px), "bottom": float(bottom_px)})

    chart_meta = {
        "image_size": [float(fig_w_px), float(fig_h_px)],
        "plot_box": plot_box,
        "job_bands": job_bands,
    }

    if output_png is not None:
        recreate_output_file(output_png)
        plt.savefig(output_png, dpi=150)
    if output_png is None:
        plt.show()
    plt.close()

    return {"gantt": list(gantt), "idleTimeR2": list(idleTimeR2), "plan": list(plan), "chart_meta": chart_meta}


# 초보자 설명: CycleTime과 Product 입력을 함께 초기화한다.
def initialize_inputs(paths: AppPaths) -> None:
    """엑셀 입력을 로드하고, plan 파이프라인에 필요한 전역 상태를 초기화한다."""
    configure_paths(paths)
    debug_log(f"initialize_inputs 시작 | cycle_time={ct_path} | product={product_path}", preferred_dir=paths.base_dir)
    cycle_data = load_cycle_time_data(ct_path)
    apply_cycle_time_data(cycle_data)
    product_plan_data = load_product_plan_data(product_path)
    apply_product_plan_data(product_plan_data)
    debug_log("initialize_inputs 완료", preferred_dir=paths.base_dir)

# 초보자 설명: CycleTime만 먼저 읽어 초기화한다.
def initialize_cycle_time_only(paths: AppPaths) -> None:
    """수동 계획 모드용 최소 초기화: CycleTime만 읽고 cycle_index를 만든다."""
    configure_paths(paths)
    debug_log(f"initialize_cycle_time_only 시작 | cycle_time={ct_path}", preferred_dir=paths.base_dir)
    cycle_data = load_cycle_time_data(ct_path)
    apply_cycle_time_data(cycle_data)
    debug_log("initialize_cycle_time_only 완료", preferred_dir=paths.base_dir)

# 초보자 설명: 규칙 기반 1차 계획 결과를 만든다.
def build_rulebase_first_result(paths: AppPaths) -> Tuple[Dict[str, object], Dict[str, object]]:
    """Rule-base 1차 probe 결과를 만든다.

    반환값:
    - selected_first_result: legacy probe가 성공하면 legacy 결과, 아니면 current 결과
    - current_first_result: 현재 로직 기준 1차 결과
    """
    current_first_result = run_simulation(
        srtPlan_sim_rows,
        output_png=None,
        render_chart=False,
        combined_weld_hold=False,
        parallel_sub_products=False,
    )
    legacy_probe = run_legacy_short_placement_probe(paths)
    return legacy_probe or current_first_result, current_first_result


# 초보자 설명: 1차 결과를 바탕으로 최종 규칙 기반 plan을 만든다.
def build_rulebase_final_plan(first_result: Dict[str, object]):
    first_idle_r2 = first_result["idleTimeR2"]
    plan2_forced = apply_forced_cut_starts_to_plan2(plan2_sim_rows, first_idle_r2)
    return combine_final_plan_by_start_time(srtPlan_sim_rows, plan2_forced, first_result["gantt"])


# 초보자 설명: 기존 방식의 Short 배치 탐색 시뮬레이션을 수행한다.
def run_legacy_short_placement_probe(paths: AppPaths) -> Optional[Dict[str, object]]:
    """repaired 중간파일 생성을 막기 위해 legacy embedded probe는 사용하지 않는다."""
    print("[INFO] legacy embedded Short probe 비활성화: current probe 사용")
    return None


# 초보자 설명: 수동 계획 파일을 사용할 때의 출력 경로를 만든다.
def build_paths_for_manual_plan(manual_plan_path: str, cycle_time_path: Optional[str] = None,
                                output_base_dir: Optional[str] = None) -> AppPaths:
    """수동 계획 모드용 경로 생성. Product.xlsx 없이도 사용할 수 있다."""
    manual_plan_path = os.path.abspath(manual_plan_path)
    base = output_base_dir or os.path.dirname(manual_plan_path) or os.getcwd()
    cycle_src = os.path.abspath(cycle_time_path) if cycle_time_path else os.path.join(base, "CycleTime.xlsx")
    return AppPaths(
        base_dir=base,
        cycle_time=cycle_src,
        product="",
        plan=os.path.join(base, "Plan_manual.xlsx"),
        result_png=os.path.join(base, "result_manual.png"),
        process_time=os.path.join(base, "ProcessTime_manual.xlsx"),
        ga_result_png=os.path.join(base, "result_GA.png"),
        ga_plan=os.path.join(base, "Plan_GA.xlsx"),
        ga_process_time=os.path.join(base, "ProcessTime_GA.xlsx"),
    )


# 초보자 설명: 수동 입력 plan을 시뮬레이션용 내부 형식으로 바꾼다.
def convert_manual_plan_rows_to_sim_rows(full_rows):
    """Plan_man.xlsx의 10열 계획표를 시뮬레이션용 9열 plan 형식으로 변환한다."""
    sim_rows = []
    spool_fit_detail_local = {}
    for idx, row in enumerate(full_rows, start=1):
        rr = list(tuple(auto_convert(cell) for cell in row))
        if len(rr) < 10:
            rr += [None] * (10 - len(rr))

        spool = rr[1]
        inch = rr[2]
        sch = rr[3]
        prod = norm_text(rr[4])
        weld = norm_text(rr[5])
        length1 = rr[6]
        length2 = rr[7]
        fit1 = norm_text(rr[8]) if not is_zero_mark(rr[8]) else ""
        fit2 = norm_text(rr[9]) if not is_zero_mark(rr[9]) else ""

        spool_fit_detail_local[spool] = (fit1, fit2)

        if prod == "2D" and weld == "":
            weld = "2D-1"
        elif prod in ["F2D", "H3D", "F3D"] and weld == "":
            weld = "3D"
        elif prod in ["2DS", "2DL", "Short", "F2DS", "F2DL", "FF", "ff"] and weld == "":
            weld = "Short"

        rep_fit = decide_rep_fitting(rr[8], rr[9])
        length = decide_internal_length(prod, length1, length2)
        sim_rows.append((idx, spool, inch, sch, prod, weld, length, rep_fit, None))

    return renumber_plan_rows(sim_rows), spool_fit_detail_local


# 초보자 설명: 사용자가 만든 수동 계획 엑셀을 읽는다.
def load_manual_plan_from_excel(manual_plan_path: str, cycle_time_path: Optional[str] = None) -> Dict[str, object]:
    """수동 계획 파일을 읽어 기존 프로그램과 같은 메모리 구조에 적재한다."""
    global plan, spool_fit_detail, spool_length_pair, CYCLE_DICT_CACHE

    active_paths = build_paths_for_manual_plan(
        manual_plan_path=manual_plan_path,
        cycle_time_path=cycle_time_path,
        output_base_dir=os.path.dirname(os.path.abspath(manual_plan_path)) or os.getcwd(),
    )

    if not os.path.exists(active_paths.cycle_time):
        raise FileNotFoundError(
            f"CycleTime.xlsx를 찾을 수 없습니다.\n수동 계획 파일과 같은 폴더에 CycleTime.xlsx를 두어 주세요.\n현재 탐색 경로: {active_paths.cycle_time}"
        )

    initialize_cycle_time_only(active_paths)

    manual_rows, f_plan, manual_sheet = open_xlsx_reader_auto(manual_plan_path, skip_header=True)
    f_plan.close()
    print(f"[INFO] Plan_man.xlsx sheet = {manual_sheet}")

    if not manual_rows:
        raise RuntimeError("Plan_man.xlsx에 생산계획 데이터가 없습니다.")

    sim_rows, spool_fit_detail_local = convert_manual_plan_rows_to_sim_rows(manual_rows)
    spool_fit_detail = spool_fit_detail_local
    spool_length_pair = {
        norm_text(row[1]): (to_num(row[6], 0.0), to_num(row[7], 0.0))
        for row in (tuple(auto_convert(cell) for cell in r) for r in manual_rows)
        if len(row) >= 8 and norm_text(row[1]) != ""
    }
    CYCLE_DICT_CACHE.clear()
    plan = list(sim_rows)

    return {
        "mode": "manual",
        "paths": active_paths,
        "manual_plan_path": os.path.abspath(manual_plan_path),
        "final_plan": sim_rows,
        "plan_full_rows": build_full_format_plan_rows(sim_rows),
    }


GA_DEFAULT_POPULATION = 16
GA_DEFAULT_GENERATIONS = None  # 전체 Job 수의 절반으로 동적 결정
GA_DEFAULT_ELITE = 2
GA_DEFAULT_MUTATION_RATE = 0.18
GA_DEFAULT_TOURNAMENT = 3


# 초보자 설명: 간트 기록으로부터 전체 make span을 계산한다.
def compute_makespan(gantt_rows):
    if not gantt_rows:
        return 0.0
    return max(safe_ct(end, 0.0) for _, _, _, _, end in gantt_rows)


# =====================================================
# GA(유전 알고리즘) 최적화
# -----------------------------------------------------
# 규칙 기반 plan 대신, Job 순서와 2D 용접장 배정을 바꾸어 가며
# make span이 더 좋아지는 조합을 탐색하는 부분이다.
# =====================================================

# 초보자 설명: 임시 계획 행을 GA용 개별 Job 표현으로 바꾼다.
def convert_tmp_row_to_ga_plan_row(tmp_row, assigned_weld=None):
    converted_row = tuple(auto_convert(cell) for cell in tmp_row)
    internal_prod = decide_internal_prod(converted_row[4])
    rep_fit = decide_rep_fitting(converted_row[8], converted_row[9])
    length = decide_internal_length(internal_prod, converted_row[6], converted_row[7])

    if internal_prod == "2D":
        plan_weld = choose_valid_2d_weld_for_tmp_row(converted_row, preferred=assigned_weld or "2D-1")
    elif internal_prod in ["F2D", "H3D", "F3D"]:
        plan_weld = "3D"
    elif internal_prod in ["2DS", "2DL", "Short", "F2DS", "F2DL", "FF", "ff"]:
        plan_weld = "Short"
    else:
        plan_weld = converted_row[5]

    return (
        0,
        converted_row[1],
        converted_row[2],
        converted_row[3],
        internal_prod,
        plan_weld,
        length,
        rep_fit,
        None,
    )


# 초보자 설명: GA가 다룰 기본 Job 목록을 만든다.
def build_ga_base_jobs(tmp_rows):
    base_jobs = []
    two_d_gene_positions = []
    jobs_by_inch = defaultdict(list)
    jobs_by_inch_and_type = defaultdict(list)
    inch_type_order = defaultdict(list)

    for idx, row in enumerate(tmp_rows):
        prod = norm_text(row[4])
        inch = _inch_to_float(row[2])
        info = {
            "idx": idx,
            "tmp_row": tuple(row),
            "prod": prod,
            "spool": row[1],
            "inch": inch,
            "valid_welds": get_valid_2d_welds_for_tmp_row(row) if prod == "2D" else [],
        }
        base_jobs.append(info)
        jobs_by_inch[inch].append(idx)
        jobs_by_inch_and_type[(inch, prod)].append(idx)
        if prod not in inch_type_order[inch]:
            inch_type_order[inch].append(prod)
        if prod == "2D":
            two_d_gene_positions.append(idx)

    ga_meta = {
        "inches": list(jobs_by_inch.keys()),
        "types_by_inch": {inch: list(types) for inch, types in inch_type_order.items()},
        "jobs_by_inch": {inch: list(indices) for inch, indices in jobs_by_inch.items()},
        "jobs_by_inch_and_type": {key: list(indices) for key, indices in jobs_by_inch_and_type.items()},
    }
    return base_jobs, two_d_gene_positions, ga_meta


# 초보자 설명: 염색체 표현을 실제 Job 순서 정보로 해석한다.
def chromosome_to_gene_order(chromosome, ga_meta=None, inch_group_optimization=False):
    if not inch_group_optimization:
        return list(chromosome["order"])

    if ga_meta is None:
        raise RuntimeError("Inch별 GA 해석을 위한 메타 정보가 없습니다.")

    gene_order = []
    inch_order = chromosome.get("inch_order", [])
    type_orders = chromosome.get("type_orders", {})
    gene_orders = chromosome.get("gene_orders", {})

    for inch in inch_order:
        for prod in type_orders.get(inch, []):
            gene_order.extend(gene_orders.get((inch, prod), []))
    return gene_order


# 초보자 설명: 염색체를 시뮬레이션 가능한 plan으로 변환한다.
def chromosome_to_plan(chromosome, base_jobs, ga_meta=None, inch_group_optimization=False):
    rows = []
    weld_genes = chromosome["welds"]
    gene_order = chromosome_to_gene_order(chromosome, ga_meta=ga_meta, inch_group_optimization=inch_group_optimization)
    for gene_idx in gene_order:
        job_info = base_jobs[gene_idx]
        assigned_weld = weld_genes.get(gene_idx)
        rows.append(convert_tmp_row_to_ga_plan_row(job_info["tmp_row"], assigned_weld=assigned_weld))
    return renumber_plan_rows(rows)


# 초보자 설명: 초기 세대용 무작위 염색체를 만든다.
def create_random_chromosome(base_jobs, two_d_gene_positions, ga_meta=None, inch_group_optimization=False):
    welds = {}
    for idx in two_d_gene_positions:
        valid_welds = list(base_jobs[idx].get("valid_welds", [])) or ["2D-1", "2D-2", "3D"]
        welds[idx] = random.choice(valid_welds)

    if not inch_group_optimization:
        order = list(range(len(base_jobs)))
        random.shuffle(order)
        return {"order": order, "welds": welds}

    if ga_meta is None:
        raise RuntimeError("Inch별 GA 생성을 위한 메타 정보가 없습니다.")

    inch_order = list(ga_meta["inches"])
    random.shuffle(inch_order)
    type_orders = {}
    gene_orders = {}
    for inch in inch_order:
        prod_list = list(ga_meta["types_by_inch"].get(inch, []))
        random.shuffle(prod_list)
        type_orders[inch] = prod_list
        for prod in prod_list:
            key = (inch, prod)
            genes = list(ga_meta["jobs_by_inch_and_type"].get(key, []))
            random.shuffle(genes)
            gene_orders[key] = genes

    return {
        "inch_order": inch_order,
        "type_orders": type_orders,
        "gene_orders": gene_orders,
        "welds": welds,
    }


# 초보자 설명: 염색체를 깊은 복사해 안전하게 수정할 수 있게 한다.
def clone_chromosome(chromosome):
    cloned = {"welds": dict(chromosome.get("welds", {}))}
    if "order" in chromosome:
        cloned["order"] = list(chromosome["order"])
    if "inch_order" in chromosome:
        cloned["inch_order"] = list(chromosome["inch_order"])
    if "type_orders" in chromosome:
        cloned["type_orders"] = {inch: list(order) for inch, order in chromosome["type_orders"].items()}
    if "gene_orders" in chromosome:
        cloned["gene_orders"] = {key: list(order) for key, order in chromosome["gene_orders"].items()}
    return cloned


# 초보자 설명: 중복 개체 판별을 위해 염색체를 비교 가능한 문자열/튜플로 바꾼다.
def chromosome_signature(chromosome, ga_meta=None, inch_group_optimization=False):
    weld_items = tuple(sorted((int(k), str(v)) for k, v in chromosome.get("welds", {}).items()))
    if not inch_group_optimization:
        return ("basic", tuple(chromosome.get("order", [])), weld_items)

    inch_order = tuple(chromosome.get("inch_order", []))
    type_orders = tuple(
        (inch, tuple(chromosome.get("type_orders", {}).get(inch, [])))
        for inch in inch_order
    )
    gene_orders = []
    for inch in inch_order:
        for prod in chromosome.get("type_orders", {}).get(inch, []):
            key = (inch, prod)
            gene_orders.append((key, tuple(chromosome.get("gene_orders", {}).get(key, []))))
    return ("inch", inch_order, type_orders, tuple(gene_orders), weld_items)


# 초보자 설명: 염색체 하나를 plan으로 바꾸고 시뮬레이션해 성능을 계산한다.
def evaluate_chromosome(chromosome, base_jobs, ga_meta=None, inch_group_optimization=False):
    candidate_plan = chromosome_to_plan(chromosome, base_jobs, ga_meta=ga_meta,
                                        inch_group_optimization=inch_group_optimization)
    # GA의 fitness 계산과 사용자가 실제로 실행하는 생산 시뮬레이션은
    # 반드시 같은 조건을 사용해야 한다.
    #
    # 이전 코드에서는 GA 평가만 combined_weld_hold=False,
    # parallel_sub_products=False로 돌리고,
    # 실제 생산 시뮬레이션/최종 출력은 기본값(True/True)로 다시 돌렸다.
    # 그 결과 콘솔의 best makespan과 최종 실행 makespan이 서로 다른 문제가 생겼다.
    #
    # 따라서 GA 평가도 최종 실행과 동일한 시뮬레이션 조건으로 계산한다.
    result = run_simulation(
        candidate_plan,
        output_png=None,
        render_chart=False,
        combined_weld_hold=True,
        parallel_sub_products=True,
    )
    makespan = compute_makespan(result["gantt"])
    return makespan, result, candidate_plan


# 초보자 설명: GA에서 토너먼트 선택으로 부모를 고른다.
def tournament_select(scored_population, tournament_size):
    size = min(max(2, tournament_size), len(scored_population))
    sampled = random.sample(scored_population, size)
    sampled.sort(key=lambda x: x[0])
    return clone_chromosome(sampled[0][1])


# 초보자 설명: 순서형 문제에 맞는 Order Crossover를 수행한다.
def order_crossover(parent1_order, parent2_order):
    n = len(parent1_order)
    if n <= 1:
        return list(parent1_order)
    a, b = sorted(random.sample(range(n), 2))
    child = [None] * n
    child[a:b + 1] = parent1_order[a:b + 1]
    p2_filtered = [gene for gene in parent2_order if gene not in child]
    fill_positions = [i for i, gene in enumerate(child) if gene is None]
    for pos, gene in zip(fill_positions, p2_filtered):
        child[pos] = gene
    return child


# 초보자 설명: 두 부모를 교차시켜 자식 염색체를 만든다.
def crossover(parent1, parent2, two_d_gene_positions, ga_meta=None, inch_group_optimization=False):
    if not inch_group_optimization:
        child = {
            "order": order_crossover(parent1["order"], parent2["order"]),
            "welds": {},
        }
    else:
        if ga_meta is None:
            raise RuntimeError("Inch별 GA 교차를 위한 메타 정보가 없습니다.")

        child = {
            "inch_order": order_crossover(parent1["inch_order"], parent2["inch_order"]),
            "type_orders": {},
            "gene_orders": {},
            "welds": {},
        }
        for inch in child["inch_order"]:
            p1_types = parent1["type_orders"].get(inch, ga_meta["types_by_inch"].get(inch, []))
            p2_types = parent2["type_orders"].get(inch, ga_meta["types_by_inch"].get(inch, []))
            child_types = order_crossover(list(p1_types), list(p2_types)) if len(p1_types) >= 2 else list(p1_types)
            child["type_orders"][inch] = child_types
            for prod in child_types:
                key = (inch, prod)
                p1_genes = parent1["gene_orders"].get(key, ga_meta["jobs_by_inch_and_type"].get(key, []))
                p2_genes = parent2["gene_orders"].get(key, ga_meta["jobs_by_inch_and_type"].get(key, []))
                child_genes = order_crossover(list(p1_genes), list(p2_genes)) if len(p1_genes) >= 2 else list(p1_genes)
                child["gene_orders"][key] = child_genes

    for idx in two_d_gene_positions:
        child["welds"][idx] = parent1["welds"][idx] if random.random() < 0.5 else parent2["welds"][idx]
    return child


# 초보자 설명: 염색체에 돌연변이를 적용해 탐색 다양성을 높인다.
def mutate_chromosome(chromosome, two_d_gene_positions, mutation_rate, ga_meta=None, inch_group_optimization=False, base_jobs=None):
    child = clone_chromosome(chromosome)

    if not inch_group_optimization:
        n = len(child["order"])
        if n >= 2 and random.random() < mutation_rate:
            i, j = random.sample(range(n), 2)
            child["order"][i], child["order"][j] = child["order"][j], child["order"][i]
        if n >= 4 and random.random() < mutation_rate * 0.6:
            i, j = sorted(random.sample(range(n), 2))
            child["order"][i:j + 1] = reversed(child["order"][i:j + 1])
    else:
        inch_count = len(child.get("inch_order", []))
        if inch_count >= 2 and random.random() < mutation_rate:
            i, j = random.sample(range(inch_count), 2)
            child["inch_order"][i], child["inch_order"][j] = child["inch_order"][j], child["inch_order"][i]

        if inch_count >= 3 and random.random() < mutation_rate * 0.5:
            i, j = sorted(random.sample(range(inch_count), 2))
            child["inch_order"][i:j + 1] = reversed(child["inch_order"][i:j + 1])

        for inch in child.get("inch_order", []):
            type_order = child["type_orders"].get(inch, [])
            type_count = len(type_order)
            if type_count >= 2 and random.random() < mutation_rate:
                i, j = random.sample(range(type_count), 2)
                type_order[i], type_order[j] = type_order[j], type_order[i]
            if type_count >= 3 and random.random() < mutation_rate * 0.35:
                i, j = sorted(random.sample(range(type_count), 2))
                type_order[i:j + 1] = reversed(type_order[i:j + 1])
            child["type_orders"][inch] = type_order

            for prod in type_order:
                key = (inch, prod)
                genes = child["gene_orders"].get(key, [])
                gene_count = len(genes)
                if gene_count >= 2 and random.random() < mutation_rate * 0.8:
                    i, j = random.sample(range(gene_count), 2)
                    genes[i], genes[j] = genes[j], genes[i]
                if gene_count >= 4 and random.random() < mutation_rate * 0.25:
                    i, j = sorted(random.sample(range(gene_count), 2))
                    genes[i:j + 1] = reversed(genes[i:j + 1])
                child["gene_orders"][key] = genes

    for idx in two_d_gene_positions:
        if random.random() < mutation_rate:
            valid_welds = ["2D-1", "2D-2", "3D"]
            if base_jobs is not None:
                valid_welds = list(base_jobs[idx].get("valid_welds", [])) or valid_welds
            child["welds"][idx] = random.choice(valid_welds)
    return child


# 초보자 설명: 최종 plan, ProcessTime, 간트 이미지를 파일로 저장한다.
def save_final_outputs(paths, final_plan, final_result, plan_output_path=None, result_png_path=None,
                       process_time_path=None):
    plan_headers = ["번호", "Spool No.", "Inch", "Sch.", "스풀타입", "용접장", "Length1", "Length2", "피팅1", "피팅2"]
    plan_output_path = plan_output_path or plan_xlsx_path
    result_png_path = result_png_path or paths.result_png
    process_time_path = process_time_path or paths.process_time

    plan_full_rows = build_full_format_plan_rows(final_plan)
    save_rows_to_xlsx(plan_output_path, plan_headers, plan_full_rows)
    print(f"[INFO] Plan 저장 완료: {plan_output_path}")

    rerun_result = run_simulation(final_plan, output_png=result_png_path, render_chart=True)
    print(f"[INFO] 결과 간트 저장 완료: {result_png_path}")

    try:
        full_plan_rows = build_full_format_plan_rows(final_plan)
        save_process_time_multirow(process_time_path, full_plan_rows, rerun_result["gantt"])
        print(f"[INFO] ProcessTime 저장 완료: {process_time_path}")
    except Exception as exc:
        print(f"[WARN] ProcessTime 생성 실패: {exc}")

    return {
        "final_plan": final_plan,
        "final_result": rerun_result,
        "plan_full_rows": plan_full_rows,
    }


# 초보자 설명: 규칙 기반 모드의 전체 파이프라인을 실행한다.
def run_rulebase_pipeline(paths: Optional[AppPaths] = None) -> Dict[str, object]:
    """기존 Rule base 방식으로 생산 계획을 수립하고 시뮬레이션을 실행한다."""
    active_paths = paths or build_paths()
    initialize_inputs(active_paths)

    first_result, current_first_result = build_rulebase_first_result(active_paths)
    final_plan = build_rulebase_final_plan(first_result)

    saved = save_final_outputs(active_paths, final_plan, first_result, plan_output_path=plan_xlsx_path,
                               result_png_path=active_paths.result_png, process_time_path=active_paths.process_time)

    return {
        "mode": "rulebase",
        "paths": active_paths,
        "first_result": first_result,
        "current_first_result": current_first_result,
        "final_result": saved["final_result"],
        "final_plan": saved["final_plan"],
        "plan_full_rows": saved["plan_full_rows"],
    }


# 초보자 설명: GA 최적화 모드의 전체 파이프라인을 실행한다.
def run_ga_optimization(paths: Optional[AppPaths] = None, settings: Optional[UIOptimizationSettings] = None,
                        population_size=GA_DEFAULT_POPULATION, generations=GA_DEFAULT_GENERATIONS,
                        elite_count=GA_DEFAULT_ELITE, mutation_rate=GA_DEFAULT_MUTATION_RATE,
                        tournament_size=GA_DEFAULT_TOURNAMENT, seed=42) -> Dict[str, object]:
    """Genetic Algorithm으로 생산 계획을 최적화한다.

    - 기본: 전체 Job 순서 + 2D 용접장 배정 최적화
    - Inch별 최적화 ON: Inch 그룹 순서 + 동일 Inch 내 스풀 타입 순서 + 그룹 내부 Job 순서를 최적화
    """
    active_paths = paths or build_paths()
    initialize_inputs(active_paths)

    if not tmp_product_rows:
        raise RuntimeError("GA 최적화를 위한 tmp_product_rows가 비어 있습니다.")

    random.seed(seed)
    ga_settings = settings or UIOptimizationSettings()
    inch_group_optimization = bool(getattr(ga_settings, "inch_group_optimization", False))

    base_jobs, two_d_gene_positions, ga_meta = build_ga_base_jobs(tmp_product_rows)
    if not base_jobs:
        raise RuntimeError("GA 최적화를 수행할 Job이 없습니다.")

    population_size = max(4, int(population_size))
    job_count = len(base_jobs)
    generations = max(20, math.ceil(job_count / 3))
    elite_count = max(1, min(int(elite_count), population_size - 1))
    print(
        f"[GA] job 수 = {job_count} -> generation 수 = {generations} | Inch별 최적화 = {'ON' if inch_group_optimization else 'OFF'}",
        flush=True,
    )

    population = [
        create_random_chromosome(
            base_jobs,
            two_d_gene_positions,
            ga_meta=ga_meta,
            inch_group_optimization=inch_group_optimization,
        )
        for _ in range(population_size)
    ]
    best_score = None
    best_chromosome = None
    best_result = None
    best_plan = None
    history = []
    evaluation_cache = {}

    for gen in range(generations):
        scored_population = []
        for chromosome in population:
            signature = chromosome_signature(
                chromosome,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
            )
            cached_eval = evaluation_cache.get(signature)
            if cached_eval is None:
                cached_eval = evaluate_chromosome(
                    chromosome,
                    base_jobs,
                    ga_meta=ga_meta,
                    inch_group_optimization=inch_group_optimization,
                )
                evaluation_cache[signature] = cached_eval
            score, result, candidate_plan = cached_eval
            scored_population.append((score, chromosome, result, candidate_plan))
            if best_score is None or score < best_score:
                best_score = score
                best_chromosome = clone_chromosome(chromosome)
                best_result = result
                best_plan = candidate_plan

        scored_population.sort(key=lambda x: x[0])
        gen_best = scored_population[0][0]
        gen_avg = sum(item[0] for item in scored_population) / len(scored_population)
        history.append((gen + 1, gen_best, gen_avg))
        print(f"[GA] generation {gen + 1}/{generations} | best makespan = {gen_best:.1f} | avg = {gen_avg:.1f}",
              flush=True)

        if gen == generations - 1:
            break

        next_population = [clone_chromosome(item[1]) for item in scored_population[:elite_count]]
        while len(next_population) < population_size:
            parent1 = tournament_select(scored_population, tournament_size)
            parent2 = tournament_select(scored_population, tournament_size)
            child = crossover(
                parent1,
                parent2,
                two_d_gene_positions,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
            )
            child = mutate_chromosome(
                child,
                two_d_gene_positions,
                mutation_rate,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
                base_jobs=base_jobs,
            )
            next_population.append(child)
        population = next_population

    if best_plan is None:
        raise RuntimeError("GA 최적화 결과를 생성하지 못했습니다.")

    print(f"[GA] 최종 best makespan = {best_score:.1f}", flush=True)
    saved = save_final_outputs(active_paths, best_plan, best_result, plan_output_path=active_paths.ga_plan,
                               result_png_path=active_paths.ga_result_png,
                               process_time_path=active_paths.ga_process_time)

    return {
        "mode": "ga",
        "paths": active_paths,
        "history": history,
        "best_score": best_score,
        "first_result": best_result,
        "final_result": saved["final_result"],
        "final_plan": saved["final_plan"],
        "plan_full_rows": saved["plan_full_rows"],
        "best_chromosome": best_chromosome,
    }


# 초보자 설명: 실행 모드를 사용자에게 묻는다.
def ask_execution_mode():
    print("실행 모드를 선택하세요.")
    print("1. Rule base 모드")
    print("2. 최적화 모드 (Genetic Algorithm)")
    raw = input("번호를 입력하세요 [기본값: 1] : ").strip()
    if raw in ["2", "ga", "GA", "opt", "OPT", "optimize", "optimization"]:
        return "ga"
    return "rulebase"


# 초보자 설명: 입력값을 규칙 기반/GA/수동 모드 중 하나로 해석한다.
def resolve_execution_mode(argv=None):
    argv = argv if argv is not None else sys.argv[1:]
    normalized = [str(a).strip().lower() for a in argv]
    for arg in normalized:
        if arg in ["--mode=ga", "--mode=opt", "ga", "opt", "optimize", "optimization"]:
            return "ga"
        if arg in ["--mode=rule", "--mode=rulebase", "rule", "rulebase"]:
            return "rulebase"
    return ask_execution_mode()


# 초보자 설명: 입력 읽기부터 저장까지 전체 프로그램 흐름을 한 번에 실행한다.
def run_full_pipeline(paths: Optional[AppPaths] = None, mode: Optional[str] = None) -> Dict[str, object]:
    selected_mode = (mode or "").strip().lower() if mode else resolve_execution_mode()
    if selected_mode in ["ga", "opt", "optimize", "optimization"]:
        return run_ga_optimization(paths, settings=UIOptimizationSettings())
    return run_rulebase_pipeline(paths)


# =====================================================
# Tkinter UI 화면 구성
# -----------------------------------------------------
# 사용자가 버튼으로 입력/계획/시뮬레이션/출력을 실행할 수 있도록
# 창, 버튼, 상태창, 간트 차트 표시 영역을 만드는 구간이다.
# =====================================================

# 초보자 설명: print 출력 내용을 Tkinter 텍스트 박스로 보내기 위한 연결 클래스다.
class TextRedirector:
    # 초보자 설명: '__init__' 단계의 처리를 맡는 보조 함수다.
    def __init__(self, text_widget: tk.Text, root: tk.Misc):
        self.text_widget = text_widget
        self.root = root

    # 초보자 설명: 'write' 단계의 처리를 맡는 보조 함수다.
    def write(self, text: str):
        if not text:
            return
        try:
            self.text_widget.configure(state="normal")
            self.text_widget.insert("end", text)
            self.text_widget.see("end")
            self.text_widget.configure(state="disabled")
            self.root.update_idletasks()
            self.root.update()
        except Exception:
            pass

    # 초보자 설명: 'flush' 단계의 처리를 맡는 보조 함수다.
    def flush(self):
        try:
            self.root.update_idletasks()
            self.root.update()
        except Exception:
            pass


# 초보자 설명: GA 최적화 화면에서 사용하는 설정값을 담는 데이터 클래스다.
@dataclass
class UIOptimizationSettings:
    objective_makespan: bool = True
    objective_load_balance: bool = False
    scrap_max: int = 100
    inch_group_optimization: bool = False


# 초보자 설명: 사용자가 직접 고른 입력 파일 경로에서 기준 폴더를 만든다.
def build_paths_from_files(cycle_time_path: str, product_path: str, output_base_dir: Optional[str] = None) -> AppPaths:
    cycle_time_path = os.path.abspath(cycle_time_path)
    product_path = os.path.abspath(product_path)
    base = output_base_dir or os.path.dirname(cycle_time_path) or os.getcwd()
    return AppPaths(
        base_dir=base,
        cycle_time=cycle_time_path,
        product=product_path,
        plan=os.path.join(base, "Plan_rule.xlsx"),
        result_png=os.path.join(base, "result_rule.png"),
        process_time=os.path.join(base, "ProcessTime_rule.xlsx"),
        ga_result_png=os.path.join(base, "result_GA.png"),
        ga_plan=os.path.join(base, "Plan_GA.xlsx"),
        ga_process_time=os.path.join(base, "ProcessTime_GA.xlsx"),
    )


# 초보자 설명: 시뮬레이션 없이 규칙 기반 plan만 만든다.
def build_rulebase_plan_only(paths: AppPaths) -> Dict[str, object]:
    active_paths = paths
    initialize_inputs(active_paths)

    first_result, current_first_result = build_rulebase_first_result(active_paths)
    final_plan = build_rulebase_final_plan(first_result)

    return {
        "mode": "rulebase",
        "paths": active_paths,
        "first_result": first_result,
        "final_result": current_first_result,
        "final_plan": final_plan,
        "plan_full_rows": build_full_format_plan_rows(final_plan),
    }


# 초보자 설명: 시뮬레이션 없이 GA 기반 plan만 만든다.
def build_ga_plan_only(
        paths: AppPaths,
        settings: Optional[UIOptimizationSettings] = None,
        population_size=GA_DEFAULT_POPULATION,
        generations=GA_DEFAULT_GENERATIONS,
        elite_count=GA_DEFAULT_ELITE,
        mutation_rate=GA_DEFAULT_MUTATION_RATE,
        tournament_size=GA_DEFAULT_TOURNAMENT,
        seed=42,
) -> Dict[str, object]:
    active_paths = paths
    initialize_inputs(active_paths)

    if not tmp_product_rows:
        raise RuntimeError("GA 최적화를 위한 tmp_product_rows가 비어 있습니다.")

    random.seed(seed)
    ga_settings = settings or UIOptimizationSettings()
    inch_group_optimization = bool(getattr(ga_settings, "inch_group_optimization", False))

    base_jobs, two_d_gene_positions, ga_meta = build_ga_base_jobs(tmp_product_rows)
    if not base_jobs:
        raise RuntimeError("GA 최적화를 수행할 Job이 없습니다.")

    population_size = max(4, int(population_size))
    job_count = len(base_jobs)
    generations = max(20, math.ceil(job_count / 3))
    elite_count = max(1, min(int(elite_count), population_size - 1))

    print(
        f"[GA] job 수 = {job_count} -> generation 수 = {generations} | Inch별 최적화 = {'ON' if inch_group_optimization else 'OFF'}",
        flush=True,
    )

    population = [
        create_random_chromosome(
            base_jobs,
            two_d_gene_positions,
            ga_meta=ga_meta,
            inch_group_optimization=inch_group_optimization,
        )
        for _ in range(population_size)
    ]
    best_score = None
    best_chromosome = None
    best_result = None
    best_plan = None
    history = []
    evaluation_cache = {}

    for gen in range(generations):
        scored_population = []
        for chromosome in population:
            signature = chromosome_signature(
                chromosome,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
            )
            cached_eval = evaluation_cache.get(signature)
            if cached_eval is None:
                cached_eval = evaluate_chromosome(
                    chromosome,
                    base_jobs,
                    ga_meta=ga_meta,
                    inch_group_optimization=inch_group_optimization,
                )
                evaluation_cache[signature] = cached_eval
            score, result, candidate_plan = cached_eval
            scored_population.append((score, chromosome, result, candidate_plan))
            if best_score is None or score < best_score:
                best_score = score
                best_chromosome = clone_chromosome(chromosome)
                best_result = result
                best_plan = candidate_plan

        scored_population.sort(key=lambda x: x[0])
        gen_best = scored_population[0][0]
        gen_avg = sum(item[0] for item in scored_population) / len(scored_population)
        history.append((gen + 1, gen_best, gen_avg))
        print(f"[GA] generation {gen + 1}/{generations} | best makespan = {gen_best:.1f} | avg = {gen_avg:.1f}",
              flush=True)

        if gen == generations - 1:
            break

        next_population = [clone_chromosome(item[1]) for item in scored_population[:elite_count]]
        while len(next_population) < population_size:
            parent1 = tournament_select(scored_population, tournament_size)
            parent2 = tournament_select(scored_population, tournament_size)
            child = crossover(
                parent1,
                parent2,
                two_d_gene_positions,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
            )
            child = mutate_chromosome(
                child,
                two_d_gene_positions,
                mutation_rate,
                ga_meta=ga_meta,
                inch_group_optimization=inch_group_optimization,
                base_jobs=base_jobs,
            )
            next_population.append(child)
        population = next_population

    if best_plan is None:
        raise RuntimeError("GA 최적화 결과를 생성하지 못했습니다.")

    preview_result = run_simulation(best_plan, output_png=None, render_chart=False, combined_weld_hold=True,
                                    parallel_sub_products=True)

    return {
        "mode": "ga",
        "paths": active_paths,
        "history": history,
        "best_score": best_score,
        "first_result": best_result,
        "final_result": preview_result,
        "final_plan": best_plan,
        "plan_full_rows": build_full_format_plan_rows(best_plan),
        "best_chromosome": best_chromosome,
        "settings": settings,
    }


# 초보자 설명: plan 행 목록을 엑셀 파일로 저장한다.
def save_plan_rows_to_excel(plan_rows, output_path):
    headers = ["번호", "Spool No.", "Inch", "Sch.", "스풀타입", "용접장", "Length1", "Length2", "피팅1", "피팅2"]
    full_rows = build_full_format_plan_rows(plan_rows)
    save_rows_to_xlsx(output_path, headers, full_rows)
    return full_rows


# 초보자 설명: GA 관련 옵션을 입력받는 팝업 창 클래스다.
class GASettingsDialog(tk.Toplevel):
    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def __init__(self, master, settings: UIOptimizationSettings):
        super().__init__(master)
        self.title("최적화 설정")
        self.resizable(False, False)
        self.configure(bg="white")
        self.settings = settings
        self.result = None

        self.var_makespan = tk.BooleanVar(value=settings.objective_makespan)
        self.var_load_balance = tk.BooleanVar(value=settings.objective_load_balance)
        self.var_scrap = tk.StringVar(value=str(settings.scrap_max))
        self.var_inch_opt = tk.BooleanVar(value=settings.inch_group_optimization)
        self._load_balance_warned = False
        self._scrap_warned = False

        self._build_ui()
        self.transient(master)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _show_unimplemented_message(self):
        messagebox.showinfo("안내", "아직 기능이 구현되지 않았습니다.", parent=self)

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_click_load_balance(self):
        self.var_load_balance.set(False)
        self._show_unimplemented_message()

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_click_scrap_entry(self, event=None):
        self._show_unimplemented_message()
        return "break"

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _build_ui(self):
        outer = tk.Frame(self, bg="white", padx=18, pady=18)
        outer.pack(fill="both", expand=True)

        title = tk.Label(outer, text="최적화 설정", font=("Malgun Gothic", 15, "bold"), bg="white", fg="#1f2a44")
        title.pack(anchor="w", pady=(0, 12))

        boxes = tk.Frame(outer, bg="white")
        boxes.pack(fill="x")

        obj_frame = tk.LabelFrame(boxes, text="목적 함수", font=("Malgun Gothic", 11, "bold"), bg="white", fg="#1f2a44",
                                  padx=14, pady=12)
        obj_frame.grid(row=0, column=0, padx=(0, 12), sticky="nsew")
        ttk.Checkbutton(obj_frame, text="Make Span 최소화", variable=self.var_makespan).pack(anchor="w", pady=4)
        ttk.Checkbutton(obj_frame, text="부하분산율 최소화", variable=self.var_load_balance,
                        command=self._on_click_load_balance).pack(anchor="w", pady=4)

        cond_frame = tk.LabelFrame(boxes, text="제약 조건", font=("Malgun Gothic", 11, "bold"), bg="white", fg="#1f2a44",
                                   padx=14, pady=12)
        cond_frame.grid(row=0, column=1, sticky="nsew")

        row1 = tk.Frame(cond_frame, bg="white")
        row1.pack(anchor="w", pady=4)
        tk.Label(row1, text="잔재율 최대값", bg="white", font=("Malgun Gothic", 10)).pack(side="left")
        entry = ttk.Entry(row1, width=8, textvariable=self.var_scrap, justify="center")
        entry.pack(side="left", padx=8)
        entry.bind("<Button-1>", self._on_click_scrap_entry)
        entry.bind("<FocusIn>", self._on_click_scrap_entry)
        tk.Label(row1, text="%", bg="white", font=("Malgun Gothic", 10)).pack(side="left")
        ttk.Checkbutton(cond_frame, text="Inch별 최적화", variable=self.var_inch_opt).pack(anchor="w", pady=8)

        boxes.grid_columnconfigure(0, weight=1)
        boxes.grid_columnconfigure(1, weight=1)

        btn = tk.Button(outer, text="적용", command=self._apply, font=("Malgun Gothic", 11, "bold"), bg="#2f80ed",
                        fg="white", relief="flat", width=12, height=1, cursor="hand2")
        btn.pack(pady=(16, 0))

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _apply(self):
        scrap_text = self.var_scrap.get().strip()
        try:
            scrap_value = int(scrap_text)
        except Exception:
            messagebox.showerror("입력 오류", "잔재율 최대값은 0~100 사이의 정수여야 합니다.", parent=self)
            return
        if not (0 <= scrap_value <= 100):
            messagebox.showerror("입력 오류", "잔재율 최대값은 0~100 사이여야 합니다.", parent=self)
            return

        self.result = UIOptimizationSettings(
            objective_makespan=self.var_makespan.get(),
            objective_load_balance=self.var_load_balance.get(),
            scrap_max=scrap_value,
            inch_group_optimization=self.var_inch_opt.get(),
        )
        self.destroy()

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_close(self):
        self.result = None
        self.destroy()


# 초보자 설명: 생산계획/시뮬레이션 프로그램의 메인 Tkinter 화면을 구성하는 핵심 클래스다.
class SPFApp:
    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("SPF 생산 계획 & 시뮬레이션 프로그램 Ver1.5")
        self.root.configure(bg="white")
        self.root.geometry("1500x920")
        self.root.minsize(1280, 780)

        style = ttk.Style()
        try:
            style.theme_use("vista")
        except Exception:
            pass

        self.plant_var = tk.StringVar(value="SPF-대구")
        self.mode_var = tk.StringVar(value="rulebase")
        self.ga_settings = UIOptimizationSettings()

        self.cycle_time_path = None
        self.product_path = None
        self.active_paths = None
        self.input_loaded = False
        self.current_plan_rows = None
        self.current_result_bundle = None
        self.chart_photo = None
        self.last_chart_path = None
        self.last_chart_title = None
        self._last_render_width = None
        self.chart_rendered_size = (0, 0)
        self.chart_source_size = (0, 0)
        self.chart_job_bands = []
        self.chart_plot_box = None
        self.drag_enabled = False
        self.drag_start_index = None
        self.drag_current_index = None
        self.drag_preview_line = None
        self.drag_preview_rect = None
        self.drag_preview_text = None

        self.status_var = tk.StringVar(value="대기 중")
        self.file_var = tk.StringVar(value="CycleTime / Product 파일이 아직 선택되지 않았습니다.")
        self.chart_title_var = tk.StringVar(value="간트 차트가 아직 생성되지 않았습니다.")

        self._build_ui()
        self.append_output("[INFO] 프로그램이 시작되었습니다.", clear=True)
        self.append_output("[INFO] 기본 계획 수립 모드: 규칙 기반")

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _make_panel(self, master, title: str, padx: int = 12, pady: int = 10):
        return tk.LabelFrame(
            master,
            text=title,
            font=("Malgun Gothic", 11, "bold"),
            bg="white",
            fg="#1f2a44",
            padx=padx,
            pady=pady,
            bd=1,
            relief="solid",
        )

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _build_ui(self):
        container = tk.Frame(self.root, bg="white", padx=20, pady=18)
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(1, weight=1)
        container.grid_columnconfigure(0, weight=1)

        title = tk.Label(container, text="SPF 생산 계획 & 시뮬레이션 프로그램", bg="white", fg="#1f2a44",
                         font=("Malgun Gothic", 21, "bold"))
        title.grid(row=0, column=0, sticky="ew", pady=(0, 14))

        content = tk.Frame(container, bg="white")
        content.grid(row=1, column=0, sticky="nsew")
        content.grid_rowconfigure(0, weight=0)
        content.grid_rowconfigure(1, weight=0)
        content.grid_rowconfigure(2, weight=1)
        content.grid_columnconfigure(0, weight=0, minsize=390)
        content.grid_columnconfigure(1, weight=1)

        left = tk.Frame(content, bg="white")
        left.grid(row=0, column=0, rowspan=3, sticky="nsew", padx=(0, 14))
        left.grid_rowconfigure(2, weight=1)
        left.grid_columnconfigure(0, weight=1)

        top_left = tk.Frame(left, bg="white")
        top_left.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        top_left.grid_columnconfigure(0, weight=1)
        top_left.grid_columnconfigure(1, weight=1)

        plant_box = self._make_panel(top_left, "적용 공장", padx=18, pady=12)
        plant_box.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=(0, 10))
        ttk.Radiobutton(plant_box, text="SPF-대구", value="SPF-대구", variable=self.plant_var,
                        command=self._on_select_spf_daegu).pack(anchor="w", pady=4)
        ttk.Radiobutton(plant_box, text="SPF-M", value="SPF-M", variable=self.plant_var,
                        command=self._on_select_spf_m).pack(anchor="w", pady=4)

        mode_box = self._make_panel(top_left, "계획 수립 모드", padx=18, pady=12)
        mode_box.grid(row=0, column=1, sticky="nsew", pady=(0, 10))
        ttk.Radiobutton(mode_box, text="규칙 기반", value="rulebase", variable=self.mode_var,
                        command=self._on_mode_change).pack(anchor="w", pady=4)
        ttk.Radiobutton(mode_box, text="최적화(GA)", value="ga", variable=self.mode_var,
                        command=self._on_mode_change).pack(anchor="w", pady=4)
        ttk.Radiobutton(mode_box, text="수동모드", value="manual", variable=self.mode_var,
                        command=self._on_mode_change).pack(anchor="w", pady=4)

        button_grid = tk.Frame(top_left, bg="white")
        button_grid.grid(row=1, column=0, columnspan=2, sticky="ew")
        button_grid.grid_columnconfigure(0, weight=1)
        button_grid.grid_columnconfigure(1, weight=1)
        button_grid.grid_rowconfigure(0, weight=1)
        button_grid.grid_rowconfigure(1, weight=1)

        buttons = [
            ("생산정보\n입력", self.load_input_files),
            ("생산계획\n수립", self.build_plan),
            ("생산\n시뮬레이션", self.run_simulation_for_current_plan),
            ("생산계획\n출력", self.export_plan),
        ]
        for idx, (btn_text, cmd) in enumerate(buttons):
            r, c = divmod(idx, 2)
            btn = tk.Button(
                button_grid,
                text=btn_text,
                command=cmd,
                font=("Malgun Gothic", 15, "bold"),
                bg="#eaf2ff",
                fg="#1f2a44",
                activebackground="#d7e8ff",
                relief="groove",
                bd=2,
                height=3,
                cursor="hand2",
            )
            btn.grid(row=r, column=c, sticky="nsew", padx=(0 if c == 0 else 8, 0), pady=(0 if r == 0 else 8, 0))

        info_box = self._make_panel(left, "상태 정보")
        info_box.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        tk.Label(info_box, textvariable=self.file_var, bg="white", anchor="w", justify="left",
                 font=("Malgun Gothic", 10), wraplength=340).pack(fill="x", pady=(0, 8))
        tk.Label(info_box, textvariable=self.status_var, bg="white", anchor="w", justify="left",
                 font=("Malgun Gothic", 10, "bold"), fg="#2f80ed", wraplength=340).pack(fill="x")

        output_box = self._make_panel(left, "프로그램 실행 결과")
        output_box.grid(row=2, column=0, sticky="nsew")
        output_box.grid_rowconfigure(0, weight=1)
        output_box.grid_columnconfigure(0, weight=1)

        self.output_text = ScrolledText(
            output_box,
            height=24,
            wrap="word",
            font=("Consolas", 10),
            bg="#fbfcff",
            fg="#1f2a44",
            relief="solid",
            bd=1,
            state="disabled",
        )
        self.output_text.grid(row=0, column=0, sticky="nsew")

        right = tk.Frame(content, bg="white")
        right.grid(row=0, column=1, rowspan=3, sticky="nsew")
        right.grid_rowconfigure(0, weight=1)
        right.grid_columnconfigure(0, weight=1)

        chart_box = self._make_panel(right, "간트 차트")
        chart_box.grid(row=0, column=0, sticky="nsew")
        chart_box.grid_rowconfigure(1, weight=1)
        chart_box.grid_columnconfigure(0, weight=1)

        tk.Label(
            chart_box,
            textvariable=self.chart_title_var,
            bg="white",
            anchor="w",
            justify="left",
            font=("Malgun Gothic", 10, "bold"),
            fg="#4a5568",
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

        chart_frame = tk.Frame(chart_box, bg="white")
        chart_frame.grid(row=1, column=0, sticky="nsew")
        chart_frame.grid_rowconfigure(0, weight=1)
        chart_frame.grid_columnconfigure(0, weight=1)

        self.chart_canvas = tk.Canvas(
            chart_frame,
            bg="#fbfcff",
            highlightthickness=1,
            highlightbackground="#808080",
            bd=0,
        )
        self.chart_scrollbar = ttk.Scrollbar(chart_frame, orient="vertical", command=self.chart_canvas.yview)
        self.chart_canvas.configure(yscrollcommand=self.chart_scrollbar.set)

        self.chart_canvas.grid(row=0, column=0, sticky="nsew")
        self.chart_scrollbar.grid(row=0, column=1, sticky="ns")

        self.chart_image_id = None
        self.chart_empty_text_id = self.chart_canvas.create_text(
            20,
            20,
            anchor="nw",
            text="간트 차트가 아직 생성되지 않았습니다.",
            fill="#4a5568",
            font=("Malgun Gothic", 11),
        )

        self.chart_drag_hint = tk.Label(
            chart_box,
            text="간트 차트 생성 후 Job 행을 마우스로 드래그하면 순서가 바뀌고 즉시 재시뮬레이션됩니다.",
            bg="white",
            anchor="w",
            justify="left",
            font=("Malgun Gothic", 9),
            fg="#6b7280",
        )
        self.chart_drag_hint.grid(row=2, column=0, sticky="ew", pady=(8, 0))

        self.chart_canvas.bind("<Configure>", self._on_chart_canvas_configure)
        self.chart_canvas.bind("<MouseWheel>", self._on_chart_mousewheel)
        self.chart_canvas.bind("<ButtonPress-1>", self._on_chart_drag_start)
        self.chart_canvas.bind("<B1-Motion>", self._on_chart_drag_motion)
        self.chart_canvas.bind("<ButtonRelease-1>", self._on_chart_drag_release)

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _clear_drag_preview(self):
        for attr in ("drag_preview_rect", "drag_preview_line", "drag_preview_text"):
            item = getattr(self, attr, None)
            if item is not None:
                try:
                    self.chart_canvas.delete(item)
                except Exception:
                    pass
                setattr(self, attr, None)

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _get_job_band_bbox(self, idx):
        if idx is None or not self.current_plan_rows:
            return None
        img_w, img_h = self.chart_rendered_size
        if img_h <= 0:
            return None

        if self.chart_job_bands and self.chart_source_size[0] > 0 and self.chart_source_size[1] > 0:
            if 0 <= idx < len(self.chart_job_bands):
                band = self.chart_job_bands[idx]
                sx = img_w / max(1.0, float(self.chart_source_size[0]))
                sy = img_h / max(1.0, float(self.chart_source_size[1]))
                plot_box = self.chart_plot_box
                if plot_box:
                    left = float(plot_box.get("left", 10.0)) * sx
                    right = float(plot_box.get("right", self.chart_source_size[0] - 10.0)) * sx
                else:
                    left = 10.0
                    right = max(40.0, img_w - 10.0)
                top = float(band.get("top", 0.0)) * sy
                bottom = float(band.get("bottom", img_h)) * sy
                return left, top, right, bottom

        n = len(self.current_plan_rows)
        top_pad = img_h * 0.06
        bottom_pad = img_h * 0.04
        usable_h = max(1.0, img_h - top_pad - bottom_pad)
        row_h = usable_h / max(1, n)
        top = top_pad + (idx * row_h)
        bottom = top + row_h
        left = 10
        right = max(40, img_w - 10)
        return left, top, right, bottom

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _draw_drag_preview(self, canvas_y):
        self._clear_drag_preview()
        try:
            idx = self._get_job_index_from_chart_y(canvas_y)
            if idx is None:
                return
            bbox = self._get_job_band_bbox(idx)
            if bbox is None:
                return
            left, top, right, bottom = bbox
            x1 = left
            x2 = right
            y1 = top
            y2 = bottom
            mid_y = (y1 + y2) / 2.0
            self.drag_preview_rect = self.chart_canvas.create_rectangle(
                x1,
                y1,
                x2,
                y2,
                fill="#93c5fd",
                outline="#2563eb",
                width=2,
                stipple="gray25",
                tags=("drag_preview",),
            )
            self.drag_preview_line = self.chart_canvas.create_line(
                x1,
                mid_y,
                x2,
                mid_y,
                fill="#ef4444",
                width=2,
                dash=(6, 4),
                tags=("drag_preview",),
            )
            self.drag_preview_text = self.chart_canvas.create_text(
                x1 + 10,
                y1 + 10,
                anchor="nw",
                text=f"이 위치로 이동: {idx + 1}번 행",
                fill="#1d4ed8",
                font=("Malgun Gothic", 10, "bold"),
                tags=("drag_preview",),
            )
            self.chart_canvas.tag_raise("drag_preview")
        except Exception:
            self._clear_drag_preview()

    # 초보자 설명: 간트 차트 화면 표시나 스크롤/드래그 동작을 담당하는 UI 메서드다.
    def _get_job_index_from_chart_y(self, canvas_y):
        if not self.current_plan_rows:
            return None
        img_w, img_h = self.chart_rendered_size
        if img_h <= 0:
            return None
        try:
            y = float(canvas_y)
        except Exception:
            return None
        if y < 0:
            y = 0.0
        if y > img_h - 1:
            y = img_h - 1

        if self.chart_job_bands and self.chart_source_size[0] > 0 and self.chart_source_size[1] > 0:
            sy = img_h / max(1.0, float(self.chart_source_size[1]))
            scaled_bands = [
                (idx, float(b.get("top", 0.0)) * sy, float(b.get("bottom", 0.0)) * sy)
                for idx, b in enumerate(self.chart_job_bands)
            ]
            for idx, top, bottom in scaled_bands:
                if top <= y <= bottom:
                    return idx
            centers = [(idx, (top + bottom) / 2.0) for idx, top, bottom in scaled_bands]
            if centers:
                return min(centers, key=lambda t: abs(t[1] - y))[0]

        n = len(self.current_plan_rows)
        top_pad = img_h * 0.06
        bottom_pad = img_h * 0.04
        usable_h = max(1.0, img_h - top_pad - bottom_pad)
        relative = (y - top_pad) / usable_h
        relative = min(0.999999, max(0.0, relative))
        idx = int(relative * n)
        return min(max(idx, 0), n - 1)

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _renumber_current_plan_rows(self):
        if not self.current_plan_rows:
            return
        self.current_plan_rows = renumber_plan_rows(self.current_plan_rows)
        if self.current_result_bundle is not None:
            self.current_result_bundle["final_plan"] = list(self.current_plan_rows)

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_chart_drag_start(self, event):
        if not self.drag_enabled:
            return "break"
        idx = self._get_job_index_from_chart_y(self.chart_canvas.canvasy(event.y))
        if idx is None:
            return "break"
        self.drag_start_index = idx
        self.drag_current_index = idx
        self._draw_drag_preview(self.chart_canvas.canvasy(event.y))
        self.set_status(f"드래그 시작 | Job {idx + 1} 선택")
        return "break"

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_chart_drag_motion(self, event):
        if self.drag_start_index is None:
            return "break"
        idx = self._get_job_index_from_chart_y(self.chart_canvas.canvasy(event.y))
        if idx is not None:
            self.drag_current_index = idx
        self._draw_drag_preview(self.chart_canvas.canvasy(event.y))
        return "break"

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_chart_drag_release(self, event):
        if self.drag_start_index is None:
            return "break"
        try:
            target_idx = self._get_job_index_from_chart_y(self.chart_canvas.canvasy(event.y))
            if target_idx is None:
                target_idx = self.drag_current_index
            source_idx = self.drag_start_index
        finally:
            self.drag_start_index = None
            self.drag_current_index = None
            self._clear_drag_preview()

        if target_idx is None or source_idx is None:
            return "break"
        if target_idx == source_idx:
            self.set_status("드래그 순서 변경 없음")
            return "break"
        if not self.current_plan_rows or len(self.current_plan_rows) < 2:
            return "break"

        moved = list(self.current_plan_rows)
        row = moved.pop(source_idx)
        moved.insert(target_idx, row)
        self.current_plan_rows = renumber_plan_rows(moved)
        if self.current_result_bundle is None:
            self.current_result_bundle = {}
        self.current_result_bundle["final_plan"] = list(self.current_plan_rows)
        self.set_status(f"Job 순서 변경 완료 | {source_idx + 1} → {target_idx + 1} | 재시뮬레이션 중...")
        self.root.after(50, self._rerun_after_drag)
        return "break"

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _rerun_after_drag(self):
        try:
            temp_png = os.path.join(tempfile.gettempdir(), "spf_ui_gantt_drag.png")
            result = self._run_with_console(
                lambda: run_simulation(self.current_plan_rows, output_png=temp_png, render_chart=True),
                clear_output=False)
            self.current_result_bundle["final_result"] = result
            makespan = compute_makespan(result["gantt"])
            self.show_chart_image(temp_png, f"간트 차트 | Make Span: {makespan:.1f}분 | 드래그 재정렬 반영")
            self.set_status(f"드래그 재정렬 후 재시뮬레이션 완료 | Make Span: {makespan:.1f}분")
        except Exception as exc:
            self.set_status("드래그 재정렬 후 재시뮬레이션에 실패했습니다.")
            self.append_output("[ERROR] 드래그 재정렬 후 재시뮬레이션 실패")
            self.append_output(str(exc))
            messagebox.showerror("오류", format_exception_with_debug_log(RuntimeError(f"드래그 재정렬 후 재시뮬레이션 중 오류가 발생했습니다.\n\n{exc}"), preferred_dir=(self.active_paths.base_dir if self.active_paths else None)))

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _show_unimplemented_message(self):
        messagebox.showinfo("안내", "아직 기능이 구현되지 않았습니다.")

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_select_spf_daegu(self):
        self.plant_var.set("SPF-대구")

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_select_spf_m(self):
        self.plant_var.set("SPF-대구")
        self._show_unimplemented_message()

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def append_output(self, text: str, clear: bool = False):
        self.output_text.configure(state="normal")
        if clear:
            self.output_text.delete("1.0", "end")
        if text:
            self.output_text.insert("end", text)
            if not text.endswith("\n"):
                self.output_text.insert("end", "\n")
        self.output_text.see("end")
        self.output_text.configure(state="disabled")
        self.root.update_idletasks()

    # 초보자 설명: 간트 차트 화면 표시나 스크롤/드래그 동작을 담당하는 UI 메서드다.
    def clear_chart(self, message: str = "간트 차트가 아직 생성되지 않았습니다."):
        self.chart_title_var.set(message)
        self.chart_photo = None
        self.last_chart_path = None
        self.last_chart_title = message
        self.chart_rendered_size = (0, 0)
        self.chart_source_size = (0, 0)
        self.chart_job_bands = []
        self.chart_plot_box = None
        self._clear_drag_preview()
        if hasattr(self, "chart_canvas"):
            self.chart_canvas.delete("all")
            self.chart_image_id = None
            self.chart_empty_text_id = self.chart_canvas.create_text(
                20,
                20,
                anchor="nw",
                text=message,
                fill="#4a5568",
                font=("Malgun Gothic", 11),
            )
            self.chart_canvas.yview_moveto(0)
            self.chart_canvas.configure(scrollregion=(0, 0, 0, 0))
        self.root.update_idletasks()

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_chart_canvas_configure(self, event=None):
        if not hasattr(self, "chart_canvas"):
            return
        if getattr(self, "last_chart_path", None) and Image is not None and ImageTk is not None:
            new_w = max(200, self.chart_canvas.winfo_width() - 4)
            if getattr(self, "_last_render_width", None) != new_w:
                self._last_render_width = new_w
                self.root.after_idle(self._refresh_chart_image)

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_chart_mousewheel(self, event):
        if hasattr(self, "chart_canvas"):
            delta = 0
            if getattr(event, "delta", 0):
                delta = int(-1 * (event.delta / 120))
            elif getattr(event, "num", None) == 4:
                delta = -1
            elif getattr(event, "num", None) == 5:
                delta = 1
            if delta != 0:
                self.chart_canvas.yview_scroll(delta, "units")
        return "break"

    # 초보자 설명: 간트 차트 화면 표시나 스크롤/드래그 동작을 담당하는 UI 메서드다.
    def _refresh_chart_image(self):
        if getattr(self, "last_chart_path", None):
            self.show_chart_image(self.last_chart_path, getattr(self, "last_chart_title", "간트 차트"))

    # 초보자 설명: 간트 차트 화면 표시나 스크롤/드래그 동작을 담당하는 UI 메서드다.
    def show_chart_image(self, image_path: str, title: str):
        if not os.path.exists(image_path):
            self.append_output(f"[WARN] 간트 차트 파일을 찾을 수 없습니다: {image_path}")
            self.clear_chart("간트 차트 파일을 찾을 수 없습니다.")
            return

        self.root.update_idletasks()
        canvas_w = max(200, self.chart_canvas.winfo_width() - 4) if hasattr(self, "chart_canvas") else 800

        try:
            if Image is not None and ImageTk is not None:
                img = Image.open(image_path)
                src_w, src_h = img.size
                ratio = canvas_w / max(1, src_w)
                new_w = max(1, int(src_w * ratio))
                new_h = max(1, int(src_h * ratio))
                img = img.resize((new_w, new_h), Image.LANCZOS)
                self.chart_photo = ImageTk.PhotoImage(img)
                self.chart_rendered_size = (new_w, new_h)
            else:
                self.chart_photo = tk.PhotoImage(file=image_path)
                self.chart_rendered_size = (self.chart_photo.width(), self.chart_photo.height())

            self.chart_canvas.delete("all")
            self.chart_image_id = self.chart_canvas.create_image(0, 0, anchor="nw", image=self.chart_photo)
            self.chart_canvas.configure(scrollregion=(0, 0, self.chart_rendered_size[0], self.chart_rendered_size[1]))

            self.chart_source_size = (0, 0)
            self.chart_job_bands = []
            self.chart_plot_box = None
            try:
                result_bundle = getattr(self, "current_result_bundle", None)
                final_result = None
                if isinstance(result_bundle, dict):
                    if "chart_meta" in result_bundle:
                        final_result = result_bundle
                    elif isinstance(result_bundle.get("final_result"), dict):
                        final_result = result_bundle.get("final_result")
                chart_meta = final_result.get("chart_meta") if isinstance(final_result, dict) else None
                if isinstance(chart_meta, dict):
                    img_size = chart_meta.get("image_size")
                    if isinstance(img_size, (list, tuple)) and len(img_size) == 2:
                        self.chart_source_size = (safe_ct(img_size[0], 0.0), safe_ct(img_size[1], 0.0))
                    bands = chart_meta.get("job_bands") or []
                    if isinstance(bands, list):
                        self.chart_job_bands = list(bands)
                    plot_box = chart_meta.get("plot_box")
                    if isinstance(plot_box, dict):
                        self.chart_plot_box = dict(plot_box)
            except Exception:
                self.chart_source_size = (0, 0)
                self.chart_job_bands = []
                self.chart_plot_box = None

            self.chart_title_var.set(title)
            self.last_chart_path = image_path
            self.last_chart_title = title
            self._last_render_width = canvas_w
            self.drag_enabled = bool(self.current_plan_rows) and len(self.current_plan_rows) >= 2
            self.chart_canvas.yview_moveto(0)
            self.root.update_idletasks()
        except Exception as exc:
            self.chart_title_var.set(title)
            self.chart_canvas.delete("all")
            self.chart_empty_text_id = self.chart_canvas.create_text(
                20,
                20,
                anchor="nw",
                text="이미지 표시 실패",
                fill="#b91c1c",
                font=("Malgun Gothic", 11, "bold"),
            )
            self.append_output(f"[WARN] 간트 차트 이미지를 UI에 표시하지 못했습니다.\n{exc}")

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _run_with_console(self, func, *, clear_output: bool = False):
        if clear_output:
            self.append_output("", clear=True)
        redirector = TextRedirector(self.output_text, self.root)
        debug_log("UI _run_with_console 시작", preferred_dir=(self.active_paths.base_dir if self.active_paths else None))
        with redirect_stdout(redirector), redirect_stderr(redirector):
            try:
                result = func()
                debug_log("UI _run_with_console 완료", preferred_dir=(self.active_paths.base_dir if self.active_paths else None))
                return result
            except Exception as exc:
                debug_log_exception("UI _run_with_console 예외", exc, preferred_dir=(self.active_paths.base_dir if self.active_paths else None))
                raise

    # 초보자 설명: 화면에서 특정 사용자 동작이 발생했을 때 실행되는 이벤트 처리 메서드다.
    def _on_mode_change(self):
        mode = self.mode_var.get()
        if mode == "ga":
            self.open_ga_settings_dialog()
        elif mode == "manual":
            self.set_status("수동모드가 선택되었습니다. 생산정보 입력 없이 생산계획 수립 버튼으로 Plan_man.xlsx를 바로 읽을 수 있습니다.")
        else:
            self.set_status("규칙 기반 모드가 선택되었습니다.")

    # 초보자 설명: 화면의 상태 표시 문구를 바꾸는 UI 메서드다.
    def set_status(self, text: str):
        self.status_var.set(text)
        self.append_output(f"[STATUS] {text}")
        self.root.update_idletasks()

    # 초보자 설명: 화면에서 다른 설정 창이나 파일 선택 창을 여는 UI 메서드다.
    def open_ga_settings_dialog(self):
        dialog = GASettingsDialog(self.root, self.ga_settings)
        self.root.wait_window(dialog)
        if dialog.result is not None:
            self.ga_settings = dialog.result
            desc = []
            if self.ga_settings.objective_makespan:
                desc.append("Make Span")
            if self.ga_settings.objective_load_balance:
                desc.append("부하분산율")
            objective_text = ", ".join(desc) if desc else "선택 없음"
            self.set_status(
                f"GA 설정 적용 완료 | 목적함수: {objective_text} | 잔재율 최대값: {self.ga_settings.scrap_max}% | Inch별 최적화: {'ON' if self.ga_settings.inch_group_optimization else 'OFF'}"
            )

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _validate_ga_selection(self):
        if not self.ga_settings.objective_makespan and not self.ga_settings.objective_load_balance:
            messagebox.showwarning("설정 확인", "목적함수를 하나 이상 선택해 주세요.")
            return False
        return True

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _require_input_files(self):
        if self.mode_var.get() == "manual":
            return True
        if self.cycle_time_path and self.product_path:
            return True
        messagebox.showwarning("입력 필요", "먼저 생산정보 입력 버튼으로 CycleTime.xlsx와 Product.xlsx를 선택해 주세요.")
        return False

    # 초보자 설명: UI에서 선택한 파일을 읽어 프로그램 상태에 반영하는 메서드다.
    def load_input_files(self):
        cycle_path = filedialog.askopenfilename(
            title="CycleTime.xlsx 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not cycle_path:
            return

        product_path = filedialog.askopenfilename(
            title="Product.xlsx 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not product_path:
            return

        self.set_status("입력 파일을 읽는 중입니다...")
        self.clear_chart("입력 파일을 읽는 중입니다.")
        try:
            self.cycle_time_path = cycle_path
            self.product_path = product_path
            out_base = os.path.dirname(product_path) or os.path.dirname(cycle_path) or os.getcwd()
            self.active_paths = build_paths_from_files(cycle_path, product_path, output_base_dir=out_base)
            self._run_with_console(lambda: initialize_inputs(self.active_paths), clear_output=False)
            self.input_loaded = True
            self.current_plan_rows = None
            self.current_result_bundle = None
            self.file_var.set(f"CycleTime: {self.cycle_time_path}\nProduct: {self.product_path}")
            self.set_status("생산정보 입력이 완료되었습니다.")
            messagebox.showinfo("완료", "CycleTime.xlsx와 Product.xlsx를 정상적으로 읽었습니다.")
        except Exception as exc:
            self.input_loaded = False
            self.current_plan_rows = None
            self.current_result_bundle = None
            self.set_status("입력 파일 읽기에 실패했습니다.")
            messagebox.showerror("오류", format_exception_with_debug_log(RuntimeError(f"입력 파일을 읽는 중 오류가 발생했습니다.\n\n{exc}"), preferred_dir=(os.path.dirname(self.cycle_time_path) if self.cycle_time_path else None)))

    # 초보자 설명: 화면에 필요한 위젯이나 레이아웃을 구성하는 UI 메서드다.
    def build_plan(self):
        if not self._require_input_files():
            return
        if self.mode_var.get() == "ga" and not self._validate_ga_selection():
            return

        self.set_status("생산계획을 메모리에 생성하는 중입니다...")
        self.clear_chart("생산계획이 생성되면 이 영역에 간트 차트가 표시됩니다.")
        try:
            if self.mode_var.get() == "manual":
                manual_plan_path = filedialog.askopenfilename(
                    title="Plan_man.xlsx 선택",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                )
                if not manual_plan_path:
                    self.set_status("수동 계획 파일 선택이 취소되었습니다.")
                    return
                result = self._run_with_console(lambda: load_manual_plan_from_excel(manual_plan_path),
                                                clear_output=False)
                self.active_paths = result["paths"]
                self.file_var.set(
                    f"Manual Plan: {result['manual_plan_path']}\nCycleTime: {self.active_paths.cycle_time}")
                self.input_loaded = True
            elif self.mode_var.get() == "ga":
                result = self._run_with_console(
                    lambda: build_ga_plan_only(self.active_paths, settings=self.ga_settings), clear_output=False)
                if self.ga_settings.objective_load_balance and not self.ga_settings.objective_makespan:
                    messagebox.showinfo("참고",
                                        "부하분산율 목적함수는 현재 변수만 연결되어 있으며, 실제 최적화 평가는 현재 코드 기준으로 Make Span 최소화로 수행됩니다.")
            else:
                result = self._run_with_console(lambda: build_rulebase_plan_only(self.active_paths), clear_output=False)

            self.current_result_bundle = result
            self.current_plan_rows = list(result["final_plan"])
            makespan = compute_makespan(result["final_result"]["gantt"]) if result.get("final_result") else 0.0
            if self.mode_var.get() == "manual":
                self.set_status(f"수동 계획 불러오기 완료 | Job 수: {len(self.current_plan_rows)}")
                messagebox.showinfo("완료", "Plan_man.xlsx를 메모리에 불러왔습니다.\n이제 생산 시뮬레이션 버튼으로 바로 실행할 수 있습니다.")
            else:
                self.set_status(f"생산계획 수립 완료 | Mode: {result['mode']} | Make Span: {makespan:.1f}분")
                messagebox.showinfo("완료", "생산계획이 메모리에 생성되었습니다.\n파일로 저장되지는 않았습니다.")
        except Exception as exc:
            self.current_plan_rows = None
            self.current_result_bundle = None
            self.set_status("생산계획 수립에 실패했습니다.")
            messagebox.showerror("오류", format_exception_with_debug_log(RuntimeError(f"생산계획 수립 중 오류가 발생했습니다.\n\n{exc}"), preferred_dir=(self.active_paths.base_dir if self.active_paths else None)))

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def run_simulation_for_current_plan(self):
        if not self.current_plan_rows:
            messagebox.showwarning("계획 필요", "먼저 생산계획 수립을 실행해 주세요.")
            return

        self.set_status("생산 시뮬레이션을 수행하는 중입니다...")
        self.clear_chart("생산 시뮬레이션 수행 중입니다...")
        try:
            self._renumber_current_plan_rows()
            temp_png = os.path.join(tempfile.gettempdir(), "spf_ui_gantt.png")
            result = self._run_with_console(
                lambda: run_simulation(self.current_plan_rows, output_png=temp_png, render_chart=True),
                clear_output=False)
            if self.current_result_bundle is None:
                self.current_result_bundle = {}
            self.current_result_bundle["final_result"] = result
            self.current_result_bundle["final_plan"] = list(self.current_plan_rows)
            makespan = compute_makespan(result["gantt"])
            self.show_chart_image(temp_png, f"간트 차트 | Make Span: {makespan:.1f}분")
            self.set_status(f"생산 시뮬레이션 완료 | Make Span: {makespan:.1f}분")
        except Exception as exc:
            self.set_status("생산 시뮬레이션에 실패했습니다.")
            self.clear_chart("생산 시뮬레이션에 실패했습니다.")
            messagebox.showerror("오류", format_exception_with_debug_log(RuntimeError(f"생산 시뮬레이션 중 오류가 발생했습니다.\n\n{exc}"), preferred_dir=(self.active_paths.base_dir if self.active_paths else None)))

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def _derive_export_paths(self, plan_path):
        plan_path = os.path.abspath(plan_path)
        base_dir = os.path.dirname(plan_path)
        stem = os.path.splitext(os.path.basename(plan_path))[0]
        lower_stem = stem.lower()

        if lower_stem.startswith("plan"):
            suffix = stem[4:]
            process_stem = f"ProcessTime{suffix}"
            chart_stem = f"result{suffix}"
        else:
            process_stem = f"{stem}_ProcessTime"
            chart_stem = f"{stem}_gantt"

        return {
            "plan": plan_path,
            "process_time": os.path.join(base_dir, f"{process_stem}.xlsx"),
            "chart": os.path.join(base_dir, f"{chart_stem}.png"),
        }

    # 초보자 설명: 화면 동작을 처리하기 위한 UI 메서드다.
    def export_plan(self):
        if not self.current_plan_rows:
            messagebox.showwarning("계획 필요", "먼저 생산계획 수립을 실행해 주세요.")
            return

        initial_dir = self.active_paths.base_dir if self.active_paths else os.getcwd()
        if self.mode_var.get() == "ga":
            default_name = "Plan_GA.xlsx"
        elif self.mode_var.get() == "manual":
            default_name = "Plan_manual.xlsx"
        else:
            default_name = "Plan_rule.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="생산계획 저장",
            defaultextension=".xlsx",
            initialdir=initial_dir,
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not save_path:
            return

        export_paths = self._derive_export_paths(save_path)
        self.set_status("생산계획 / ProcessTime / 간트 차트 파일을 저장하는 중입니다...")
        try:
            self._renumber_current_plan_rows()
            self._run_with_console(lambda: save_plan_rows_to_excel(self.current_plan_rows, export_paths["plan"]),
                                   clear_output=False)

            # 현재 화면에 반영된 최종 Job 순서 기준으로 다시 시뮬레이션하여 저장한다.
            result = self._run_with_console(
                lambda: run_simulation(self.current_plan_rows, output_png=export_paths["chart"], render_chart=True),
                clear_output=False)
            save_process_time_multirow(export_paths["process_time"],
                                       build_full_format_plan_rows(self.current_plan_rows), result["gantt"])

            if self.current_result_bundle is None:
                self.current_result_bundle = {}
            self.current_result_bundle["final_result"] = result

            if os.path.exists(export_paths["chart"]):
                makespan = compute_makespan(result["gantt"])
                self.show_chart_image(export_paths["chart"], f"간트 차트 | Make Span: {makespan:.1f}분")

            self.set_status(
                f"파일 저장 완료 | Plan: {os.path.basename(export_paths['plan'])} | "
                f"ProcessTime: {os.path.basename(export_paths['process_time'])} | "
                f"Chart: {os.path.basename(export_paths['chart'])}"
            )
            messagebox.showinfo(
                "완료",
                "파일 저장이 완료되었습니다.\n\n"
                f"Plan: {export_paths['plan']}\n"
                f"ProcessTime: {export_paths['process_time']}\n"
                f"간트 차트: {export_paths['chart']}"
            )
        except Exception as exc:
            self.set_status("파일 저장에 실패했습니다.")
            messagebox.showerror("오류", format_exception_with_debug_log(RuntimeError(f"파일 저장 중 오류가 발생했습니다.\n\n{exc}"), preferred_dir=(self.active_paths.base_dir if self.active_paths else None)))


# 초보자 설명: 아직 구현되지 않은 GA 옵션을 눌렀을 때 안내 메시지를 띄운다.
def _ga_dialog_unimplemented(self, event=None):
    messagebox.showinfo("안내", "아직 기능이 구현되지 않았습니다.", parent=self)
    return "break"


# 초보자 설명: GA 설정 대화상자의 위젯을 배치한다.
def _ga_dialog_build_ui(self):
    outer = tk.Frame(self, bg="white", padx=18, pady=18)
    outer.pack(fill="both", expand=True)

    title = tk.Label(
        outer,
        text="최적화 설정",
        font=("Malgun Gothic", 15, "bold"),
        bg="white",
        fg="#1f2a44",
    )
    title.pack(anchor="w", pady=(0, 12))

    boxes = tk.Frame(outer, bg="white")
    boxes.pack(fill="x")

    obj_frame = tk.LabelFrame(
        boxes,
        text="목적 함수",
        font=("Malgun Gothic", 11, "bold"),
        bg="white",
        fg="#1f2a44",
        padx=14,
        pady=12,
    )
    obj_frame.grid(row=0, column=0, padx=(0, 12), sticky="nsew")

    ttk.Checkbutton(obj_frame, text="Make Span 최소화", variable=self.var_makespan).pack(anchor="w", pady=4)
    ttk.Checkbutton(
        obj_frame,
        text="부하분산율 최소화",
        variable=self.var_load_balance,
        command=lambda: (self.var_load_balance.set(False), _ga_dialog_unimplemented(self)),
    ).pack(anchor="w", pady=4)

    cond_frame = tk.LabelFrame(
        boxes,
        text="제약 조건",
        font=("Malgun Gothic", 11, "bold"),
        bg="white",
        fg="#1f2a44",
        padx=14,
        pady=12,
    )
    cond_frame.grid(row=0, column=1, sticky="nsew")

    row1 = tk.Frame(cond_frame, bg="white")
    row1.pack(anchor="w", pady=4)
    tk.Label(row1, text="잔재율 최대값", bg="white", font=("Malgun Gothic", 10)).pack(side="left")
    entry = ttk.Entry(row1, width=8, textvariable=self.var_scrap, justify="center")
    entry.pack(side="left", padx=8)
    entry.bind("<FocusIn>", _ga_dialog_unimplemented)
    entry.bind("<Button-1>", _ga_dialog_unimplemented)
    tk.Label(row1, text="%", bg="white", font=("Malgun Gothic", 10)).pack(side="left")

    ttk.Checkbutton(cond_frame, text="Inch별 최적화", variable=self.var_inch_opt).pack(anchor="w", pady=8)

    boxes.grid_columnconfigure(0, weight=1)
    boxes.grid_columnconfigure(1, weight=1)

    btn = tk.Button(
        outer,
        text="적용",
        command=self._apply,
        font=("Malgun Gothic", 11, "bold"),
        bg="#2f80ed",
        fg="white",
        relief="flat",
        width=12,
        height=1,
        cursor="hand2",
    )
    btn.pack(pady=(16, 0))


# 초보자 설명: GA 설정 창에서 입력한 값을 적용한다.
def _ga_dialog_apply(self):
    scrap_text = self.var_scrap.get().strip()
    try:
        scrap_value = int(scrap_text)
    except Exception:
        messagebox.showerror("입력 오류", "잔재율 최대값은 0~100 사이의 정수여야 합니다.", parent=self)
        return
    if not (0 <= scrap_value <= 100):
        messagebox.showerror("입력 오류", "잔재율 최대값은 0~100 사이여야 합니다.", parent=self)
        return

    self.result = UIOptimizationSettings(
        objective_makespan=self.var_makespan.get(),
        objective_load_balance=self.var_load_balance.get(),
        scrap_max=scrap_value,
        inch_group_optimization=self.var_inch_opt.get(),
    )
    self.destroy()


# 초보자 설명: GA 설정 창이 닫힐 때 정리 작업을 한다.
def _ga_dialog_on_close(self):
    self.result = None
    self.destroy()


GASettingsDialog._build_ui = _ga_dialog_build_ui
GASettingsDialog._apply = _ga_dialog_apply
GASettingsDialog._on_close = _ga_dialog_on_close
GASettingsDialog._unimplemented = _ga_dialog_unimplemented

SPFAPP = SPFApp


# =====================================================
# 프로그램 시작점
# =====================================================

# 초보자 설명: 프로그램 실행 진입점이다. 보통 이 함수에서 UI를 시작한다.
def main() -> None:
    set_debug_log_path()
    debug_log(f"프로그램 시작 | platform={platform.platform()} | python={sys.version}")
    root = tk.Tk()
    app = SPFApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
